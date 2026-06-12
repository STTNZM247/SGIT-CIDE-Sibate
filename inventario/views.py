import base64
import csv
import io
import os
import re
import secrets
import textwrap
import unicodedata
from collections import defaultdict
from datetime import date, timedelta

from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Exists, OuterRef
from django.http import Http404, HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from .models import AuditoriaLog, Catalogo, DetallePedido, Disponibilidad, ImportacionInventarioLog, Notificacion, Pedido, PedidoEvidencia, Producto, ProductoFoto, Rol, Subcategoria, Usuario, VerificacionSenaToken
from .forms import ProductoForm


DEVOLUCION_CODIGO_SEGUNDOS = 60
VALIDACION_MANUAL_VENCE_HORAS = 4
SUBCAT_MAX_WORD_LENGTH = 20


def _safe_code_fragment(raw_value, fallback_id):
    value = (raw_value or '').strip().upper()
    if value:
        return re.sub(r'[^A-Z0-9]+', '', value)
    return str(fallback_id).zfill(4)


def _generar_codigo_producto(subcategoria):
    categoria = subcategoria.subcategoria_padre
    catalogo = subcategoria.id_cat_fk

    macro_code = _safe_code_fragment(catalogo.codigo_macro, catalogo.id_cat)
    categoria_code = _safe_code_fragment(getattr(categoria, 'codigo_clasificacion', ''), categoria.id_subcat)
    subcategoria_code = _safe_code_fragment(subcategoria.codigo_clasificacion, subcategoria.id_subcat)
    prefix = f'{macro_code}-{categoria_code}-{subcategoria_code}'

    last_codigo = (
        Producto.objects
        .filter(codigo_producto__startswith=f'{prefix}-')
        .order_by('-codigo_producto')
        .values_list('codigo_producto', flat=True)
        .first()
    )

    consecutivo = 1
    if last_codigo:
        tail = (last_codigo or '').split('-')[-1]
        if tail.isdigit():
            consecutivo = int(tail) + 1

    return f'{prefix}-{str(consecutivo).zfill(4)}'


def _generar_codigo_producto_aleatorio():
    # Fallback para productos sin subcategoría válida: código corto, legible y único.
    for _ in range(20):
        candidate = f"PRD-{secrets.token_hex(3).upper()}"
        if not Producto.objects.filter(codigo_producto__iexact=candidate).exists():
            return candidate
    return None


def _asignar_codigo_producto_si_falta(producto):
    if (producto.codigo_producto or '').strip():
        return producto.codigo_producto

    subcategoria = producto.subcategorias.order_by('id_subcat').first()
    if subcategoria:
        for _ in range(5):
            try:
                codigo = _generar_codigo_producto(subcategoria)
                producto.codigo_producto = codigo
                producto.fch_ult_act = timezone.now()
                producto.save(update_fields=['codigo_producto', 'fch_ult_act'])
                return codigo
            except IntegrityError:
                continue

    codigo_aleatorio = _generar_codigo_producto_aleatorio()
    if not codigo_aleatorio:
        return None

    producto.codigo_producto = codigo_aleatorio
    producto.fch_ult_act = timezone.now()
    producto.save(update_fields=['codigo_producto', 'fch_ult_act'])
    return codigo_aleatorio


def _expirar_solicitudes_validacion_manual():
    """Caduca solicitudes manuales SENA en estado 'solicitada' con más de 4 horas."""
    limite = timezone.now() - timedelta(hours=VALIDACION_MANUAL_VENCE_HORAS)
    return (
        Usuario.objects
        .filter(verificacion_sena_estado='solicitada', verificacion_sena_solicitada_en__lte=limite)
        .update(
            verificacion_sena_estado='pendiente',
            verificacion_sena_solicitada_en=None,
        )
    )


def _reabrir_solicitudes_con_enlace_vencido():
    """Reabre solicitudes en enlace_enviado cuando no existe token vigente sin usar."""
    ahora = timezone.now()
    token_vigente_qs = VerificacionSenaToken.objects.filter(
        usuario=OuterRef('pk'),
        usado_en__isnull=True,
        expira_en__gte=ahora,
    )
    return (
        Usuario.objects
        .filter(verificacion_sena_estado='enlace_enviado')
        .annotate(tiene_token_vigente=Exists(token_vigente_qs))
        .filter(tiene_token_vigente=False)
        .update(verificacion_sena_estado='solicitada')
    )


def _crear_notificacion(usuario, tipo, titulo, mensaje, pedido_id=None):
    """Crea una notificación para el usuario de forma segura (nunca lanza excepción)."""
    try:
        Notificacion.objects.create(
            id_usuario_fk=usuario,
            tipo=tipo,
            titulo=titulo,
            mensaje=mensaje,
            id_pedido_ref=pedido_id,
        )
    except Exception:
        pass


def _notificar_staff(tipo, titulo, mensaje, pedido_id=None):
    """Envía una notificación a todos los usuarios con rol admin o almacenista."""
    try:
        from .models import Usuario as _Usuario
        staff = _Usuario.objects.filter(
            id_rol_fk__nombre_rol__in=['admin', 'almacenista'],
            is_active=True,
        )
        Notificacion.objects.bulk_create([
            Notificacion(
                id_usuario_fk=u,
                tipo=tipo,
                titulo=titulo,
                mensaje=mensaje,
                id_pedido_ref=pedido_id,
            )
            for u in staff
        ])  
    except Exception:
        pass


def _tiempo_vencido(fecha_devolucion, ahora):
    """Devuelve texto humanizado de cuánto tiempo lleva vencido. Ej: 'hace 2 horas', 'hace 3 días'."""
    diff = ahora - fecha_devolucion
    total_seg = int(diff.total_seconds())
    if total_seg < 60:
        return 'hace unos segundos'
    minutos = total_seg // 60
    if minutos < 60:
        return f'hace {minutos} min'
    horas = minutos // 60
    if horas < 24:
        return f'hace {horas} h {minutos % 60} min' if minutos % 60 else f'hace {horas} h'
    dias = horas // 24
    horas_rest = horas % 24
    if dias == 1:
        return f'hace 1 día' + (f' y {horas_rest} h' if horas_rest else '')
    return f'hace {dias} días' + (f' y {horas_rest} h' if horas_rest else '')


def _tiempo_restante(fecha_devolucion, ahora):
    """Devuelve texto humanizado del tiempo que queda. Ej: '2 h 30 min', '3 días'."""
    diff = fecha_devolucion - ahora
    total_seg = int(diff.total_seconds())
    if total_seg <= 0:
        return ''
    minutos = total_seg // 60
    if minutos < 60:
        return f'{minutos} min'
    horas = minutos // 60
    if horas < 24:
        return f'{horas} h {minutos % 60} min' if minutos % 60 else f'{horas} h'
    dias = horas // 24
    horas_rest = horas % 24
    if dias == 1:
        return '1 día' + (f' y {horas_rest} h' if horas_rest else '')
    return f'{dias} días' + (f' y {horas_rest} h' if horas_rest else '')


def _registrar_auditoria(request, accion, entidad, entidad_id=None, descripcion=''):
    usuario = None
    if request and getattr(request, 'user', None) and request.user.is_authenticated:
        usuario = request.user
    rol = None
    if usuario and getattr(usuario, 'id_rol_fk', None):
        rol = usuario.id_rol_fk.nombre_rol

    actor = 'sistema'
    if usuario and usuario.is_authenticated:
        nombre = f'{getattr(usuario, "nombre", "") or ""} {getattr(usuario, "apellido", "") or ""}'.strip()
        actor = nombre or getattr(usuario, 'correo', None) or f'usuario#{getattr(usuario, "pk", "")}'

    descripcion_final = (descripcion or '').strip()
    actor_tag = f'Actor: {actor}' + (f' ({rol})' if rol else '')
    if descripcion_final:
        descripcion_final = f'{descripcion_final} | {actor_tag}'
    else:
        descripcion_final = actor_tag

    ip = ''
    if request:
        ip = request.META.get('HTTP_X_FORWARDED_FOR', '') or request.META.get('REMOTE_ADDR', '')
    if ',' in ip:
        ip = ip.split(',')[0].strip()

    try:
        AuditoriaLog.objects.create(
            accion=accion,
            entidad=entidad,
            entidad_id=str(entidad_id) if entidad_id is not None else None,
            descripcion=descripcion_final,
            id_usuario_fk=usuario if usuario and usuario.is_authenticated else None,
            rol_usuario=rol,
            ip_origen=ip[:45] if ip else None,
        )
    except Exception:
        # Evita romper el flujo principal si la tabla de auditoria aun no existe.
        pass


def _auto_cancelar_pedidos_pendientes_vencidos():
    now = timezone.localtime()
    with transaction.atomic():
        pedidos = list(
            Pedido.objects
            .select_for_update()
            .select_related('id_usuario_fk')
            .filter(
                estado='pendiente',
                fecha_devolucion__isnull=False,
                fecha_devolucion__lte=now,
            )
        )

        if not pedidos:
            return 0

        pedido_ids = [p.id_pedido for p in pedidos]
        Pedido.objects.filter(id_pedido__in=pedido_ids).update(
            estado='cancelado',
            fch_ult_act=now,
        )
        DetallePedido.objects.filter(id_pedido_fk_id__in=pedido_ids).update(
            estado_detalle='cancelado',
            fch_ult_act=now,
        )

    for pedido in pedidos:
        _crear_notificacion(
            usuario=pedido.id_usuario_fk,
            tipo='rechazado',
            titulo='Pedido cancelado automáticamente',
            mensaje=(
                f'Tu pedido #{pedido.id_pedido} fue cancelado automáticamente porque '
                'la hora/fecha límite de entrega se venció antes de ser aprobado por almacén.'
            ),
            pedido_id=pedido.id_pedido,
        )
        _registrar_auditoria(
            None,
            accion='actualizar',
            entidad='pedido',
            entidad_id=pedido.id_pedido,
            descripcion=f'Pedido #{pedido.id_pedido} cancelado automáticamente por vencimiento en estado pendiente.',
        )

    return len(pedidos)


def _auto_marcar_prestamos_vencidos():
    """Marca como 'vencido' los préstamos entregados cuya fecha de devolución ya pasó."""
    now = timezone.localtime()
    with transaction.atomic():
        pedidos = list(
            Pedido.objects
            .select_for_update()
            .select_related('id_usuario_fk')
            .filter(
                estado='entregado',
                tipo_devolucion__in=['global', None, ''],
                fecha_devolucion__isnull=False,
                fecha_devolucion__lte=now,
            )
        )
        if not pedidos:
            return 0

        pedido_ids = [p.id_pedido for p in pedidos]
        Pedido.objects.filter(id_pedido__in=pedido_ids).update(
            estado='vencido',
            fch_ult_act=now,
        )

    for pedido in pedidos:
        _crear_notificacion(
            usuario=pedido.id_usuario_fk,
            tipo='prestamo_vencido',
            titulo='¡Préstamo vencido!',
            mensaje=(
                f'Tu préstamo #{pedido.id_pedido} ha vencido. '
                'Por favor acércate a almacén para devolver los productos a la brevedad posible.'
            ),
            pedido_id=pedido.id_pedido,
        )
        _registrar_auditoria(
            None,
            accion='actualizar',
            entidad='prestamo',
            entidad_id=pedido.id_pedido,
            descripcion=f'Préstamo #{pedido.id_pedido} marcado automáticamente como vencido.',
        )

    return len(pedidos)


def _sumar_stock_disponibilidad(detalle, now):
    _ajustar_stock_disponibilidad(detalle, now, detalle.cantidad_solicitada)


def _ajustar_stock_disponibilidad(detalle, now, delta):
    if not detalle.id_prod_fk_id:
        return

    disp = (
        Disponibilidad.objects
        .select_for_update()
        .filter(id_prod_fk_id=detalle.id_prod_fk_id)
        .order_by('-id_disp')
        .first()
    )

    if not disp:
        if delta <= 0:
            return
        Disponibilidad.objects.create(
            id_prod_fk=detalle.id_prod_fk,
            cantidad=delta,
            stock=delta,
            descr_dispo='Stock restaurado por devolución de préstamo.',
            fch_registro=now,
            fch_ult_act=now,
        )
        return

    update_fields = ['fch_ult_act']

    if disp.cantidad is not None:
        disp.cantidad = max((disp.cantidad or 0) + delta, 0)
        update_fields.append('cantidad')
    if disp.stock is not None:
        disp.stock = max((disp.stock or 0) + delta, 0)
        update_fields.append('stock')
    if disp.cantidad is None and disp.stock is None and delta > 0:
        disp.cantidad = delta
        disp.stock = delta
        update_fields.extend(['cantidad', 'stock'])

    disp.fch_ult_act = now
    disp.save(update_fields=update_fields)


def _renovar_codigo_devolucion(pedido, now):
    pedido.codigo_entrega = f'{secrets.randbelow(1000000):06d}'
    pedido.codigo_expira_en = now + timedelta(seconds=DEVOLUCION_CODIGO_SEGUNDOS)
    pedido.fch_ult_act = now
    pedido.save(update_fields=['codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])


def _estado_pedido_canonico(estado):
    estado_limpio = (estado or '').strip().lower().replace('_', ' ')
    aliases = {
        'deuelto': 'devuelto',
        'debuelto': 'devuelto',
        'devolvido': 'devuelto',
        'esperandoentrega': 'esperando entrega',
    }
    return aliases.get(estado_limpio, estado_limpio)


def _parse_subcategorias_text(raw_text):
    text = (raw_text or '').replace('\n', ',').replace(';', ',')
    values = []
    for part in text.split(','):
        item = (part or '').strip()
        if item:
            values.append(item)
    unique = []
    seen = set()
    for value in values:
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(value)
    return unique


def _validar_palabras_subcategoria(nombre):
    for palabra in re.findall(r'\S+', (nombre or '').strip()):
        if len(palabra) > SUBCAT_MAX_WORD_LENGTH:
            return (
                f'Cada palabra del nombre de la subcategoria debe tener '
                f'maximo {SUBCAT_MAX_WORD_LENGTH} caracteres.'
            )
    return None


def _sync_subcategorias_producto(producto, catalogo_id, selected_ids=None, raw_new=''):
    selected_ids = selected_ids or []
    selected_qs = Subcategoria.objects.filter(pk__in=selected_ids)
    related = list(selected_qs)

    catalogo = Catalogo.objects.filter(pk=catalogo_id).first() if catalogo_id else None
    for nombre in _parse_subcategorias_text(raw_new):
        if not catalogo:
            continue
        ruta = [segmento.strip() for segmento in nombre.split('/') if segmento.strip()]
        subcat = Subcategoria.ensure_path(catalogo, ruta)
        related.append(subcat)

    dedup = list({s.pk: s for s in related}.values())
    producto.subcategorias.set(dedup)


def _build_subcategoria_tree(subcategorias_qs):
    nodes = {}
    roots = []

    for subcategoria in subcategorias_qs:
        nodes[subcategoria.id_subcat] = {
            'id': subcategoria.id_subcat,
            'nombre': subcategoria.nombre_subcategoria,
            'ruta': subcategoria.ruta_completa,
            'padre_id': subcategoria.subcategoria_padre_id,
            'children': [],
        }

    for subcategoria in subcategorias_qs:
        node = nodes[subcategoria.id_subcat]
        padre_id = subcategoria.subcategoria_padre_id
        if padre_id and padre_id in nodes:
            nodes[padre_id]['children'].append(node)
        else:
            roots.append(node)

    def _sort(node_list):
        node_list.sort(key=lambda item: item['nombre'].lower())
        for item in node_list:
            _sort(item['children'])

    _sort(roots)
    return roots


def _subcategoria_descendants_ids(root_subcat):
    ids = [root_subcat.id_subcat]
    frontier = [root_subcat.id_subcat]

    while frontier:
        children = list(
            Subcategoria.objects
            .filter(subcategoria_padre_id__in=frontier)
            .values_list('id_subcat', flat=True)
        )
        if not children:
            break
        ids.extend(children)
        frontier = children

    return ids


def _subcategoria_delete_state(subcategoria):
    child_count = Subcategoria.objects.filter(subcategoria_padre=subcategoria).count()
    descendants_ids = _subcategoria_descendants_ids(subcategoria)
    products_count = (
        Producto.objects
        .filter(subcategorias__id_subcat__in=descendants_ids)
        .distinct()
        .count()
    )

    can_delete = products_count == 0
    reason = ''
    if products_count > 0:
        reason = 'No se puede eliminar porque tiene productos asociados en esta rama.'

    return {
        'can_delete': can_delete,
        'children_count': child_count,
        'products_count': products_count,
        'reason': reason,
    }


@login_required
def subcategoria_crear_rapida(request, cat_id):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    nombre = (request.POST.get('nombre_subcategoria') or '').strip()
    parent_id = (request.POST.get('subcategoria_padre_id') or '').strip()

    if not nombre:
        return JsonResponse({'ok': False, 'error': 'Debes escribir un nombre.'}, status=400)

    ruta = [segmento.strip() for segmento in nombre.split('/') if segmento.strip()]
    for segmento in ruta:
        long_word_error = _validar_palabras_subcategoria(segmento)
        if long_word_error:
            return JsonResponse({'ok': False, 'error': long_word_error}, status=400)

    parent = None
    if parent_id:
        parent = get_object_or_404(Subcategoria, pk=parent_id, id_cat_fk=catalogo)

    before_ids = set(
        Subcategoria.objects
        .filter(id_cat_fk=catalogo)
        .values_list('id_subcat', flat=True)
    )
    try:
        leaf = Subcategoria.ensure_path(catalogo, ruta, parent=parent)
    except ValidationError as exc:
        error_msg = 'Se ha alcanzado el límite máximo de 30 niveles de profundidad.'
        if hasattr(exc, 'messages') and exc.messages:
            error_msg = exc.messages[0]
        return JsonResponse({'ok': False, 'error': error_msg}, status=400)

    after_ids = set(
        Subcategoria.objects
        .filter(id_cat_fk=catalogo)
        .values_list('id_subcat', flat=True)
    )
    created_count = len(after_ids - before_ids)
    if created_count > 0:
        _registrar_auditoria(
            request,
            accion='crear',
            entidad='subcategoria',
            entidad_id=leaf.id_subcat,
            descripcion=(
                f'Se creó la subcategoría "{leaf.nombre_subcategoria}" '
                f'en catálogo "{catalogo.nombre_catalogo}". '
                f'Ruta: {leaf.ruta_completa}. Nodos creados: {created_count}.'
            ),
        )

    return JsonResponse({
        'ok': True,
        'subcategoria': {
            'id': leaf.id_subcat,
            'nombre': leaf.nombre_subcategoria,
            'ruta': leaf.ruta_completa,
            'padre_id': leaf.subcategoria_padre_id,
        },
    })


@login_required
def subcategoria_renombrar(request, cat_id, subcat_id):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    subcategoria = get_object_or_404(Subcategoria, pk=subcat_id, id_cat_fk=catalogo)
    nuevo_nombre = (request.POST.get('nombre_subcategoria') or '').strip()

    if not nuevo_nombre:
        return JsonResponse({'ok': False, 'error': 'Debes escribir un nombre.'}, status=400)

    long_word_error = _validar_palabras_subcategoria(nuevo_nombre)
    if long_word_error:
        return JsonResponse({'ok': False, 'error': long_word_error}, status=400)

    old_name = subcategoria.nombre_subcategoria
    old_route = subcategoria.ruta_completa
    subcategoria.nombre_subcategoria = nuevo_nombre
    subcategoria.fch_ult_act = timezone.now()

    try:
        subcategoria.full_clean()
        subcategoria.save(update_fields=['nombre_subcategoria', 'fch_ult_act'])
    except ValidationError as exc:
        error_msg = 'No fue posible actualizar el nombre de la subcategoría.'
        if hasattr(exc, 'messages') and exc.messages:
            error_msg = exc.messages[0]
        return JsonResponse({'ok': False, 'error': error_msg}, status=400)
    except IntegrityError:
        return JsonResponse(
            {'ok': False, 'error': 'Ya existe una subcategoría con ese nombre en este nivel.'},
            status=400,
        )

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='subcategoria',
        entidad_id=subcategoria.id_subcat,
        descripcion=(
            f'Se renombró subcategoría de "{old_name}" a "{subcategoria.nombre_subcategoria}" '
            f'en catálogo "{catalogo.nombre_catalogo}". '
            f'Ruta anterior: {old_route}. Ruta nueva: {subcategoria.ruta_completa}.'
        ),
    )

    return JsonResponse({
        'ok': True,
        'subcategoria': {
            'id': subcategoria.id_subcat,
            'nombre': subcategoria.nombre_subcategoria,
            'ruta': subcategoria.ruta_completa,
            'padre_id': subcategoria.subcategoria_padre_id,
        },
    })


@login_required
def subcategoria_delete_estado(request, cat_id, subcat_id):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    if request.method != 'GET':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    subcategoria = get_object_or_404(Subcategoria, pk=subcat_id, id_cat_fk=catalogo)
    state = _subcategoria_delete_state(subcategoria)
    return JsonResponse({'ok': True, **state})


@login_required
def subcategoria_eliminar(request, cat_id, subcat_id):
    wants_json = (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or request.headers.get('sec-fetch-dest') == 'empty'
        or 'application/json' in (request.headers.get('accept') or '')
    )

    if not _is_admin(request):
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)
        messages.error(request, 'No autorizado.')
        return redirect('productos_catalogo', cat_id=cat_id)

    if request.method not in ['POST', 'DELETE']:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)
        messages.error(request, 'Método no permitido.')
        return redirect('productos_catalogo', cat_id=cat_id)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    subcategoria = get_object_or_404(Subcategoria, pk=subcat_id, id_cat_fk=catalogo)

    state = _subcategoria_delete_state(subcategoria)
    if not state['can_delete']:
        if wants_json:
            return JsonResponse({'ok': False, 'error': state['reason']}, status=400)
        messages.error(request, state['reason'])
        return redirect('productos_catalogo', cat_id=cat_id)

    deleted_id = subcategoria.id_subcat
    deleted_name = subcategoria.nombre_subcategoria
    deleted_route = subcategoria.ruta_completa
    subcategoria.delete()

    _registrar_auditoria(
        request,
        accion='eliminar',
        entidad='subcategoria',
        entidad_id=deleted_id,
        descripcion=(
            f'Se eliminó subcategoría "{deleted_name}" en catálogo "{catalogo.nombre_catalogo}". '
            f'Ruta eliminada: {deleted_route}.'
        ),
    )

    if wants_json:
        return JsonResponse({
            'ok': True,
            'subcategoria': {
                'id': deleted_id,
                'nombre': deleted_name,
            },
        })

    messages.success(request, f'Subcategoría "{deleted_name}" eliminada correctamente.')
    return redirect('productos_catalogo', cat_id=cat_id)


@login_required
def producto_mover_subcategoria(request, cat_id, prod_id):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    producto = get_object_or_404(Producto, pk=prod_id, id_cat_fk=catalogo)

    prev_ids = list(producto.subcategorias.values_list('id_subcat', flat=True))
    prev_rutas = list(
        Subcategoria.objects
        .filter(id_cat_fk=catalogo, id_subcat__in=prev_ids)
        .order_by('nombre_subcategoria')
        .values_list('nombre_subcategoria', flat=True)
    )
    destino_raw = (request.POST.get('subcategoria_destino_id') or '').strip()
    if not destino_raw:
        # Permite soltar sobre la ruta raíz del catálogo (sin subcategoría).
        producto.subcategorias.clear()
        _registrar_auditoria(
            request,
            accion='actualizar',
            entidad='producto',
            entidad_id=producto.id_prod,
            descripcion=(
                f'Se movió producto "{producto.nombre_producto}" a raíz del catálogo '
                f'"{catalogo.nombre_catalogo}". '
                f'Subcategorías anteriores: {", ".join(prev_rutas) if prev_rutas else "ninguna"}.'
            ),
        )
        return JsonResponse({
            'ok': True,
            'producto': {
                'id': producto.id_prod,
                'prev_subcategoria_ids': prev_ids,
                'new_subcategoria_ids': [],
            },
            'destino': {
                'id': None,
                'ruta': catalogo.nombre_catalogo,
            },
        })

    try:
        destino_id = int(destino_raw)
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Subcategoría destino inválida.'}, status=400)

    destino = get_object_or_404(Subcategoria, pk=destino_id, id_cat_fk=catalogo)
    producto.subcategorias.set([destino])

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='producto',
        entidad_id=producto.id_prod,
        descripcion=(
            f'Se movió producto "{producto.nombre_producto}" a subcategoría "{destino.nombre_subcategoria}" '
            f'en catálogo "{catalogo.nombre_catalogo}". '
            f'Ruta destino: {destino.ruta_completa}. '
            f'Subcategorías anteriores: {", ".join(prev_rutas) if prev_rutas else "ninguna"}.'
        ),
    )

    return JsonResponse({
        'ok': True,
        'producto': {
            'id': producto.id_prod,
            'prev_subcategoria_ids': prev_ids,
            'new_subcategoria_ids': [destino.id_subcat],
        },
        'destino': {
            'id': destino.id_subcat,
            'ruta': destino.ruta_completa,
        },
    })


@login_required
def producto_restaurar_subcategorias(request, cat_id, prod_id):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    producto = get_object_or_404(Producto, pk=prod_id, id_cat_fk=catalogo)
    prev_ids = list(producto.subcategorias.values_list('id_subcat', flat=True))

    raw_ids = (request.POST.get('subcategoria_ids') or '').strip()
    if not raw_ids:
        producto.subcategorias.clear()
        _registrar_auditoria(
            request,
            accion='actualizar',
            entidad='producto',
            entidad_id=producto.id_prod,
            descripcion=(
                f'Se restauró (Ctrl+Z) producto "{producto.nombre_producto}" a raíz del catálogo '
                f'"{catalogo.nombre_catalogo}". IDs anteriores: {prev_ids}.'
            ),
        )
        return JsonResponse({'ok': True, 'producto': {'id': producto.id_prod, 'subcategoria_ids': []}})

    parsed_ids = []
    for part in raw_ids.split(','):
        part = part.strip()
        if not part:
            continue
        try:
            parsed_ids.append(int(part))
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'Lista de subcategorías inválida.'}, status=400)

    valid_ids = list(
        Subcategoria.objects
        .filter(id_cat_fk=catalogo, id_subcat__in=parsed_ids)
        .values_list('id_subcat', flat=True)
    )
    valid_set = set(valid_ids)
    if len(valid_set) != len(set(parsed_ids)):
        return JsonResponse({'ok': False, 'error': 'Hay subcategorías inválidas para restaurar.'}, status=400)

    ordered_unique_ids = []
    seen = set()
    for item in parsed_ids:
        if item in seen:
            continue
        seen.add(item)
        ordered_unique_ids.append(item)

    producto.subcategorias.set(ordered_unique_ids)
    restored_rutas = list(
        Subcategoria.objects
        .filter(id_cat_fk=catalogo, id_subcat__in=ordered_unique_ids)
        .order_by('nombre_subcategoria')
        .values_list('nombre_subcategoria', flat=True)
    )
    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='producto',
        entidad_id=producto.id_prod,
        descripcion=(
            f'Se restauró (Ctrl+Z) producto "{producto.nombre_producto}" en catálogo '
            f'"{catalogo.nombre_catalogo}". IDs anteriores: {prev_ids}. '
            f'IDs restaurados: {ordered_unique_ids}. '
            f'Subcategorías restauradas: {", ".join(restored_rutas) if restored_rutas else "ninguna"}.'
        ),
    )
    return JsonResponse({
        'ok': True,
        'producto': {
            'id': producto.id_prod,
            'subcategoria_ids': ordered_unique_ids,
        },
    })


@login_required
def producto_editar(request, prod_id):
    # Solo admin puede editar
    if not (request.user.is_authenticated and request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'No tienes permisos para editar productos.')
        return redirect('producto_detalle', prod_id=prod_id)

    producto = get_object_or_404(Producto, pk=prod_id)
    catalogos = Catalogo.objects.all().order_by('nombre_catalogo')
    ubicaciones_producto = UbicacionProducto.objects.order_by('nombre')
    subcategorias = Subcategoria.objects.select_related('id_cat_fk').order_by('id_cat_fk__nombre_catalogo', 'nombre_subcategoria')
    disp = (
        Disponibilidad.objects
        .filter(id_prod_fk=producto)
        .order_by('-id_disp')
        .first()
    )
    if request.method == 'POST':
        if request.POST.get('generar_codigo_producto') == '1':
            if (producto.codigo_producto or '').strip():
                messages.info(request, f'El producto ya tiene código: {producto.codigo_producto}.')
            else:
                codigo_generado = _asignar_codigo_producto_si_falta(producto)
                if codigo_generado:
                    _registrar_auditoria(
                        request,
                        accion='editar',
                        entidad='producto',
                        entidad_id=producto.id_prod,
                        descripcion=(
                            f'Se generó código para el producto "{producto.nombre_producto}": {codigo_generado}.'
                        ),
                    )
                    messages.success(request, f'Código generado correctamente: {codigo_generado}.')
                else:
                    messages.error(request, 'No se pudo generar un código único en este momento. Intenta de nuevo.')
            return redirect('producto_editar', prod_id=producto.id_prod)

        nombre = request.POST.get('nombre_producto', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        id_cat_fk = request.POST.get('id_cat_fk')
        unidad_medida = request.POST.get('unidad_medida', 'unidad').strip() or 'unidad'
        id_ubicacion_fk = (request.POST.get('id_ubicacion_fk') or '').strip()
        ubicacion_obj = None
        if id_ubicacion_fk.isdigit():
            ubicacion_obj = UbicacionProducto.objects.filter(pk=int(id_ubicacion_fk)).first()
        ubicacion = (ubicacion_obj.nombre if ubicacion_obj else request.POST.get('ubicacion', '').strip())
        tipo_bien = request.POST.get('tipo_bien', 'devolutivo').strip() or 'devolutivo'
        numero_placa = request.POST.get('numero_placa', '').strip()
        cuentadante = request.POST.get('cuentadante', '').strip()
        stock = request.POST.get('stock')
        cantidad = request.POST.get('cantidad')
        descr_dispo = request.POST.get('descr_dispo', '').strip()
        fot_prod = request.FILES.get('fot_prod')
        clear_fot_prod = request.POST.get('clear_fot_prod')

        # Validaciones mínimas
        if not nombre or not id_cat_fk or not ubicacion or stock is None or cantidad is None:
            messages.error(request, 'Completa todos los campos obligatorios.')
        elif tipo_bien == 'devolutivo' and (not numero_placa or not cuentadante):
            messages.error(request, 'Para bienes devolutivos debes registrar número de placa y cuentadante.')
        else:
            producto.nombre_producto = nombre
            producto.descripcion = descripcion
            producto.id_cat_fk_id = id_cat_fk
            producto.unidad_medida = unidad_medida
            producto.ubicacion = ubicacion
            producto.tipo_bien = tipo_bien
            producto.numero_placa = numero_placa if tipo_bien == 'devolutivo' else ''
            producto.cuentadante = cuentadante if tipo_bien == 'devolutivo' else ''

            # Gestión foto principal
            if fot_prod:
                if producto.fot_prod:
                    producto.fot_prod.delete(save=False)
                producto.fot_prod = fot_prod
            elif clear_fot_prod and producto.fot_prod:
                producto.fot_prod.delete(save=False)
                producto.fot_prod = None

            producto.fch_ult_act = timezone.now()
            producto.save()

            # Completa código histórico ausente para productos viejos.
            if not (producto.codigo_producto or '').strip():
                _asignar_codigo_producto_si_falta(producto)

            # Fotos adicionales nuevas (respetando máximo de 5 total)
            fotos_nuevas = request.FILES.getlist('fotos_nuevas')
            if fotos_nuevas:
                total_actuales = producto.fotos.count() + (1 if producto.fot_prod else 0)
                cupo = max(0, 5 - total_actuales)
                orden_base = (producto.fotos.order_by('-orden').values_list('orden', flat=True).first() or 0) + 1
                for i, f in enumerate(fotos_nuevas[:cupo]):
                    ProductoFoto.objects.create(id_prod_fk=producto, foto=f, orden=orden_base + i)

            # Actualizar disponibilidad
            if disp:
                disp.stock = stock
                disp.cantidad = cantidad
                disp.descr_dispo = descr_dispo
                disp.fch_ult_act = timezone.now()
                disp.save()
            else:
                Disponibilidad.objects.create(
                    id_prod_fk=producto,
                    stock=stock,
                    cantidad=cantidad,
                    descr_dispo=descr_dispo,
                    fch_registro=timezone.now(),
                    fch_ult_act=timezone.now(),
                )
            messages.success(request, 'Producto actualizado correctamente.')
            return redirect('producto_editar', prod_id=producto.id_prod)

    fotos_extra = producto.fotos.all()
    total_fotos = fotos_extra.count() + (1 if producto.fot_prod else 0)
    ubicacion_actual_en_lista = ubicaciones_producto.filter(nombre=producto.ubicacion).exists()
    return render(request, 'inventario/catalogo/producto_editar.html', {
        'producto': producto,
        'catalogos': catalogos,
        'ubicaciones_producto': ubicaciones_producto,
        'ubicacion_actual_en_lista': ubicacion_actual_en_lista,
        'subcategorias': subcategorias,
        'disponibilidad': disp,
        'fotos_extra': fotos_extra,
        'total_fotos': total_fotos,
        'cupo_fotos': max(0, 5 - total_fotos),
    })


@login_required
def eliminar_foto_producto(request, prod_id, foto_id):
    if not _is_admin(request):
        messages.error(request, 'No tienes permisos para eliminar fotos.')
        return redirect('producto_detalle', prod_id=prod_id)

    foto = get_object_or_404(ProductoFoto, pk=foto_id, id_prod_fk_id=prod_id)
    if request.method == 'POST':
        if foto.foto:
            foto.foto.delete(save=False)
        foto.delete()
        messages.success(request, 'Foto eliminada correctamente.')
    return redirect('producto_editar', prod_id=prod_id)


@login_required
def producto_detalle(request, prod_id):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    from .models import Producto, Disponibilidad
    try:
        producto = Producto.objects.select_related('id_cat_fk').prefetch_related('subcategorias').get(pk=prod_id)
    except Producto.DoesNotExist:
        raise Http404('Producto no encontrado')
    disp = (
        Disponibilidad.objects
        .filter(id_prod_fk=producto)
        .order_by('-id_disp')
        .first()
    )
    fotos_extra = producto.fotos.all()
    return render(request, 'inventario/catalogo/producto_detalle.html', {
        'producto': producto,
        'catalogo': producto.id_cat_fk,
        'disponibilidad': disp,
        'fotos_extra': fotos_extra,
    })

# Panel de almacenista
from django.contrib.auth.decorators import login_required

@login_required
def panel_almacenista(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')
    return redirect('inventario_panel')
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.db import models
from django.db import transaction
from django.db.models import OuterRef, Subquery
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .db_compat import usuario_supports_tipo_doc
from .forms import CambioPasswordPerfilForm, CatalogoForm, ProductoForm, UbicacionProductoForm, UsuarioPerfilForm
from .models import AuditoriaLog, Catalogo, DetallePedido, Disponibilidad, Pedido, PedidoEvidencia, Producto, ProductoFoto, Subcategoria, UbicacionProducto, Usuario, Rol, VerificacionSenaToken
from .views_catalogo_panel import build_catalogo_cards, build_catalogo_tree


def _user_role(request):
    if not request.user.is_authenticated:
        return None
    if getattr(request.user, 'is_superuser', False) or getattr(request.user, 'is_staff', False):
        return 'admin'

    rol = (getattr(getattr(request.user, 'id_rol_fk', None), 'nombre_rol', '') or '').strip().lower()
    if rol in {'admin', 'administrador'}:
        return 'admin'
    if rol in {'almacenista', 'almacen'}:
        return 'almacenista'
    if rol in {'', 'usuario', 'aprendiz', 'instructor'}:
        return 'usuario'
    return rol


def _is_admin(request):
    return _user_role(request) == 'admin'


def _is_admin_or_almacenista(request):
    return _user_role(request) in ['admin', 'almacenista']


@login_required
def catalogo(request):
    if not _is_admin_or_almacenista(request):
        return redirect('dashboard')

    catalogos = (
        Catalogo.objects
        .select_related('id_ubicacion_fk')
        .annotate(total_productos=models.Count('producto'))
        .order_by('nombre_catalogo')
    )
    macro_cards = build_catalogo_cards(catalogos)
    catalogo_ubicaciones_map = {
        str(cat.id_cat): (cat.id_ubicacion_fk.nombre if cat.id_ubicacion_fk else '')
        for cat in catalogos
    }
    cat_form = CatalogoForm()
    ubi_form = UbicacionProductoForm()
    ubicaciones_producto = UbicacionProducto.objects.order_by('nombre')
    prod_form = ProductoForm()
    return render(
        request,
        'inventario/catalogo/catalogo.html',
        {
            'catalogos': catalogos,
            'macro_cards': macro_cards,
            'cat_form': cat_form,
            'ubi_form': ubi_form,
            'ubicaciones_producto': ubicaciones_producto,
            'prod_form': prod_form,
            'catalogo_ubicaciones_map': catalogo_ubicaciones_map,
            'puede_gestionar_catalogo': _is_admin(request),
        },
    )


@login_required
def registrar_catalogo(request):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede registrar catalogos.')
        return redirect('catalogo')

    if request.method == 'POST':
        form = CatalogoForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.fch_registro = timezone.now()
            obj.fch_ult_act = timezone.now()
            obj.save()
            _registrar_auditoria(
                request,
                accion='crear',
                entidad='catalogo',
                entidad_id=obj.id_cat,
                descripcion=f'Se creó el catálogo "{obj.nombre_catalogo}".',
            )
            messages.success(request, f'Catálogo "{obj.nombre_catalogo}" registrado correctamente.')
        else:
            messages.error(request, 'Error al registrar el catálogo. Revisa los campos.')
    return redirect('catalogo')


@login_required
def editar_catalogo(request):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede editar catalogos.')
        return redirect('catalogo')

    if request.method != 'POST':
        return redirect('catalogo')

    cat_id = (request.POST.get('cat_id') or '').strip()
    if not cat_id.isdigit():
        messages.error(request, 'No se pudo identificar la macro categoría a editar.')
        return redirect('catalogo')

    catalogo = get_object_or_404(Catalogo, pk=int(cat_id))
    codigo_actual = (catalogo.codigo_macro or '').strip().upper()
    codigo_enviado = (request.POST.get('codigo_macro') or '').strip().upper()

    # Si la macro ya tiene código, ese código queda inmutable.
    if codigo_actual and codigo_enviado != codigo_actual:
        messages.error(request, 'El código de una macro ya registrada no se puede modificar.')
        return redirect('catalogo')

    form = CatalogoForm(request.POST, instance=catalogo)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.fch_ult_act = timezone.now()
        obj.save()
        _registrar_auditoria(
            request,
            accion='editar',
            entidad='catalogo',
            entidad_id=obj.id_cat,
            descripcion=f'Se editó la macro categoría "{obj.nombre_catalogo}".',
        )
        messages.success(request, f'Macro categoría "{obj.nombre_catalogo}" actualizada correctamente.')
    else:
        messages.error(request, 'Error al editar la macro categoría. Revisa los campos.')

    return redirect('catalogo')


@login_required
def registrar_ubicacion_producto(request):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede registrar ubicaciones de productos.')
        return redirect('catalogo')

    if request.method == 'POST':
        form = UbicacionProductoForm(request.POST)
        if form.is_valid():
            ubicacion = form.save(commit=False)
            ubicacion.fch_registro = timezone.now()
            ubicacion.fch_ult_act = timezone.now()
            ubicacion.save()
            _registrar_auditoria(
                request,
                accion='crear',
                entidad='ubicacion_producto',
                entidad_id=ubicacion.id_ubicacion,
                descripcion=f'Se creó la ubicación "{ubicacion.nombre}".',
            )
            messages.success(request, f'Ubicación "{ubicacion.nombre}" registrada correctamente.')
        else:
            messages.error(request, 'Error al registrar la ubicación. Revisa el formulario.')

    return redirect('catalogo')


@login_required
def editar_ubicacion_producto(request, ubicacion_id):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede editar ubicaciones de productos.')
        return redirect('catalogo')

    if request.method != 'POST':
        return redirect('catalogo')

    ubicacion = get_object_or_404(UbicacionProducto, pk=ubicacion_id)
    form = UbicacionProductoForm(request.POST, instance=ubicacion)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.fch_ult_act = timezone.now()
        obj.save()
        _registrar_auditoria(
            request,
            accion='editar',
            entidad='ubicacion_producto',
            entidad_id=obj.id_ubicacion,
            descripcion=f'Se editó la ubicación "{obj.nombre}".',
        )
        messages.success(request, f'Ubicación "{obj.nombre}" actualizada correctamente.')
    else:
        messages.error(request, 'No se pudo editar la ubicación. Verifica el nombre (único y válido).')

    return redirect('catalogo')


@login_required
def registrar_producto(request):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede registrar productos.')
        return redirect('catalogo')

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            subcategoria_seleccionada = form.cleaned_data.get('subcategoria')
            obj = form.save(commit=False)
            obj.fch_registro = timezone.now()
            obj.fch_ult_act = timezone.now()

            # Múltiples fotos: primera → fot_prod, resto → ProductoFoto (máx 5 total)
            fotos = request.FILES.getlist('fotos')[:5]
            if fotos:
                obj.fot_prod = fotos[0]

            with transaction.atomic():
                obj.save()
                if subcategoria_seleccionada:
                    _sync_subcategorias_producto(
                        obj,
                        obj.id_cat_fk_id,
                        [str(subcategoria_seleccionada.pk)],
                        '',
                    )

                    # Se genera al final para garantizar prefijo correcto por macro/categoría/subcategoría.
                    for _ in range(3):
                        try:
                            obj.codigo_producto = _generar_codigo_producto(subcategoria_seleccionada)
                            obj.save(update_fields=['codigo_producto'])
                            break
                        except IntegrityError:
                            continue

            # Fotos adicionales (índices 1-4)
            for i, f in enumerate(fotos[1:], start=1):
                ProductoFoto.objects.create(id_prod_fk=obj, foto=f, orden=i)

            _registrar_auditoria(
                request,
                accion='crear',
                entidad='producto',
                entidad_id=obj.id_prod,
                descripcion=(
                    f'Se creó el producto "{obj.nombre_producto}" '
                    f'({obj.get_tipo_bien_display()}, unidad: {obj.get_unidad_medida_display()}, ubicación: {obj.ubicacion}, código: {obj.codigo_producto or "N/A"}).'
                ),
            )

            stock_inicial = form.cleaned_data.get('stock_inicial') or 0
            descr_dispo = form.cleaned_data.get('descr_dispo') or ''
            Disponibilidad.objects.create(
                id_prod_fk=obj,
                cantidad=stock_inicial,
                stock=stock_inicial,
                descr_dispo=descr_dispo,
                fch_registro=timezone.now(),
                fch_ult_act=timezone.now(),
            )

            messages.success(request, f'Producto "{obj.nombre_producto}" registrado correctamente.')
        else:
            messages.error(request, 'Error al registrar el producto. Revisa los campos.')
    return redirect('catalogo')


@login_required
def eliminar_catalogo(request, cat_id):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede eliminar catalogos.')
        return redirect('catalogo')

    catalogo = get_object_or_404(Catalogo, pk=cat_id)

    if request.method == 'POST':
        if Producto.objects.filter(id_cat_fk=catalogo).exists():
            messages.error(
                request,
                f'No se puede eliminar el catálogo "{catalogo.nombre_catalogo}" porque tiene productos registrados.',
            )
        else:
            nombre = catalogo.nombre_catalogo
            catalogo_id = catalogo.id_cat
            catalogo.delete()
            _registrar_auditoria(
                request,
                accion='eliminar',
                entidad='catalogo',
                entidad_id=catalogo_id,
                descripcion=f'Se eliminó el catálogo "{nombre}".',
            )
            messages.success(request, f'Catálogo "{nombre}" eliminado correctamente.')

    return redirect('catalogo')


@login_required
def productos_catalogo(request, cat_id):
    if not _is_admin_or_almacenista(request):
        return redirect('dashboard')

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    selected_subcat_id = (request.GET.get('subcat') or '').strip()
    selected_subcat = None
    if selected_subcat_id:
        try:
            selected_subcat = Subcategoria.objects.get(pk=int(selected_subcat_id), id_cat_fk=catalogo)
        except (ValueError, Subcategoria.DoesNotExist):
            selected_subcat = None

    disp_qs = Disponibilidad.objects.filter(id_prod_fk=OuterRef('pk')).order_by('-id_disp')
    productos_qs = (
        Producto.objects
        .filter(id_cat_fk=catalogo)
        .annotate(
            stock_actual=Subquery(disp_qs.values('stock')[:1]),
            cantidad_total=Subquery(disp_qs.values('cantidad')[:1]),
            descr_dispo_actual=Subquery(disp_qs.values('descr_dispo')[:1]),
        )
        .order_by('nombre_producto')
        .prefetch_related('fotos', 'subcategorias')
    )

    if selected_subcat:
        # Muestra solo productos del nivel actual para evitar que el padre
        # replique visualmente los productos de hijas/descendientes.
        productos = productos_qs.filter(subcategorias=selected_subcat).distinct()
    else:
        # En raíz solo se listan productos sin subcategoría para que el movimiento
        # a carpetas sea real (no se vea como copia).
        productos = productos_qs.filter(subcategorias__isnull=True).distinct()

    subcategoria_tree = build_catalogo_tree(catalogo, selected_subcat)
    breadcrumb_subcats = []
    if selected_subcat:
        node = selected_subcat
        while node:
            breadcrumb_subcats.append(node)
            node = node.subcategoria_padre
        breadcrumb_subcats.reverse()

    if selected_subcat and selected_subcat.subcategoria_padre_id:
        back_url = f"{reverse('productos_catalogo', kwargs={'cat_id': catalogo.id_cat})}?subcat={selected_subcat.subcategoria_padre_id}"
    elif selected_subcat:
        back_url = reverse('productos_catalogo', kwargs={'cat_id': catalogo.id_cat})
    else:
        back_url = reverse('catalogo')

    return render(
        request,
        'inventario/catalogo/productos_catalogo.html',
        {
            'catalogo': catalogo,
            'productos': productos,
            'subcategoria_tree': subcategoria_tree,
            'selected_subcat_id': selected_subcat.id_subcat if selected_subcat else None,
            'breadcrumb_subcats': breadcrumb_subcats,
            'back_url': back_url,
            'puede_gestionar_catalogo': _is_admin(request),
        },
    )


@login_required
def eliminar_producto(request, cat_id, prod_id):
    if not _is_admin(request):
        messages.error(request, 'Solo el administrador puede eliminar productos.')
        return redirect('productos_catalogo', cat_id=cat_id)

    catalogo = get_object_or_404(Catalogo, pk=cat_id)
    producto = get_object_or_404(Producto, pk=prod_id, id_cat_fk=catalogo)

    if request.method == 'POST':
        nombre = producto.nombre_producto
        producto_id = producto.id_prod
        producto.delete()
        _registrar_auditoria(
            request,
            accion='eliminar',
            entidad='producto',
            entidad_id=producto_id,
            descripcion=f'Se eliminó el producto "{nombre}".',
        )
        messages.success(request, f'Producto "{nombre}" eliminado correctamente.')

    return redirect('productos_catalogo', cat_id=cat_id)



@login_required
def dashboard(request):
    if not _is_admin(request):
        if _user_role(request) == 'almacenista':
            return redirect('inventario_panel')
        return redirect('panel_usuario')

    _auto_cancelar_pedidos_pendientes_vencidos()
    _expirar_solicitudes_validacion_manual()
    _reabrir_solicitudes_con_enlace_vencido()

    ahora = timezone.localtime()
    anio_actual = ahora.year
    mes_actual = ahora.month
    disp_qs = Disponibilidad.objects.filter(id_prod_fk=OuterRef('pk')).order_by('-id_disp')

    alertas_stock_bajo_raw = list(
        Producto.objects
        .select_related('id_cat_fk')
        .annotate(stock_actual=Subquery(disp_qs.values('stock')[:1]))
        .filter(stock_actual__isnull=False, stock_actual__lt=5)
        .order_by('stock_actual', 'nombre_producto')[:8]
    )

    alertas_stock_bajo = []
    for item in alertas_stock_bajo_raw:
        stock = int(item.stock_actual or 0)
        nivel = 'Critico' if stock <= 2 else 'Bajo'
        detalle = 'Reposicion urgente' if stock <= 2 else 'Planificar reposicion'
        alertas_stock_bajo.append({
            'nombre_producto': item.nombre_producto,
            'catalogo': item.id_cat_fk.nombre_catalogo if item.id_cat_fk else 'Sin catalogo',
            'stock_actual': stock,
            'nivel': nivel,
            'detalle': detalle,
        })

    alertas_cantidad_baja_raw = list(
        Producto.objects
        .select_related('id_cat_fk')
        .annotate(cantidad_actual=Subquery(disp_qs.values('cantidad')[:1]))
        .filter(cantidad_actual__isnull=False, cantidad_actual__lt=5)
        .order_by('cantidad_actual', 'nombre_producto')[:8]
    )

    alertas_cantidad_baja = []
    for item in alertas_cantidad_baja_raw:
        cantidad = int(item.cantidad_actual or 0)
        nivel = 'Critico' if cantidad <= 2 else 'Bajo'
        detalle = 'Revisar disponibilidad inmediata' if cantidad <= 2 else 'Programar abastecimiento'
        alertas_cantidad_baja.append({
            'nombre_producto': item.nombre_producto,
            'catalogo': item.id_cat_fk.nombre_catalogo if item.id_cat_fk else 'Sin catalogo',
            'cantidad_actual': cantidad,
            'nivel': nivel,
            'detalle': detalle,
        })

    productos_con_existencia = list(
        Producto.objects
        .select_related('id_cat_fk')
        .annotate(
            stock_actual=Subquery(disp_qs.values('stock')[:1]),
            cantidad_actual=Subquery(disp_qs.values('cantidad')[:1]),
        )
        .order_by('nombre_producto')
    )

    total_stock_general = 0
    total_cantidad_general = 0
    productos_deficit_base = []
    for item in productos_con_existencia:
        stock = max(int(item.stock_actual or 0), 0)
        cantidad = max(int(item.cantidad_actual or 0), 0)
        total_stock_general += stock
        total_cantidad_general += cantidad

        if cantidad < stock:
            productos_deficit_base.append({
                'producto_id': item.id_prod,
                'nombre_producto': item.nombre_producto or f'Producto {item.id_prod}',
                'catalogo': item.id_cat_fk.nombre_catalogo if item.id_cat_fk else 'Sin catalogo',
                'stock_actual': stock,
                'cantidad_actual': cantidad,
                'faltante': stock - cantidad,
            })

    deficit_producto_ids = [item['producto_id'] for item in productos_deficit_base]
    detalle_por_producto = defaultdict(dict)
    if deficit_producto_ids:
        resumen_detalles = (
            DetallePedido.objects
            .filter(
                id_prod_fk_id__in=deficit_producto_ids,
                id_pedido_fk__estado__in=['entregado', 'pendiente', 'esperando entrega'],
            )
            .values('id_prod_fk_id', 'id_pedido_fk__estado')
            .annotate(
                total=models.Sum('cantidad_solicitada'),
                pedido_ref=models.Max('id_pedido_fk_id'),
            )
        )
        for item in resumen_detalles:
            detalle_por_producto[item['id_prod_fk_id']][item['id_pedido_fk__estado']] = {
                'total': int(item['total'] or 0),
                'pedido_ref': item['pedido_ref'],
            }

    productos_deficit = []
    for base in productos_deficit_base:
        estados = detalle_por_producto.get(base['producto_id'], {})
        entregado = estados.get('entregado', {})
        pendiente = estados.get('pendiente', {})
        esperando = estados.get('esperando entrega', {})

        motivo = f'Diferencia inventario: faltan {base["faltante"]} und por ajuste de disponibilidad.'
        pedido_ref = None

        if int(entregado.get('total', 0)) > 0:
            motivo = f'En prestamos activos: {entregado["total"]} und comprometidas.'
            pedido_ref = entregado.get('pedido_ref')
        elif int(pendiente.get('total', 0)) > 0 or int(esperando.get('total', 0)) > 0:
            total_comprometido = int(pendiente.get('total', 0)) + int(esperando.get('total', 0))
            motivo = f'Comprometido en pedidos por entregar: {total_comprometido} und.'
            pedido_ref = esperando.get('pedido_ref') or pendiente.get('pedido_ref')

        productos_deficit.append({
            **base,
            'motivo': motivo,
            'pedido_ref': pedido_ref,
        })

    productos_deficit.sort(key=lambda item: (-item['faltante'], item['cantidad_actual'], item['nombre_producto']))

    pie_stock_cantidad_segmentos = []
    pie_stock_cantidad_tramos = []
    total_stock_cantidad = total_stock_general + total_cantidad_general
    tramo_acumulado = 0.0
    if total_stock_cantidad > 0:
        segmentos = [
            ('Stock total', total_stock_general, '#2d6cdf'),
            ('Cantidad total', total_cantidad_general, '#22a06b'),
        ]
        for etiqueta, cantidad, color in segmentos:
            if cantidad <= 0:
                continue
            porcentaje = round((cantidad / total_stock_cantidad) * 100, 1)
            inicio = tramo_acumulado
            tramo_acumulado += porcentaje
            pie_stock_cantidad_tramos.append(f'{color} {inicio:.2f}% {tramo_acumulado:.2f}%')
            pie_stock_cantidad_segmentos.append({
                'label': etiqueta,
                'cantidad': cantidad,
                'porcentaje': porcentaje,
                'color': color,
            })

    pie_stock_cantidad_conic = (
        'conic-gradient(' + (', '.join(pie_stock_cantidad_tramos) if pie_stock_cantidad_tramos else '#dce5de 0% 100%') + ')'
    )

    total_productos = Producto.objects.count()
    prestamos_activos = Pedido.objects.filter(estado='entregado').count()
    month_keys, resumen_mensual = _resumen_pedidos_mensual(ahora, meses=12)
    key_actual = (anio_actual, mes_actual)
    estado_conteos = resumen_mensual.get(key_actual, {
        'pendiente': 0,
        'esperando entrega': 0,
        'entregado': 0,
        'devuelto': 0,
        'cancelado': 0,
    })
    pedidos_mes_actual = sum(estado_conteos.values())
    pendientes_preview_limit = 6
    pedidos_pendientes_total = Pedido.objects.filter(estado='pendiente').count()
    pedidos_pendientes_qs = (
        Pedido.objects
        .filter(estado='pendiente')
        .select_related('id_usuario_fk')
        .prefetch_related('detalles')
        .order_by('fch_registro', 'id_pedido')[:pendientes_preview_limit]
    )

    resumen_pendientes = []
    for pedido in pedidos_pendientes_qs:
        detalles = list(pedido.detalles.all())
        if detalles:
            producto_label = detalles[0].nombre_producto or f'Producto {detalles[0].id_prod_fk_id}'
            if len(detalles) > 1:
                producto_label = f'{producto_label} +{len(detalles)-1}'
        else:
            producto_label = 'Sin productos'

        usuario_label = (
            f'{pedido.id_usuario_fk.nombre or ""} {pedido.id_usuario_fk.apellido or ""}'.strip()
            if pedido.id_usuario_fk_id else ''
        )
        if not usuario_label:
            usuario_label = pedido.id_usuario_fk.correo if pedido.id_usuario_fk_id else 'Sin usuario'

        resumen_pendientes.append({
            'id_pedido': pedido.id_pedido,
            'usuario': usuario_label,
            'producto': producto_label,
            'fecha_solicitud': pedido.fch_registro,
            'fecha_entrega': pedido.fecha_devolucion,
            'codigo_confirmacion': pedido.codigo_entrega or '--',
            'estado': 'Pendiente',
        })

    usuarios_solicitud_manual_qs = (
        Usuario.objects
        .filter(verificacion_sena_estado='solicitada')
        .order_by('verificacion_sena_solicitada_en', 'id_usu')
    )
    usuarios_solicitud_manual_total = usuarios_solicitud_manual_qs.count()
    usuarios_solicitud_manual = []
    for usuario in usuarios_solicitud_manual_qs[:8]:
        nombre_usuario = (f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() or usuario.correo or f'Usuario {usuario.pk}')
        intento_url = ''
        if getattr(usuario, 'verificacion_sena_imagen', None):
            try:
                intento_url = usuario.verificacion_sena_imagen.url
            except Exception:
                intento_url = ''

        usuarios_solicitud_manual.append({
            'id': usuario.pk,
            'nombre': nombre_usuario,
            'correo': usuario.correo or '-',
            'documento': usuario.cc or '-',
            'estado': usuario.get_verificacion_sena_estado_display(),
            'observacion': (usuario.verificacion_sena_observacion or '').strip(),
            'solicitada_en': usuario.verificacion_sena_solicitada_en,
            'intento_url': intento_url,
        })

    usuarios_solicitud_manual_modal = []
    for usuario in usuarios_solicitud_manual_qs:
        nombre_usuario = (f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() or usuario.correo or f'Usuario {usuario.pk}')
        intento_url = ''
        if getattr(usuario, 'verificacion_sena_imagen', None):
            try:
                intento_url = usuario.verificacion_sena_imagen.url
            except Exception:
                intento_url = ''

        usuarios_solicitud_manual_modal.append({
            'id': usuario.pk,
            'nombre': nombre_usuario,
            'correo': usuario.correo or '-',
            'documento': usuario.cc or '-',
            'estado': usuario.get_verificacion_sena_estado_display(),
            'observacion': (usuario.verificacion_sena_observacion or '').strip(),
            'solicitada_en': usuario.verificacion_sena_solicitada_en,
            'intento_url': intento_url,
        })

    usuarios_documento_validacion_qs = (
        Usuario.objects
        .filter(verificacion_sena_estado='documento_cargado')
        .order_by('verificacion_sena_solicitada_en', 'id_usu')
    )
    usuarios_documento_validacion_total = usuarios_documento_validacion_qs.count()
    usuarios_documento_validacion = []
    for usuario in usuarios_documento_validacion_qs[:12]:
        nombre_usuario = (f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() or usuario.correo or f'Usuario {usuario.pk}')

        intento_url = ''
        if getattr(usuario, 'verificacion_sena_imagen', None):
            try:
                intento_url = usuario.verificacion_sena_imagen.url
            except Exception:
                intento_url = ''

        documento_url = ''
        if getattr(usuario, 'verificacion_sena_documento', None):
            try:
                documento_url = usuario.verificacion_sena_documento.url
            except Exception:
                documento_url = ''

        usuarios_documento_validacion.append({
            'id': usuario.pk,
            'nombre': nombre_usuario,
            'correo': usuario.correo or '-',
            'documento': usuario.cc or '-',
            'estado': usuario.get_verificacion_sena_estado_display(),
            'observacion': (usuario.verificacion_sena_observacion or '').strip(),
            'solicitada_en': usuario.verificacion_sena_solicitada_en,
            'intento_url': intento_url,
            'documento_url': documento_url,
        })
    productos_en_mora = (
        DetallePedido.objects
        .filter(
            id_pedido_fk__estado='entregado',
            fecha_devolucion__isnull=False,
            fecha_devolucion__lt=ahora,
        )
        .exclude(estado_detalle__in=['devuelto', 'cancelado', 'rechazado', 'no_disponible'])
        .aggregate(total=models.Sum('cantidad_solicitada'))
        .get('total') or 0
    )

    estados_pedido = [
        ('pendiente', 'Pendientes', '#2d6cdf'),
        ('esperando entrega', 'Esperando entrega', '#26a7c6'),
        ('entregado', 'Prestados', '#57c271'),
        ('devuelto', 'Devueltos', '#e88a2a'),
        ('cancelado', 'Cancelados', '#cf3f5b'),
    ]
    prestamos_mes_actual = estado_conteos.get('entregado', 0)

    total_pedidos_mes = sum(estado_conteos.values())
    pie_segmentos = []
    pie_tramos = []
    acumulado = 0.0
    if total_pedidos_mes > 0:
        for clave, etiqueta, color in estados_pedido:
            cantidad = estado_conteos.get(clave, 0)
            if cantidad <= 0:
                continue
            porcentaje = round((cantidad / total_pedidos_mes) * 100, 1)
            inicio = acumulado
            acumulado += porcentaje
            pie_tramos.append(f'{color} {inicio:.2f}% {acumulado:.2f}%')
            pie_segmentos.append({
                'label': etiqueta,
                'cantidad': cantidad,
                'porcentaje': porcentaje,
                'color': color,
            })

    pie_conic = 'conic-gradient(' + (', '.join(pie_tramos) if pie_tramos else '#dce5de 0% 100%') + ')'

    tendencia = _construir_tendencia_mensual(ahora, meses=12)
    nombres_meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    return render(
        request,
        'inventario/dashboard/index.html',
        {
            'total_productos': total_productos,
            'total_stock_general': total_stock_general,
            'total_cantidad_general': total_cantidad_general,
            'pie_stock_cantidad_segmentos': pie_stock_cantidad_segmentos,
            'pie_stock_cantidad_conic': pie_stock_cantidad_conic,
            'productos_deficit': productos_deficit[:20],
            'total_productos_deficit': len(productos_deficit),
            'prestamos_activos': prestamos_activos,
            'pedidos_mes_actual': pedidos_mes_actual,
            'pedidos_pendientes_total': pedidos_pendientes_total,
            'hay_mas_pendientes': pedidos_pendientes_total > pendientes_preview_limit,
            'resumen_pendientes': resumen_pendientes,
            'usuarios_solicitud_manual': usuarios_solicitud_manual,
            'usuarios_solicitud_manual_modal': usuarios_solicitud_manual_modal,
            'usuarios_solicitud_manual_total': usuarios_solicitud_manual_total,
            'usuarios_documento_validacion': usuarios_documento_validacion,
            'usuarios_documento_validacion_total': usuarios_documento_validacion_total,
            'prestamos_mes_actual': prestamos_mes_actual,
            'productos_en_mora': productos_en_mora,
            'alertas_stock_bajo': alertas_stock_bajo,
            'alertas_cantidad_baja': alertas_cantidad_baja,
            'pie_segmentos': pie_segmentos,
            'pie_conic': pie_conic,
            'total_pedidos_mes': total_pedidos_mes,
            'tendencia': tendencia,
            'mes_reporte': ahora.strftime('%Y-%m'),
            'mes_actual_label': f'{nombres_meses[mes_actual - 1]} {anio_actual}',
        },
    )


def _construir_tendencia_mensual(ahora, meses=12):
    month_keys, resumen_mensual = _resumen_pedidos_mensual(ahora, meses=meses)
    nombres_meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    tendencia = []
    for y, m in month_keys:
        data = resumen_mensual.get((y, m), {})
        tendencia.append({
            'year': y,
            'month': m,
            'label': f'{nombres_meses[m - 1]} {str(y)[2:]}',
            'prestamos': data.get('entregado', 0),
            'pendientes': data.get('pendiente', 0),
            'devueltos': data.get('devuelto', 0),
            'cancelados': data.get('cancelado', 0),
        })
    return tendencia


def _resumen_pedidos_mensual(ahora, meses=12):
    def _mes_menos(base_year, base_month, minus_steps):
        total = base_year * 12 + (base_month - 1) - minus_steps
        return total // 12, (total % 12) + 1

    month_keys = [_mes_menos(ahora.year, ahora.month, offset) for offset in range(meses - 1, -1, -1)]
    month_keys = [key for key in month_keys if key[0] >= 2026]
    if not month_keys:
        month_keys = [(ahora.year, ahora.month)]
    month_set = set(month_keys)

    base = {
        'pendiente': 0,
        'esperando entrega': 0,
        'entregado': 0,
        'devuelto': 0,
        'cancelado': 0,
    }
    resumen = {key: dict(base) for key in month_keys}

    pedidos = Pedido.objects.exclude(fch_registro__isnull=True).only('estado', 'fch_registro')
    for pedido in pedidos:
        dt = pedido.fch_registro
        if not dt:
            continue
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        dt_local = timezone.localtime(dt)
        key = (dt_local.year, dt_local.month)
        if key not in month_set:
            continue
        estado = (pedido.estado or '').strip().lower()
        if estado == 'rechazado':
            estado = 'cancelado'
        if estado in resumen[key]:
            resumen[key][estado] += 1

    return month_keys, resumen


@login_required
def dashboard_tendencia_data(request):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    _auto_cancelar_pedidos_pendientes_vencidos()

    ahora = timezone.localtime()
    tendencia = _construir_tendencia_mensual(ahora, meses=12)
    return JsonResponse({
        'ok': True,
        'tendencia': tendencia,
        'updated_at': ahora.strftime('%Y-%m-%d %H:%M:%S'),
    })


@login_required
def dashboard_tendencia_detalle(request):
    if not _is_admin(request):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    _auto_cancelar_pedidos_pendientes_vencidos()

    try:
        year = int(request.GET.get('year', '0'))
        month = int(request.GET.get('month', '0'))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Parámetros inválidos.'}, status=400)

    serie = (request.GET.get('serie') or '').strip().lower()
    if not (1900 <= year <= 2200 and 1 <= month <= 12):
        return JsonResponse({'ok': False, 'error': 'Periodo inválido.'}, status=400)

    if serie not in {'pendientes', 'prestamos', 'devueltos', 'cancelados'}:
        return JsonResponse({'ok': False, 'error': 'Serie inválida.'}, status=400)

    estado_target = {
        'pendientes': {'pendiente'},
        'prestamos': {'entregado', 'vencido'},
        'devueltos': {'devuelto'},
        'cancelados': {'cancelado', 'rechazado'},
    }[serie]

    pedidos = (
        Pedido.objects
        .exclude(fch_registro__isnull=True)
        .select_related('id_usuario_fk__id_rol_fk')
        .only('id_pedido', 'estado', 'fch_registro', 'id_usuario_fk__nombre', 'id_usuario_fk__apellido', 'id_usuario_fk__correo', 'id_usuario_fk__id_rol_fk__nombre_rol')
    )

    pedidos_filtrados = []
    for pedido in pedidos:
        dt = pedido.fch_registro
        if not dt:
            continue
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        dt_local = timezone.localtime(dt)
        if dt_local.year != year or dt_local.month != month:
            continue
        if (pedido.estado or '').strip().lower() not in estado_target:
            continue
        pedidos_filtrados.append(pedido)

    ids_pedido = [p.id_pedido for p in pedidos_filtrados]
    logs = AuditoriaLog.objects.none()
    if ids_pedido:
        logs = (
            AuditoriaLog.objects
            .filter(entidad_id__in=[str(pid) for pid in ids_pedido], entidad__in=['pedido', 'prestamo'])
            .select_related('id_usuario_fk__id_rol_fk')
            .order_by('-fch_registro', '-id_log')
        )

    ultimo_log_por_pedido = {}
    for log in logs:
        try:
            pid = int(log.entidad_id)
        except (TypeError, ValueError):
            continue
        if pid not in ultimo_log_por_pedido:
            ultimo_log_por_pedido[pid] = log

    detalle = []
    for pedido in pedidos_filtrados:
        log = ultimo_log_por_pedido.get(pedido.id_pedido)
        if log and log.id_usuario_fk:
            nombre = f'{log.id_usuario_fk.nombre or ""} {log.id_usuario_fk.apellido or ""}'.strip()
            actor = nombre or log.id_usuario_fk.correo or f'Usuario #{log.id_usuario_fk_id}'
            rol = log.rol_usuario or (log.id_usuario_fk.id_rol_fk.nombre_rol if getattr(log.id_usuario_fk, 'id_rol_fk', None) else '-')
            hora = timezone.localtime(log.fch_registro).strftime('%d/%m/%Y %H:%M') if log.fch_registro else '-'
            descripcion = log.descripcion or '-'
        else:
            usuario = pedido.id_usuario_fk
            nombre = f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() if usuario else ''
            actor = nombre or (usuario.correo if usuario else '-')
            rol = usuario.id_rol_fk.nombre_rol if usuario and getattr(usuario, 'id_rol_fk', None) else '-'
            hora = timezone.localtime(pedido.fch_registro).strftime('%d/%m/%Y %H:%M') if pedido.fch_registro else '-'
            descripcion = f'Pedido #{pedido.id_pedido} en estado {pedido.estado or "-"}.'

        detalle.append({
            'pedido_id': pedido.id_pedido,
            'usuario': actor,
            'rol': rol,
            'hora': hora,
            'descripcion': descripcion,
            'estado': pedido.estado,
        })

    return JsonResponse({
        'ok': True,
        'serie': serie,
        'year': year,
        'month': month,
        'total': len(detalle),
        'items': detalle,
    })


def inventario_panel(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    q = (request.GET.get('q') or '').strip()
    cat_id = (request.GET.get('categoria') or '').strip()
    bajo_stock = (request.GET.get('bajo_stock') or '').strip() == '1'
    disp_qs = Disponibilidad.objects.filter(id_prod_fk=OuterRef('pk')).order_by('-id_disp')

    productos_qs = (
        Producto.objects
        .select_related('id_cat_fk')
        .prefetch_related('fotos', 'subcategorias')
        .annotate(
            stock_actual=Subquery(disp_qs.values('stock')[:1]),
            cantidad_actual=Subquery(disp_qs.values('cantidad')[:1]),
        )
    )

    if cat_id.isdigit():
        productos_qs = productos_qs.filter(id_cat_fk_id=int(cat_id))

    if q:
        productos_qs = productos_qs.filter(
            models.Q(nombre_producto__icontains=q)
            | models.Q(id_cat_fk__nombre_catalogo__icontains=q)
            | models.Q(codigo_producto__icontains=q)
            | models.Q(id_cat_fk__codigo_macro__icontains=q)
            | models.Q(subcategorias__nombre_subcategoria__icontains=q)
            | models.Q(subcategorias__codigo_clasificacion__icontains=q)
            | models.Q(subcategorias__subcategoria_padre__nombre_subcategoria__icontains=q)
            | models.Q(subcategorias__subcategoria_padre__codigo_clasificacion__icontains=q)
        ).distinct()

    if bajo_stock:
        productos_qs = productos_qs.filter(stock_actual__lte=5)

    productos = list(productos_qs.order_by('-fch_registro', '-id_prod'))

    catalogo_ids_en_resultado = sorted({prod.id_cat_fk_id for prod in productos if prod.id_cat_fk_id})
    subcat_map = {}
    if catalogo_ids_en_resultado:
        for subcat in (
            Subcategoria.objects
            .filter(id_cat_fk_id__in=catalogo_ids_en_resultado)
            .values('id_subcat', 'nombre_subcategoria', 'subcategoria_padre_id', 'id_cat_fk_id')
        ):
            subcat_map[subcat['id_subcat']] = subcat

    def _ruta_subcategoria(subcat_id):
        partes = []
        visited = set()
        current = subcat_map.get(subcat_id)
        while current and current['id_subcat'] not in visited:
            visited.add(current['id_subcat'])
            partes.append(current['nombre_subcategoria'])
            current = subcat_map.get(current['subcategoria_padre_id'])
        return ' / '.join(reversed(partes))

    for prod in productos:
        trazas = []
        for subcat in prod.subcategorias.all():
            ruta = _ruta_subcategoria(subcat.id_subcat)
            trazas.append({
                'id': subcat.id_subcat,
                'nombre': subcat.nombre_subcategoria,
                'ruta': ruta or subcat.nombre_subcategoria,
            })
        trazas.sort(key=lambda item: item['ruta'].lower())
        prod.subcategoria_trazas = trazas
        prod.subcategoria_ruta_resumen = ' | '.join(item['ruta'] for item in trazas) if trazas else 'Raíz del catálogo'

    catalogos = (
        Catalogo.objects.annotate(
            total_productos=models.Count('producto')
        ).order_by('nombre_catalogo')
    )

    productos_por_catalogo = {}
    for prod in productos:
        productos_por_catalogo.setdefault(prod.id_cat_fk_id, []).append(prod)

    secciones_catalogo = [
        {
            'catalogo': cat,
            'productos': productos_por_catalogo.get(cat.id_cat, []),
        }
        for cat in catalogos
        if productos_por_catalogo.get(cat.id_cat)
    ]

    return render(
        request,
        'inventario/dashboard/inventario_panel.html',
        {
            'q': q,
            'categoria_activa': cat_id,
            'bajo_stock': bajo_stock,
            'catalogos': catalogos,
            'productos': productos,
            'secciones_catalogo': secciones_catalogo,
        },
    )


def _mes_reporte_desde_request(request):
    mes_param = (request.GET.get('mes') or '').strip()
    ahora = timezone.localtime()
    if len(mes_param) == 7 and mes_param[4] == '-':
        try:
            anio = int(mes_param[:4])
            mes = int(mes_param[5:7])
            if 1 <= mes <= 12:
                return anio, mes
        except (TypeError, ValueError):
            pass
    return ahora.year, ahora.month


def _obtener_prestamos_mes(anio, mes):
    pedidos_qs = (
        Pedido.objects
        .exclude(fch_registro__isnull=True)
        .select_related('id_usuario_fk__id_rol_fk')
        .prefetch_related('detalles')
        .order_by('-fch_registro', '-id_pedido')
    )

    pedidos_filtrados = []
    for pedido in pedidos_qs:
        dt = pedido.fch_registro
        if not dt:
            continue
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        dt_local = timezone.localtime(dt)
        if dt_local.year == anio and dt_local.month == mes:
            pedidos_filtrados.append(pedido)

    return pedidos_filtrados


def _categoria_pedido_reporte(estado):
    estado_limpio = (estado or '').strip().lower()
    if estado_limpio in ['entregado', 'vencido', 'devuelto']:
        return 'REALIZADO'
    if estado_limpio in ['cancelado', 'rechazado']:
        return 'CANCELADO'
    return 'EN PROCESO'


def _resumen_productos_pedido(pedido, max_items=None, multiline=False):
    detalles = list(getattr(pedido, 'detalles', []).all()) if hasattr(getattr(pedido, 'detalles', None), 'all') else []
    if not detalles:
        return 'Sin detalle'

    nombres = []
    for idx, det in enumerate(detalles, start=1):
        nombre = (det.nombre_producto or '').strip() or f'Producto {det.id_prod_fk_id or "-"}'
        cantidad = int(det.cantidad_solicitada or 0)
        estado = (det.estado_detalle or '').strip()
        sufijo_estado = f' [{estado}]' if estado else ''
        nombres.append(f'{idx}. {nombre} x{cantidad}{sufijo_estado}')

    if max_items is not None and max_items >= 0:
        nombres = nombres[:max_items]
        if len(detalles) > max_items:
            nombres.append(f'+{len(detalles) - max_items} más')

    separador = '\n' if multiline else ' | '
    return separador.join(nombres)


def _build_pdf_text_report(lines):
    def _escape(texto):
        return (texto or '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')

    contenido = ['BT', '/F1 10 Tf', '14 TL', '40 800 Td']
    for idx, linea in enumerate(lines[:55]):
        if idx == 0:
            contenido.append(f'({_escape(linea)}) Tj')
        else:
            contenido.append('T*')
            contenido.append(f'({_escape(linea)}) Tj')
    contenido.append('ET')

    stream_data = '\n'.join(contenido).encode('latin-1', errors='replace')
    obj1 = b'1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n'
    obj2 = b'2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n'
    obj3 = (
        b'3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] '
        b'/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>endobj\n'
    )
    obj4 = b'4 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n'
    obj5 = b'5 0 obj<< /Length ' + str(len(stream_data)).encode('ascii') + b' >>stream\n' + stream_data + b'\nendstream endobj\n'

    objects = [obj1, obj2, obj3, obj4, obj5]
    pdf = b'%PDF-1.4\n'
    offsets = []
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj

    xref_start = len(pdf)
    pdf += b'xref\n0 6\n0000000000 65535 f \n'
    for off in offsets:
        pdf += f'{off:010d} 00000 n \n'.encode('ascii')
    pdf += b'trailer<< /Size 6 /Root 1 0 R >>\nstartxref\n'
    pdf += str(xref_start).encode('ascii') + b'\n%%EOF'
    return pdf


@login_required
def reporte_prestamos_excel(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    anio, mes = _mes_reporte_desde_request(request)
    prestamos = _obtener_prestamos_mes(anio, mes)
    secciones = {
        'REALIZADO': [],
        'CANCELADO': [],
        'EN PROCESO': [],
    }
    for pedido in prestamos:
        secciones[_categoria_pedido_reporte(pedido.estado)].append(pedido)

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from PIL import Image as PILImage
    except Exception:
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="reporte_prestamos_{anio}_{mes:02d}.csv"'
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow([f'Reporte mensual de pedidos - {anio}-{mes:02d}'])
        writer.writerow(['Generado', timezone.localtime().strftime('%d/%m/%Y %H:%M')])
        writer.writerow([])
        writer.writerow(['Resumen'])
        writer.writerow(['Realizados', len(secciones['REALIZADO'])])
        writer.writerow(['Cancelados', len(secciones['CANCELADO'])])
        writer.writerow(['En proceso', len(secciones['EN PROCESO'])])
        writer.writerow(['Total', len(prestamos)])
        writer.writerow([])

        encabezado = [
            'Categoria', 'Pedido', 'Fecha registro', 'Estado pedido', 'Usuario', 'Rol',
            'Total productos', 'Total unidades', 'Item', 'Producto', 'Cantidad solicitada',
            'Estado detalle', 'Area', 'Fecha devolucion'
        ]

        for nombre_seccion in ['REALIZADO', 'CANCELADO', 'EN PROCESO']:
            items = secciones[nombre_seccion]
            writer.writerow([f'SECCION: {nombre_seccion} ({len(items)})'])
            writer.writerow(encabezado)

            for pedido in items:
                usuario = pedido.id_usuario_fk
                nombre_usuario = ((usuario.nombre or '') + ' ' + (usuario.apellido or '')).strip() or (usuario.correo or '')
                rol = usuario.id_rol_fk.nombre_rol if usuario.id_rol_fk else 'sin rol'
                fecha_registro = timezone.localtime(pedido.fch_registro).strftime('%d/%m/%Y %H:%M') if pedido.fch_registro else '-'
                fecha_devolucion = timezone.localtime(pedido.fecha_devolucion).strftime('%d/%m/%Y %H:%M') if pedido.fecha_devolucion else '-'
                area = (pedido.area_ubicacion or '').replace('\n', ' ').replace(';', ',')

                detalles = list(pedido.detalles.all()) if hasattr(pedido, 'detalles') else []
                if not detalles:
                    writer.writerow([
                        nombre_seccion,
                        pedido.id_pedido,
                        fecha_registro,
                        pedido.estado,
                        nombre_usuario,
                        rol,
                        pedido.total_productos,
                        pedido.total_unidades,
                        '-',
                        'Sin detalle',
                        0,
                        '-',
                        area,
                        fecha_devolucion,
                    ])
                    continue

                for idx, det in enumerate(detalles, start=1):
                    producto = ((det.nombre_producto or '').strip() or f'Producto {det.id_prod_fk_id or "-"}').replace(';', ',')
                    writer.writerow([
                        nombre_seccion,
                        pedido.id_pedido,
                        fecha_registro,
                        pedido.estado,
                        nombre_usuario,
                        rol,
                        pedido.total_productos,
                        pedido.total_unidades,
                        idx,
                        producto,
                        int(det.cantidad_solicitada or 0),
                        (det.estado_detalle or '-').replace(';', ','),
                        area,
                        fecha_devolucion,
                    ])
            writer.writerow([])
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte pedidos'

    widths = [15, 10, 19, 16, 22, 12, 14, 14, 8, 34, 16, 16, 20, 19]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    thin = Side(style='thin', color='CFE4D2')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_title = PatternFill('solid', fgColor='DFF4E5')
    fill_section = PatternFill('solid', fgColor='ECF8F0')
    fill_head = PatternFill('solid', fgColor='E5F4E9')
    fill_alt = PatternFill('solid', fgColor='F7FCF8')

    font_title = Font(name='Calibri', size=16, bold=True, color='1D6B3A')
    font_sub = Font(name='Calibri', size=11, bold=True, color='2A5E3F')
    font_head = Font(name='Calibri', size=10, bold=True, color='235438')
    font_cell = Font(name='Calibri', size=10, color='1F4330')

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value=f'Reporte mensual de pedidos - {anio}-{mes:02d}')
    ws.cell(row=row, column=1).font = font_title
    ws.cell(row=row, column=1).fill = fill_title
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 28

    row += 1
    ws.cell(row=row, column=1, value='Generado')
    ws.cell(row=row, column=2, value=timezone.localtime().strftime('%d/%m/%Y %H:%M'))
    ws.cell(row=row, column=1).font = font_sub
    ws.cell(row=row, column=2).font = font_sub

    logo_candidates = [
        os.path.join(settings.BASE_DIR, 'logoSena.png'),
        os.path.join(settings.BASE_DIR, 'inventario', 'static', 'inventario', 'img', 'logoSena.png'),
    ]
    for logo_path in logo_candidates:
        if not os.path.exists(logo_path):
            continue
        try:
            logo = XLImage(logo_path)
            logo.width = 64
            logo.height = 64
            ws.add_image(logo, 'N1')
            break
        except Exception:
            continue

    row += 2
    ws.cell(row=row, column=1, value='Resumen').font = font_sub
    resumen_items = [
        ('Realizados', len(secciones['REALIZADO'])),
        ('Cancelados', len(secciones['CANCELADO'])),
        ('En proceso', len(secciones['EN PROCESO'])),
        ('Total', len(prestamos)),
    ]
    for nombre, cantidad in resumen_items:
        row += 1
        ws.cell(row=row, column=1, value=nombre).font = font_cell
        ws.cell(row=row, column=2, value=cantidad).font = font_cell

    row += 2
    encabezado = [
        'Categoria', 'Pedido', 'Fecha registro', 'Estado pedido', 'Usuario', 'Rol',
        'Total productos', 'Total unidades', 'Item', 'Producto', 'Cantidad solicitada',
        'Estado detalle', 'Area', 'Fecha devolucion'
    ]

    for nombre_seccion in ['REALIZADO', 'CANCELADO', 'EN PROCESO']:
        items = secciones[nombre_seccion]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        ws.cell(row=row, column=1, value=f'SECCION: {nombre_seccion} ({len(items)})')
        ws.cell(row=row, column=1).font = font_sub
        ws.cell(row=row, column=1).fill = fill_section
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

        for col, name in enumerate(encabezado, start=1):
            c = ws.cell(row=row, column=col, value=name)
            c.font = font_head
            c.fill = fill_head
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

        data_start = row
        for pedido in items:
            usuario = pedido.id_usuario_fk
            nombre_usuario = ((usuario.nombre or '') + ' ' + (usuario.apellido or '')).strip() or (usuario.correo or '')
            rol = usuario.id_rol_fk.nombre_rol if usuario.id_rol_fk else 'sin rol'
            fecha_registro = timezone.localtime(pedido.fch_registro).strftime('%d/%m/%Y %H:%M') if pedido.fch_registro else '-'
            fecha_devolucion = timezone.localtime(pedido.fecha_devolucion).strftime('%d/%m/%Y %H:%M') if pedido.fecha_devolucion else '-'
            area = (pedido.area_ubicacion or '').replace('\n', ' ')
            detalles = list(pedido.detalles.all()) if hasattr(pedido, 'detalles') else []

            if not detalles:
                detalles = [None]

            for idx, det in enumerate(detalles, start=1):
                producto = 'Sin detalle'
                cantidad_det = 0
                estado_det = '-'
                if det is not None:
                    producto = (det.nombre_producto or '').strip() or f'Producto {det.id_prod_fk_id or "-"}'
                    cantidad_det = int(det.cantidad_solicitada or 0)
                    estado_det = det.estado_detalle or '-'

                row_values = [
                    nombre_seccion,
                    pedido.id_pedido,
                    fecha_registro,
                    pedido.estado,
                    nombre_usuario,
                    rol,
                    pedido.total_productos,
                    pedido.total_unidades,
                    idx if det is not None else '-',
                    producto,
                    cantidad_det,
                    estado_det,
                    area,
                    fecha_devolucion,
                ]

                for col, value in enumerate(row_values, start=1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_cell
                    c.border = border
                    c.alignment = Alignment(vertical='top', wrap_text=True)
                if (row - data_start) % 2 == 1:
                    for col in range(1, 15):
                        ws.cell(row=row, column=col).fill = fill_alt
                row += 1

        row += 1

    ws.freeze_panes = 'A12'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_prestamos_{anio}_{mes:02d}.xlsx"'
    return response


@login_required
def exportar_inventario_excel(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    from django.db.models import Subquery, TextField, Value
    from django.db.models.functions import Coalesce

    disp_qs = Disponibilidad.objects.filter(id_prod_fk=OuterRef('pk')).order_by('-id_disp')
    productos = list(
        Producto.objects
        .select_related('id_cat_fk')
        .prefetch_related('subcategorias', 'fotos')
        .annotate(
            stock_actual=Coalesce(Subquery(disp_qs.values('stock')[:1]), 0),
            cantidad_actual=Coalesce(Subquery(disp_qs.values('cantidad')[:1]), 0),
            descr_dispo_actual=Coalesce(
                Subquery(disp_qs.values('descr_dispo')[:1]),
                Value('', output_field=TextField()),
                output_field=TextField(),
            ),
        )
        .order_by('id_cat_fk__nombre_catalogo', 'nombre_producto', 'id_prod')
    )

    if not productos:
        messages.info(request, 'No hay productos para exportar en este momento.')
        return redirect('dashboard')

    catalogos = defaultdict(list)
    for producto in productos:
        catalogos[producto.id_cat_fk].append(producto)

    generated_at = timezone.localtime().strftime('%d/%m/%Y %H:%M')

    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from PIL import Image as PILImage
    except Exception:
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="inventario_completo.csv"'
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['Inventario completo por categorias'])
        writer.writerow(['Generado', generated_at])
        writer.writerow([])
        writer.writerow([
            'Categoria', 'Producto', 'Descripcion', 'Unidad de medida', 'Ubicacion', 'Tipo de bien',
            'Numero de placa', 'Cuentadante', 'Subcategorias', 'Stock actual', 'Cantidad actual',
            'Detalle disponibilidad', 'Imagen principal', 'Imagenes secundarias'
        ])

        for catalogo, items in catalogos.items():
            writer.writerow([f'CATEGORIA: {catalogo.nombre_catalogo}', '', '', '', '', '', '', '', '', '', '', '', '', ''])
            for producto in items:
                sec_names = ', '.join(
                    os.path.basename(getattr(foto.foto, 'name', '') or '')
                    for foto in producto.fotos.all()
                )
                writer.writerow([
                    catalogo.nombre_catalogo,
                    producto.nombre_producto or f'Producto {producto.id_prod}',
                    (producto.descripcion or '').replace(';', ','),
                    producto.get_unidad_medida_display(),
                    producto.ubicacion or '',
                    producto.get_tipo_bien_display(),
                    producto.numero_placa or '',
                    producto.cuentadante or '',
                    ' | '.join(sub.ruta_completa for sub in producto.subcategorias.all()),
                    int(producto.stock_actual or 0),
                    int(producto.cantidad_actual or 0),
                    (producto.descr_dispo_actual or '').replace(';', ','),
                    producto.fot_prod.url if producto.fot_prod else '',
                    sec_names,
                ])
            writer.writerow([])
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.sheet_view.showGridLines = False

    widths = {
        'A': 24,
        'B': 16,
        'C': 28,
        'D': 50,
        'E': 15,
        'F': 22,
        'G': 15,
        'H': 16,
        'I': 18,
        'J': 24,
        'K': 10,
        'L': 10,
        'M': 28,
        'N': 34,
        'O': 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style='thin', color='C9DFC9')
    medium = Side(style='medium', color='7CAF7D')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_strong = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_title = PatternFill('solid', fgColor='D9F2D9')
    fill_meta = PatternFill('solid', fgColor='EEF9EE')
    fill_category = PatternFill('solid', fgColor='DFF1DF')
    fill_head = PatternFill('solid', fgColor='CFEBCF')
    fill_alt = PatternFill('solid', fgColor='F7FCF7')
    fill_image_empty = PatternFill('solid', fgColor='F4F8F4')
    fill_stock = PatternFill('solid', fgColor='E8F7E8')

    font_title = Font(name='Calibri', size=18, bold=True, color='1E5C30')
    font_sub = Font(name='Calibri', size=11, bold=True, color='2A5E3F')
    font_meta = Font(name='Calibri', size=10, color='345B3D')
    font_head = Font(name='Calibri', size=10, bold=True, color='1F4A2B')
    font_cell = Font(name='Calibri', size=10, color='1F4330')
    font_placeholder = Font(name='Calibri', size=10, italic=True, color='7E8C7F')

    total_categorias = len(catalogos)
    total_productos = len(productos)
    total_con_foto = sum(1 for producto in productos if producto.fot_prod)
    imagenes_insertadas = 0

    def _excel_image_from_path(image_path):
        normalized = io.BytesIO()
        with PILImage.open(image_path) as pil_image:
            pil_image = pil_image.convert('RGBA')
            background = PILImage.new('RGBA', pil_image.size, (255, 255, 255, 255))
            background.alpha_composite(pil_image)
            background.convert('RGB').save(normalized, format='PNG')
        normalized.seek(0)
        excel_image = XLImage(normalized)
        excel_image._normalized_buffer = normalized
        return excel_image

    row = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=11)
    title_cell = ws.cell(row=1, column=1, value='Inventario completo por categorias')
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.border = border_strong
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    for merged_row in range(1, 3):
        for merged_col in range(1, 12):
            ws.cell(row=merged_row, column=merged_col).fill = fill_title
            ws.cell(row=merged_row, column=merged_col).border = border_strong

    logo_candidates = [
        os.path.join(settings.BASE_DIR, 'logoSena.png'),
        os.path.join(settings.BASE_DIR, 'inventario', 'static', 'inventario', 'img', 'logoSena.png'),
    ]
    for logo_path in logo_candidates:
        if not os.path.exists(logo_path):
            continue
        try:
            logo = _excel_image_from_path(logo_path)
            logo.width = 120
            logo.height = 78
            ws.add_image(logo, 'M1')
            break
        except Exception:
            continue

    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 26

    meta_rows = [
        ('Generado', generated_at),
        ('Categorias exportadas', total_categorias),
        ('Productos exportados', total_productos),
        ('Productos con foto registrada', total_con_foto),
        ('Imagenes embebidas en Excel', ''),
    ]
    meta_start = 4
    for idx, (label, value) in enumerate(meta_rows, start=meta_start):
        left = ws.cell(row=idx, column=1, value=label)
        right = ws.cell(row=idx, column=2, value=value)
        left.font = font_sub
        right.font = font_meta
        left.fill = fill_meta
        right.fill = fill_meta
        left.border = border
        right.border = border
        left.alignment = Alignment(vertical='center')
        right.alignment = Alignment(vertical='center')

    row = 10
    headers = [
        'Categoria', 'Imagen', 'Producto', 'Descripcion', 'Unidad', 'Ubicacion',
        'Tipo', 'Placa', 'Cuentadante', 'Subcategorias', 'Stock', 'Cantidad', 'Detalle',
        'Archivo imagen', 'Archivos secundarios'
    ]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = font_head
        cell.fill = fill_head
        cell.border = border_strong
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.auto_filter.ref = f'A{row}:O{row}'

    data_start_row = row + 1
    for catalogo, items in catalogos.items():
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        category_cell = ws.cell(row=row, column=1, value=f'Categoria: {catalogo.nombre_catalogo} ({len(items)} productos)')
        category_cell.font = font_sub
        category_cell.fill = fill_category
        category_cell.border = border_strong
        category_cell.alignment = Alignment(horizontal='left', vertical='center')
        for merged_col in range(1, 16):
            ws.cell(row=row, column=merged_col).fill = fill_category
            ws.cell(row=row, column=merged_col).border = border_strong
        ws.row_dimensions[row].height = 24

        for producto in items:
            row += 1
            ws.row_dimensions[row].height = 92

            subcategorias = ' | '.join(sub.ruta_completa for sub in producto.subcategorias.all())
            image_name = ''
            if producto.fot_prod:
                image_name = os.path.basename(getattr(producto.fot_prod, 'name', '') or '')
            secondary_names = ', '.join(
                os.path.basename(getattr(foto.foto, 'name', '') or '')
                for foto in producto.fotos.all()
            )

            values = [
                catalogo.nombre_catalogo,
                '',
                producto.nombre_producto or f'Producto {producto.id_prod}',
                producto.descripcion or '',
                producto.get_unidad_medida_display(),
                producto.ubicacion or '',
                producto.get_tipo_bien_display(),
                producto.numero_placa or '',
                producto.cuentadante or '',
                subcategorias,
                int(producto.stock_actual or 0),
                int(producto.cantidad_actual or 0),
                producto.descr_dispo_actual or '',
                image_name,
                secondary_names,
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = font_cell
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            if (row - data_start_row) % 2 == 1:
                for col in range(1, 16):
                    ws.cell(row=row, column=col).fill = fill_alt

            for stock_col in (11, 12):
                ws.cell(row=row, column=stock_col).fill = fill_stock
                ws.cell(row=row, column=stock_col).alignment = Alignment(horizontal='center', vertical='center')

            image_cell = ws.cell(row=row, column=2)
            image_cell.alignment = Alignment(horizontal='center', vertical='center')
            image_cell.border = border
            image_cell.fill = fill_image_empty

            if producto.fot_prod:
                try:
                    image_path = producto.fot_prod.path
                    if not os.path.exists(image_path):
                        raise FileNotFoundError(image_path)

                    image = _excel_image_from_path(image_path)

                    image.width = 78
                    image.height = 78
                    ws.add_image(image, f'B{row}')
                    imagenes_insertadas += 1
                    image_cell.value = None
                except Exception:
                    image_cell.value = 'Sin vista previa'
                    image_cell.font = font_placeholder
            else:
                image_cell.value = ''

    ws.cell(row=8, column=2, value=imagenes_insertadas).font = font_meta
    ws.cell(row=8, column=2).fill = fill_meta
    ws.cell(row=8, column=2).border = border
    ws.cell(row=8, column=2).alignment = Alignment(vertical='center')

    sec_ws = wb.create_sheet(title='Imagenes Secundarias')
    sec_ws.sheet_view.showGridLines = False
    sec_widths = {
        'A': 14,
        'B': 24,
        'C': 32,
        'D': 10,
        'E': 32,
        'F': 18,
    }
    for col, width in sec_widths.items():
        sec_ws.column_dimensions[col].width = width

    sec_headers = ['ID Producto', 'Categoria', 'Producto', 'Orden', 'Archivo secundario', 'Imagen']
    sec_row = 1
    for col, name in enumerate(sec_headers, start=1):
        c = sec_ws.cell(row=sec_row, column=col, value=name)
        c.font = font_head
        c.fill = fill_head
        c.border = border_strong
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sec_row += 1
    for producto in productos:
        fotos_sec = list(producto.fotos.all())
        if not fotos_sec:
            continue
        for foto in fotos_sec:
            sec_ws.row_dimensions[sec_row].height = 88
            sec_values = [
                producto.id_prod,
                producto.id_cat_fk.nombre_catalogo if producto.id_cat_fk else '',
                producto.nombre_producto or '',
                foto.orden,
                os.path.basename(getattr(foto.foto, 'name', '') or ''),
                '',
            ]
            for col, value in enumerate(sec_values, start=1):
                c = sec_ws.cell(row=sec_row, column=col, value=value)
                c.font = font_cell
                c.border = border
                c.alignment = Alignment(vertical='center', wrap_text=True)

            img_cell = sec_ws.cell(row=sec_row, column=6)
            img_cell.fill = fill_image_empty
            img_cell.border = border
            img_cell.alignment = Alignment(horizontal='center', vertical='center')

            try:
                sec_path = foto.foto.path
                if os.path.exists(sec_path):
                    sec_image = _excel_image_from_path(sec_path)
                    sec_image.width = 78
                    sec_image.height = 78
                    sec_ws.add_image(sec_image, f'F{sec_row}')
            except Exception:
                img_cell.value = 'Sin vista previa'
                img_cell.font = font_placeholder

            sec_row += 1

    sec_ws.freeze_panes = 'A2'

    ws.freeze_panes = 'A11'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="inventario_completo_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    return response


def _header_key(value):
    text = (value or '').strip().lower()
    text = text.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    return ' '.join(text.split())


def _parse_excel_subcategorias(raw_text):
    """Parsea subcategorías desde Excel soportando |, coma, ; y salto de línea."""
    text = (raw_text or '').replace('\n', '|').replace(';', '|').replace(',', '|')
    values = []
    seen = set()
    for part in text.split('|'):
        item = (part or '').strip()
        if not item:
            continue
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        values.append(item)
    return values


def _find_header_row(sheet, required_keys, max_scan_rows=60):
    for row_idx in range(1, max_scan_rows + 1):
        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
        normalized = {_header_key(str(v)) for v in row_values if v is not None and str(v).strip()}
        if required_keys.issubset(normalized):
            return row_idx
    return None


def _extract_images_by_row(sheet):
    row_map = defaultdict(list)
    for image in getattr(sheet, '_images', []):
        try:
            anchor = getattr(image, 'anchor', None)
            if not anchor or not hasattr(anchor, '_from'):
                continue
            row_1_based = int(anchor._from.row) + 1
            col_1_based = int(anchor._from.col) + 1
            raw = image._data()
            ext = (getattr(image, 'format', None) or 'png').lower()
            if ext == 'jpeg':
                ext = 'jpg'
            row_map[row_1_based].append((col_1_based, raw, ext))
        except Exception:
            continue
    return row_map


def _image_to_data_uri(raw_bytes, ext='png'):
    try:
        from PIL import Image as PILImage

        src = io.BytesIO(raw_bytes)
        dst = io.BytesIO()
        with PILImage.open(src) as img:
            img = img.convert('RGB')
            img.thumbnail((140, 140))
            img.save(dst, format='JPEG', quality=80)
        encoded = base64.b64encode(dst.getvalue()).decode('ascii')
        return f'data:image/jpeg;base64,{encoded}'
    except Exception:
        try:
            encoded = base64.b64encode(raw_bytes).decode('ascii')
            safe_ext = (ext or 'png').lower()
            if safe_ext == 'jpg':
                safe_ext = 'jpeg'
            return f'data:image/{safe_ext};base64,{encoded}'
        except Exception:
            return ''


def _build_preview_from_excel(temp_path, search=''):
    try:
        from openpyxl import load_workbook
    except Exception:
        return {'groups': [], 'rows': [], 'error': 'openpyxl no está disponible para previsualizar.'}

    wb = load_workbook(temp_path, data_only=True)
    ws = wb['Inventario'] if 'Inventario' in wb.sheetnames else wb.active

    required = {'categoria', 'producto'}
    header_row = _find_header_row(ws, required)
    if not header_row:
        return {'groups': [], 'rows': [], 'error': 'No se encontró la tabla de inventario en el Excel.'}

    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        text = str(value).strip() if value is not None else ''
        if not text:
            continue
        header_to_col[_header_key(text)] = col

    image_col = header_to_col.get('imagen')
    main_images_by_row = _extract_images_by_row(ws)

    secondary_images_map = defaultdict(list)
    if 'Imagenes Secundarias' in wb.sheetnames:
        sec_ws = wb['Imagenes Secundarias']
        sec_required = {'categoria', 'producto'}
        sec_header_row = _find_header_row(sec_ws, sec_required, max_scan_rows=20)
        if sec_header_row:
            sec_header_to_col = {}
            for col in range(1, sec_ws.max_column + 1):
                value = sec_ws.cell(row=sec_header_row, column=col).value
                if value is None:
                    continue
                sec_header_to_col[_header_key(str(value))] = col

            sec_img_col = sec_header_to_col.get('imagen')
            sec_images_by_row = _extract_images_by_row(sec_ws)
            for row_idx in range(sec_header_row + 1, sec_ws.max_row + 1):
                categoria = str(sec_ws.cell(row=row_idx, column=sec_header_to_col.get('categoria', 1)).value or '').strip()
                producto = str(sec_ws.cell(row=row_idx, column=sec_header_to_col.get('producto', 1)).value or '').strip()
                if not categoria or not producto:
                    continue
                key = (categoria.lower(), producto.lower())
                for col_1_based, raw, ext in sec_images_by_row.get(row_idx, []):
                    if sec_img_col and col_1_based != sec_img_col:
                        continue
                    data_uri = _image_to_data_uri(raw, ext)
                    if data_uri:
                        secondary_images_map[key].append(data_uri)

    rows = []
    groups_map = defaultdict(list)
    groups_order = []
    q = (search or '').strip().lower()

    def _cell_text(row_idx, key):
        col = header_to_col.get(key)
        if not col:
            return ''
        value = ws.cell(row=row_idx, column=col).value
        return str(value).strip() if value is not None else ''

    for row_idx in range(header_row + 1, ws.max_row + 1):
        categoria_text = _cell_text(row_idx, 'categoria')
        producto_text = _cell_text(row_idx, 'producto')

        if not producto_text or categoria_text.lower().startswith('categoria:'):
            continue

        main_image = ''
        if image_col and row_idx in main_images_by_row:
            for col_1_based, raw, ext in main_images_by_row[row_idx]:
                if col_1_based != image_col:
                    continue
                main_image = _image_to_data_uri(raw, ext)
                if main_image:
                    break

        key = (categoria_text.lower(), producto_text.lower())
        secondary_imgs = secondary_images_map.get(key, [])

        row_data = {
            'Categoria': categoria_text,
            'Producto': producto_text,
            'Descripcion': _cell_text(row_idx, 'descripcion'),
            'Unidad': _cell_text(row_idx, 'unidad'),
            'Ubicacion': _cell_text(row_idx, 'ubicacion'),
            'Tipo': _cell_text(row_idx, 'tipo'),
            'Subcategorias': _cell_text(row_idx, 'subcategorias'),
            'Stock': _cell_text(row_idx, 'stock'),
            'Cantidad': _cell_text(row_idx, 'cantidad'),
            'MainImage': main_image,
            'SecondaryImages': secondary_imgs[:4],
            'SecondaryCount': len(secondary_imgs),
        }

        searchable = ' '.join([
            row_data['Categoria'],
            row_data['Producto'],
            row_data['Descripcion'],
            row_data['Unidad'],
            row_data['Ubicacion'],
            row_data['Tipo'],
            row_data['Subcategorias'],
            row_data['Stock'],
            row_data['Cantidad'],
        ]).lower()
        if q and q not in searchable:
            continue

        rows.append(row_data)
        if categoria_text not in groups_map:
            groups_order.append(categoria_text)
        groups_map[categoria_text].append(row_data)

        if len(rows) >= 300:
            break

    groups = [
        {
            'categoria': categoria,
            'items': groups_map.get(categoria, []),
        }
        for categoria in groups_order
    ]

    return {
        'groups': groups,
        'rows': rows,
        'error': '',
    }


def _save_temp_excel(uploaded_file):
    uploads_dir = os.path.join(settings.MEDIA_ROOT, 'tmp', 'imports_excel')
    os.makedirs(uploads_dir, exist_ok=True)
    token = secrets.token_urlsafe(16)
    ext = os.path.splitext(uploaded_file.name or '')[1].lower() or '.xlsx'
    file_name = f'inventario_import_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}_{token}{ext}'
    temp_path = os.path.join(uploads_dir, file_name)

    with open(temp_path, 'wb') as dst:
        for chunk in uploaded_file.chunks():
            dst.write(chunk)

    return temp_path


def _importar_excel_inventario(temp_path, request):
    from django.core.files.base import ContentFile
    from django.db import transaction
    from django.db.models import Q

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return {'ok': False, 'message': f'No se pudo importar: {exc}'}

    wb = load_workbook(temp_path, data_only=True)
    ws = wb['Inventario'] if 'Inventario' in wb.sheetnames else wb.active

    required = {'categoria', 'producto'}
    header_row = _find_header_row(ws, required)
    if not header_row:
        return {'ok': False, 'message': 'No se encontró una tabla válida de inventario en el archivo.'}

    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        key = _header_key(str(value))
        if key:
            header_to_col[key] = col

    image_col = header_to_col.get('imagen')
    main_images_by_row = _extract_images_by_row(ws)

    created = 0
    updated = 0
    imported_main_images = 0
    imported_secondary_images = 0
    total_rows = 0
    errors = []
    product_key_map = {}

    def _cell_value(row_idx, key):
        col = header_to_col.get(key)
        if not col:
            return ''
        value = ws.cell(row=row_idx, column=col).value
        return str(value).strip() if value is not None else ''

    with transaction.atomic():
        for row_idx in range(header_row + 1, ws.max_row + 1):
            categoria = _cell_value(row_idx, 'categoria')
            nombre = _cell_value(row_idx, 'producto')

            if not nombre or categoria.lower().startswith('categoria:'):
                continue

            total_rows += 1

            catalogo, _ = Catalogo.objects.get_or_create(
                nombre_catalogo=categoria,
                defaults={'fch_registro': timezone.now(), 'fch_ult_act': timezone.now()},
            )

            producto = (
                Producto.objects
                .filter(id_cat_fk=catalogo)
                .filter(Q(nombre_producto__iexact=nombre) | Q(nombre_producto=nombre))
                .first()
            )
            is_created = producto is None
            if is_created:
                producto = Producto(id_cat_fk=catalogo, fch_registro=timezone.now())

            unidad = _cell_value(row_idx, 'unidad') or 'Unidad'
            unidad_map = {k.lower(): k for k, _ in Producto.UNIDAD_MEDIDA_CHOICES}
            unidad_display_map = {v.lower(): k for k, v in Producto.UNIDAD_MEDIDA_CHOICES}
            unidad_key = unidad_map.get(unidad.lower()) or unidad_display_map.get(unidad.lower()) or 'unidad'

            tipo = _cell_value(row_idx, 'tipo') or 'Devolutivo'
            tipo_map = {k.lower(): k for k, _ in Producto.TIPO_BIEN_CHOICES}
            tipo_display_map = {v.lower(): k for k, v in Producto.TIPO_BIEN_CHOICES}
            tipo_key = tipo_map.get(tipo.lower()) or tipo_display_map.get(tipo.lower()) or 'devolutivo'

            producto.nombre_producto = nombre
            producto.descripcion = _cell_value(row_idx, 'descripcion')
            producto.unidad_medida = unidad_key
            producto.ubicacion = _cell_value(row_idx, 'ubicacion') or 'Pendiente por asignar'
            producto.tipo_bien = tipo_key
            producto.numero_placa = _cell_value(row_idx, 'placa') if tipo_key == 'devolutivo' else ''
            producto.cuentadante = _cell_value(row_idx, 'cuentadante') if tipo_key == 'devolutivo' else ''
            producto.id_cat_fk = catalogo
            producto.fch_ult_act = timezone.now()
            producto.save()

            if is_created:
                created += 1
            else:
                updated += 1

            key = (categoria.lower(), nombre.lower())
            product_key_map[key] = producto

            subcats_raw = _cell_value(row_idx, 'subcategorias')
            subcats = _parse_excel_subcategorias(subcats_raw)
            selected_subcats = []
            for sub_name in subcats:
                ruta = [segmento.strip() for segmento in sub_name.split('/') if segmento.strip()]
                if not ruta:
                    continue
                sub_obj = Subcategoria.ensure_path(catalogo, ruta)
                selected_subcats.append(sub_obj)
            producto.subcategorias.set(selected_subcats)

            stock_txt = _cell_value(row_idx, 'stock')
            cantidad_txt = _cell_value(row_idx, 'cantidad')
            detalle_txt = _cell_value(row_idx, 'detalle')
            try:
                stock = int(float(stock_txt or 0))
            except Exception:
                stock = 0
            try:
                cantidad = int(float(cantidad_txt or 0))
            except Exception:
                cantidad = 0

            disponibilidad = Disponibilidad.objects.filter(id_prod_fk=producto).order_by('-id_disp').first()
            if not disponibilidad:
                disponibilidad = Disponibilidad(id_prod_fk=producto, fch_registro=timezone.now())
            disponibilidad.stock = stock
            disponibilidad.cantidad = cantidad
            disponibilidad.descr_dispo = detalle_txt
            disponibilidad.fch_ult_act = timezone.now()
            disponibilidad.save()

            if image_col and row_idx in main_images_by_row:
                for col_1_based, raw, ext in main_images_by_row[row_idx]:
                    if col_1_based != image_col:
                        continue
                    try:
                        main_name = f'producto_import_{producto.id_prod}_{row_idx}.{ext}'
                        producto.fot_prod.save(main_name, ContentFile(raw), save=True)
                        imported_main_images += 1
                        break
                    except Exception as exc:
                        errors.append(f'Imagen principal {nombre}: {exc}')

        if 'Imagenes Secundarias' in wb.sheetnames:
            sec_ws = wb['Imagenes Secundarias']
            sec_required = {'categoria', 'producto', 'orden'}
            sec_header_row = _find_header_row(sec_ws, sec_required, max_scan_rows=20)
            if sec_header_row:
                sec_header_to_col = {}
                for col in range(1, sec_ws.max_column + 1):
                    value = sec_ws.cell(row=sec_header_row, column=col).value
                    if value is None:
                        continue
                    sec_header_to_col[_header_key(str(value))] = col

                sec_image_col = sec_header_to_col.get('imagen')
                sec_images_by_row = _extract_images_by_row(sec_ws)
                cleared_products = set()

                for row_idx in range(sec_header_row + 1, sec_ws.max_row + 1):
                    categoria = str(sec_ws.cell(row=row_idx, column=sec_header_to_col.get('categoria', 1)).value or '').strip()
                    nombre = str(sec_ws.cell(row=row_idx, column=sec_header_to_col.get('producto', 1)).value or '').strip()
                    if not categoria or not nombre:
                        continue

                    orden_txt = str(sec_ws.cell(row=row_idx, column=sec_header_to_col.get('orden', 1)).value or '').strip()
                    try:
                        orden = int(float(orden_txt or 0))
                    except Exception:
                        orden = 0

                    producto = product_key_map.get((categoria.lower(), nombre.lower()))
                    if not producto:
                        continue

                    if sec_image_col and row_idx in sec_images_by_row:
                        if producto.id_prod not in cleared_products:
                            producto.fotos.all().delete()
                            cleared_products.add(producto.id_prod)

                        for col_1_based, raw, ext in sec_images_by_row[row_idx]:
                            if col_1_based != sec_image_col:
                                continue
                            try:
                                sec_name = f'producto_sec_import_{producto.id_prod}_{row_idx}_{orden}.{ext}'
                                foto_obj = ProductoFoto(id_prod_fk=producto, orden=orden)
                                foto_obj.foto.save(sec_name, ContentFile(raw), save=True)
                                imported_secondary_images += 1
                            except Exception as exc:
                                errors.append(f'Imagen secundaria {nombre}: {exc}')

    estado = 'ok' if not errors else 'ok_parcial'
    resumen = (
        f'Productos procesados: {total_rows}. '
        f'Creados: {created}. Actualizados: {updated}. '
        f'Imágenes principales: {imported_main_images}. '
        f'Imágenes secundarias: {imported_secondary_images}. '
        f'Errores: {len(errors)}.'
    )

    ImportacionInventarioLog.objects.create(
        id_usuario_fk=request.user if request.user.is_authenticated else None,
        nombre_archivo=os.path.basename(temp_path),
        estado=estado,
        total_productos=total_rows,
        total_creados=created,
        total_actualizados=updated,
        total_imagenes_principales=imported_main_images,
        total_imagenes_secundarias=imported_secondary_images,
        total_errores=len(errors),
        resumen=resumen + (f' Detalles: {" | ".join(errors[:5])}' if errors else ''),
    )

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='inventario_importacion',
        entidad_id='excel',
        descripcion=resumen,
    )

    return {
        'ok': True,
        'message': resumen,
        'errors': errors,
    }


@login_required
def importar_inventario_panel(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    logs = list(ImportacionInventarioLog.objects.select_related('id_usuario_fk')[:30])
    return render(
        request,
        'inventario/dashboard/importar_inventario_panel.html',
        {
            'logs_importacion': logs,
        },
    )


@login_required
def importar_inventario_carga(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    session_file_key = 'inventario_import_temp_path'
    session_name_key = 'inventario_import_file_name'

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip().lower()

        if action == 'preview':
            file_obj = request.FILES.get('excel_file')
            if not file_obj:
                messages.error(request, 'Selecciona un archivo Excel para continuar.')
            elif not file_obj.name.lower().endswith('.xlsx'):
                messages.error(request, 'El archivo debe ser .xlsx')
            else:
                old_path = request.session.get(session_file_key)
                if old_path and os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception:
                        pass

                temp_path = _save_temp_excel(file_obj)
                request.session[session_file_key] = temp_path
                request.session[session_name_key] = file_obj.name
                messages.success(request, 'Archivo cargado. Revisa la visualización antes de subir a la base de datos.')

        elif action == 'import':
            temp_path = request.session.get(session_file_key)
            if not temp_path or not os.path.exists(temp_path):
                messages.error(request, 'No hay un archivo cargado para importar. Primero usa "Cargar y visualizar".')
            else:
                result = _importar_excel_inventario(temp_path, request)
                if result.get('ok'):
                    messages.success(request, result.get('message') or 'Importación completada.')
                else:
                    messages.error(request, result.get('message') or 'No fue posible completar la importación.')

                try:
                    os.remove(temp_path)
                except Exception:
                    pass
                request.session.pop(session_file_key, None)
                request.session.pop(session_name_key, None)

        elif action == 'cancel':
            temp_path = request.session.get(session_file_key)
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass
            request.session.pop(session_file_key, None)
            request.session.pop(session_name_key, None)
            messages.info(request, 'Previsualización cancelada. Puedes seleccionar otro archivo.')
            return redirect('importar_inventario_panel')

        return redirect('importar_inventario_carga')

    q = (request.GET.get('q') or '').strip()
    temp_path = request.session.get(session_file_key)
    preview = {'groups': [], 'rows': [], 'error': ''}
    active_file_name = request.session.get(session_name_key, '')

    if temp_path and os.path.exists(temp_path):
        preview = _build_preview_from_excel(temp_path, q)
    else:
        request.session.pop(session_file_key, None)
        request.session.pop(session_name_key, None)

    return render(
        request,
        'inventario/dashboard/importar_inventario.html',
        {
            'preview_groups': preview.get('groups', []),
            'preview_rows': preview.get('rows', []),
            'preview_error': preview.get('error', ''),
            'preview_search': q,
            'active_file_name': active_file_name,
        },
    )


@login_required
def reporte_prestamos_pdf(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    anio, mes = _mes_reporte_desde_request(request)
    prestamos = list(_obtener_prestamos_mes(anio, mes))

    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        messages.error(request, 'No se pudo generar el PDF porque Pillow no está disponible en el servidor.')
        return redirect('dashboard')

    secciones = {
        'REALIZADO': [],
        'CANCELADO': [],
        'EN PROCESO': [],
    }
    for pedido in prestamos:
        secciones[_categoria_pedido_reporte(pedido.estado)].append(pedido)

    page_w, page_h = 1754, 1240
    margin_x = 36
    header_h = 126
    row_h = 70

    col_pedido = 70
    col_fecha = 150
    col_estado = 120
    col_usuario = 220
    col_rol = 110
    col_productos = 70
    col_unidades = 80
    col_detalle = 360
    col_devolucion = 150
    col_area = 352

    x_pedido = margin_x
    x_fecha = x_pedido + col_pedido
    x_estado = x_fecha + col_fecha
    x_usuario = x_estado + col_estado
    x_rol = x_usuario + col_usuario
    x_productos = x_rol + col_rol
    x_unidades = x_productos + col_productos
    x_detalle = x_unidades + col_unidades
    x_devolucion = x_detalle + col_detalle
    x_area = x_devolucion + col_devolucion
    x_end = x_area + col_area

    def _load_font(size, bold=False):
        candidates = []
        windir = os.environ.get('WINDIR', 'C:\\Windows')
        if bold:
            candidates.extend([
                os.path.join(windir, 'Fonts', 'arialbd.ttf'),
                os.path.join(windir, 'Fonts', 'segoeuib.ttf'),
            ])
        else:
            candidates.extend([
                os.path.join(windir, 'Fonts', 'arial.ttf'),
                os.path.join(windir, 'Fonts', 'segoeui.ttf'),
            ])
        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    font_title = _load_font(32, bold=True)
    font_sub = _load_font(19, bold=False)
    font_head = _load_font(17, bold=True)
    font_cell = _load_font(15, bold=False)

    pages = []
    page = None
    draw = None
    y = 0
    page_number = 0

    def _new_page():
        nonlocal page, draw, y, page_number
        page_number += 1
        page = Image.new('RGB', (page_w, page_h), 'white')
        draw = ImageDraw.Draw(page)

        draw.rectangle((margin_x, 28, x_end, 28 + header_h), outline='#2d7a49', width=2)

        logo_candidates = [
            os.path.join(settings.BASE_DIR, 'logoSena.png'),
            os.path.join(settings.BASE_DIR, 'inventario', 'static', 'inventario', 'img', 'logoSena.png'),
        ]
        logo_pasted = False
        for logo_path in logo_candidates:
            if not os.path.exists(logo_path):
                continue
            try:
                with Image.open(logo_path) as logo_src:
                    logo = logo_src.convert('RGBA')
                    logo.thumbnail((86, 86), Image.Resampling.LANCZOS)
                    page.paste(logo, (margin_x + 12, 46), logo)
                logo_pasted = True
                break
            except Exception:
                continue

        title_x = margin_x + (118 if logo_pasted else 16)
        draw.text((title_x, 48), f'REPORTE MENSUAL DE PEDIDOS - {anio}-{mes:02d}', fill='#1b6e3a', font=font_title)
        draw.text(
            (title_x, 94),
            f'Generado: {timezone.localtime().strftime("%d/%m/%Y %H:%M")} | Página {page_number}',
            fill='#4d7f62',
            font=font_sub,
        )

        y = 28 + header_h + 16

    def _draw_table_header(top_y):
        draw.rectangle((margin_x, top_y, x_end, top_y + 30), outline='#8cb99a', width=1, fill='#ecf7ef')
        draw.text((x_pedido + 6, top_y + 7), 'ID', fill='#205335', font=font_head)
        draw.text((x_fecha + 6, top_y + 7), 'FECHA', fill='#205335', font=font_head)
        draw.text((x_estado + 6, top_y + 7), 'ESTADO', fill='#205335', font=font_head)
        draw.text((x_usuario + 6, top_y + 7), 'USUARIO', fill='#205335', font=font_head)
        draw.text((x_rol + 6, top_y + 7), 'ROL', fill='#205335', font=font_head)
        draw.text((x_productos + 6, top_y + 7), 'PROD', fill='#205335', font=font_head)
        draw.text((x_unidades + 6, top_y + 7), 'UNDS', fill='#205335', font=font_head)
        draw.text((x_detalle + 6, top_y + 7), 'DETALLE PRODUCTOS', fill='#205335', font=font_head)
        draw.text((x_devolucion + 6, top_y + 7), 'DEVOLUCION', fill='#205335', font=font_head)
        draw.text((x_area + 6, top_y + 7), 'AREA', fill='#205335', font=font_head)
        return top_y + 30

    def _wrap_lines(texto, width_chars):
        bruto = (texto or '').strip() or '-'
        bloques = bruto.splitlines() or ['-']
        lines = []
        for bloque in bloques:
            limpio = (bloque or '').strip() or '-'
            lines.extend(textwrap.wrap(limpio, width=width_chars) or ['-'])
        return lines

    def _draw_wrapped_text(texto, x, top_y, width_chars, max_lines=2):
        lines = _wrap_lines(texto, width_chars)
        if max_lines is not None:
            lines = lines[:max_lines]
        for idx, line in enumerate(lines):
            draw.text((x, top_y + 7 + idx * 18), line, fill='#244f35', font=font_cell)

    _new_page()
    resumen_realizados = len(secciones['REALIZADO'])
    resumen_cancelados = len(secciones['CANCELADO'])
    resumen_proceso = len(secciones['EN PROCESO'])
    draw.text(
        (margin_x, y),
        f'Resumen -> Realizados: {resumen_realizados}  |  Cancelados: {resumen_cancelados}  |  En proceso: {resumen_proceso}  |  Total: {len(prestamos)}',
        fill='#2f6844',
        font=font_sub,
    )
    y += 34

    for sec_nombre in ['REALIZADO', 'CANCELADO', 'EN PROCESO']:
        items = secciones[sec_nombre]
        if y + 64 > page_h - 30:
            pages.append(page)
            _new_page()

        draw.rectangle((margin_x, y, x_end, y + 34), outline='#b7d5bf', width=1, fill='#f3fbf5')
        draw.text((margin_x + 10, y + 8), f'SECCION: {sec_nombre} ({len(items)})', fill='#245a39', font=font_head)
        y += 34
        y = _draw_table_header(y)

        if not items:
            if y + row_h > page_h - 30:
                pages.append(page)
                _new_page()
                y = _draw_table_header(y)
            draw.rectangle((margin_x, y, x_end, y + row_h), outline='#d2e2d6', width=1)
            draw.text((margin_x + 12, y + 24), 'Sin pedidos en esta sección.', fill='#5f7d67', font=font_cell)
            y += row_h
            y += 10
            continue

        for idx, pedido in enumerate(items):
            detalle_completo = _resumen_productos_pedido(pedido, multiline=True)
            detalle_lines = _wrap_lines(detalle_completo, 45)
            area_lines = _wrap_lines((pedido.area_ubicacion or '-').replace('\n', ' '), 41)
            fecha_lines = _wrap_lines(
                timezone.localtime(pedido.fch_registro).strftime('%Y-%m-%d %H:%M') if pedido.fch_registro else '-',
                16,
            )

            lineas_necesarias = max(len(detalle_lines), len(area_lines), len(fecha_lines), 2)
            row_h_actual = max(row_h, 12 + lineas_necesarias * 18)

            if y + row_h_actual > page_h - 30:
                pages.append(page)
                _new_page()
                y = _draw_table_header(y)

            bg = '#ffffff' if idx % 2 == 0 else '#f8fcf9'
            draw.rectangle((margin_x, y, x_end, y + row_h_actual), outline='#d2e2d6', width=1, fill=bg)
            for x_line in [x_fecha, x_estado, x_usuario, x_rol, x_productos, x_unidades, x_detalle, x_devolucion, x_area]:
                draw.line((x_line, y, x_line, y + row_h_actual), fill='#dcebdd', width=1)

            usuario = pedido.id_usuario_fk
            nombre_usuario = ((usuario.nombre or '') + ' ' + (usuario.apellido or '')).strip() or (usuario.correo or '-')
            rol = usuario.id_rol_fk.nombre_rol if usuario and usuario.id_rol_fk else '-'
            fecha_txt = timezone.localtime(pedido.fch_registro).strftime('%Y-%m-%d %H:%M') if pedido.fch_registro else '-'
            devolucion_txt = timezone.localtime(pedido.fecha_devolucion).strftime('%Y-%m-%d %H:%M') if pedido.fecha_devolucion else '-'

            draw.text((x_pedido + 6, y + 24), str(pedido.id_pedido), fill='#23543a', font=font_cell)
            _draw_wrapped_text(fecha_txt, x_fecha + 6, y, width_chars=16, max_lines=2)
            _draw_wrapped_text((pedido.estado or '-').title(), x_estado + 6, y, width_chars=14, max_lines=2)
            _draw_wrapped_text(nombre_usuario, x_usuario + 6, y, width_chars=25, max_lines=2)
            _draw_wrapped_text(rol, x_rol + 6, y, width_chars=11, max_lines=2)
            draw.text((x_productos + 20, y + 24), str(pedido.total_productos or 0), fill='#23543a', font=font_cell)
            draw.text((x_unidades + 20, y + 24), str(pedido.total_unidades or 0), fill='#23543a', font=font_cell)
            _draw_wrapped_text(detalle_completo, x_detalle + 6, y, width_chars=45, max_lines=None)
            _draw_wrapped_text(devolucion_txt, x_devolucion + 6, y, width_chars=15, max_lines=2)
            _draw_wrapped_text((pedido.area_ubicacion or '-').replace('\n', ' '), x_area + 6, y, width_chars=41, max_lines=None)

            y += row_h_actual

        y += 10

    if page is not None:
        pages.append(page)

    buffer = io.BytesIO()
    pages[0].save(buffer, format='PDF', save_all=True, append_images=pages[1:])
    pdf_bytes = buffer.getvalue()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_prestamos_{anio}_{mes:02d}.pdf"'
    return response


@login_required
def reporte_stock_bajo_pdf(request):
    if not _is_admin_or_almacenista(request):
        return redirect('panel_usuario')

    from django.db.models.functions import Coalesce

    try:
        from PIL import Image, ImageDraw, ImageFont, ImageOps
    except Exception:
        messages.error(request, 'No se pudo generar el PDF porque Pillow no está disponible en el servidor.')
        return redirect('dashboard')

    disp_qs = Disponibilidad.objects.filter(id_prod_fk=OuterRef('pk')).order_by('-id_disp')
    productos_qs = (
        Producto.objects
        .select_related('id_cat_fk')
        .annotate(
            stock_actual=Coalesce(Subquery(disp_qs.values('stock')[:1]), 0),
            cantidad_actual=Coalesce(Subquery(disp_qs.values('cantidad')[:1]), 0),
        )
        .filter(
            models.Q(stock_actual__lt=5)
            | models.Q(cantidad_actual__lt=5)
        )
        .order_by('stock_actual', 'cantidad_actual', 'nombre_producto')
    )

    productos = list(productos_qs)
    if not productos:
        messages.info(request, 'No hay productos con stock o cantidad baja para exportar.')
        return redirect('dashboard')

    # A4 horizontal para que la tabla salga amplia y legible en impresión.
    page_w, page_h = 1754, 1240
    margin_x = 44
    y_start = 200
    row_h = 148

    col_check = 58
    col_img = 130
    col_title = 300
    col_desc = 560
    col_cant = 120
    col_stock = 120
    col_estado = 220
    col_compra = 158

    x_check = margin_x
    x_img = x_check + col_check
    x_title = x_img + col_img
    x_desc = x_title + col_title
    x_cant = x_desc + col_desc
    x_stock = x_cant + col_cant
    x_estado = x_stock + col_stock
    x_compra = x_estado + col_estado
    x_end = x_compra + col_compra

    def _load_font(size, bold=False):
        candidates = []
        windir = os.environ.get('WINDIR', 'C:\\Windows')
        if bold:
            candidates.extend([
                os.path.join(windir, 'Fonts', 'arialbd.ttf'),
                os.path.join(windir, 'Fonts', 'segoeuib.ttf'),
            ])
        else:
            candidates.extend([
                os.path.join(windir, 'Fonts', 'arial.ttf'),
                os.path.join(windir, 'Fonts', 'segoeui.ttf'),
            ])

        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    font_title = _load_font(42, bold=True)
    font_sub = _load_font(22, bold=False)
    font_head = _load_font(20, bold=True)
    font_cell = _load_font(18, bold=False)
    font_small = _load_font(16, bold=False)

    pages = []
    page = None
    draw = None
    y = y_start
    page_number = 0

    def _new_page():
        nonlocal page, draw, y, page_number
        page_number += 1
        page = Image.new('RGB', (page_w, page_h), 'white')
        draw = ImageDraw.Draw(page)

        header_top = 48
        header_bottom = 170
        draw.rectangle((margin_x, header_top, x_end, header_bottom), outline='#1f3f67', width=3)

        logo_candidates = [
            os.path.join(settings.BASE_DIR, 'logoSena.png'),
            os.path.join(settings.BASE_DIR, 'inventario', 'static', 'inventario', 'img', 'logoSena.png'),
        ]
        logo_pasted = False
        for logo_path in logo_candidates:
            if not os.path.exists(logo_path):
                continue
            try:
                with Image.open(logo_path) as logo_src:
                    logo = logo_src.convert('RGBA')
                    logo.thumbnail((98, 98), Image.Resampling.LANCZOS)
                    logo_x = margin_x + 16
                    logo_y = header_top + 12
                    page.paste(logo, (logo_x, logo_y), logo)
                logo_pasted = True
                break
            except Exception:
                continue

        title_x = margin_x + (132 if logo_pasted else 18)
        draw.text(
            (title_x, header_top + 28),
            'RECIBO SENA DE PRODUCTOS EN ALERTA',
            fill='#00843d',
            font=font_title,
        )
        draw.text(
            (title_x, header_top + 82),
            f'Generado: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}  |  Página {page_number}',
            fill='#3e5f82',
            font=font_sub,
        )

        draw.rectangle((margin_x, y_start - 38, x_end, y_start), outline='#7f96b2', width=1, fill='#f2f7fd')
        draw.text((x_check + 18, y_start - 29), 'X', fill='#25496c', font=font_head)
        draw.text((x_img + 20, y_start - 29), 'IMAGEN', fill='#25496c', font=font_head)
        draw.text((x_title + 10, y_start - 29), 'TITULO', fill='#25496c', font=font_head)
        draw.text((x_desc + 10, y_start - 29), 'DESCRIPCION', fill='#25496c', font=font_head)
        draw.text((x_cant + 15, y_start - 29), 'CANTIDAD', fill='#25496c', font=font_head)
        draw.text((x_stock + 24, y_start - 29), 'STOCK', fill='#25496c', font=font_head)
        draw.text((x_estado + 12, y_start - 29), 'MOTIVO', fill='#25496c', font=font_head)
        draw.text((x_compra + 12, y_start - 29), 'COMPRA', fill='#25496c', font=font_head)

        y = y_start

    def _draw_cell_box(top_y):
        draw.rectangle((margin_x, top_y, x_end, top_y + row_h), outline='#c7d4e4', width=1)
        draw.line((x_img, top_y, x_img, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_title, top_y, x_title, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_desc, top_y, x_desc, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_cant, top_y, x_cant, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_stock, top_y, x_stock, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_estado, top_y, x_estado, top_y + row_h), fill='#d2ddec', width=1)
        draw.line((x_compra, top_y, x_compra, top_y + row_h), fill='#d2ddec', width=1)

        # Cuadro de check para marcar con esfero.
        draw.rectangle((x_check + 18, top_y + 54, x_check + 42, top_y + 78), outline='#2a4b6f', width=2)

    def _draw_wrapped(texto, x, top_y, width_chars=42, max_lines=4, fill='#1f3856', font=None, line_h=22):
        limpio = (texto or '').strip()
        if not limpio:
            limpio = '-'
        lines = textwrap.wrap(limpio, width=width_chars)[:max_lines]
        use_font = font or font_cell
        for idx, line in enumerate(lines):
            draw.text((x, top_y + 10 + idx * line_h), line, fill=fill, font=use_font)

    _new_page()
    for prod in productos:
        if y + row_h > page_h - 80:
            pages.append(page)
            _new_page()

        _draw_cell_box(y)

        stock = int(prod.stock_actual or 0)
        cantidad = int(prod.cantidad_actual or 0)
        if cantidad < stock:
            estado_txt = f'Faltan {stock - cantidad} und'
        elif stock < 5 and cantidad < 5:
            estado_txt = 'Stock y cantidad bajos'
        elif stock < 5:
            estado_txt = 'Stock bajo'
        else:
            estado_txt = 'Cantidad baja'

        # Imagen del producto.
        if getattr(prod, 'fot_prod', None):
            try:
                ruta_imagen = prod.fot_prod.path
                if os.path.exists(ruta_imagen):
                    with Image.open(ruta_imagen) as img_src:
                        thumb = ImageOps.fit(img_src.convert('RGB'), (106, 106), method=Image.Resampling.LANCZOS)
                        page.paste(thumb, (x_img + 12, y + 20))
            except Exception:
                draw.rectangle((x_img + 12, y + 20, x_img + 118, y + 126), outline='#c8d6e7', width=1)
                draw.text((x_img + 28, y + 68), 'Sin img', fill='#6e839d', font=font_small)
        else:
            draw.rectangle((x_img + 12, y + 20, x_img + 118, y + 126), outline='#c8d6e7', width=1)
            draw.text((x_img + 28, y + 68), 'Sin img', fill='#6e839d', font=font_small)

        _draw_wrapped(prod.nombre_producto, x_title + 10, y, width_chars=30, max_lines=3, font=font_cell, line_h=24)
        _draw_wrapped(prod.descripcion, x_desc + 10, y, width_chars=58, max_lines=5, font=font_small, line_h=20)
        draw.text((x_cant + 44, y + 60), str(cantidad), fill='#1f3856', font=font_head)
        draw.text((x_stock + 44, y + 60), str(stock), fill='#1f3856', font=font_head)
        _draw_wrapped(estado_txt, x_estado + 10, y, width_chars=22, max_lines=4, font=font_small, line_h=20)

        # Espacio para que escriban la cantidad comprada a mano.
        draw.line((x_compra + 12, y + 70, x_end - 14, y + 70), fill='#355a82', width=1)
        draw.text((x_compra + 12, y + 82), 'cantidad comprada', fill='#6f86a2', font=font_small)

        y += row_h

    if page is not None:
        pages.append(page)

    buffer = io.BytesIO()
    pages[0].save(buffer, format='PDF', save_all=True, append_images=pages[1:])
    pdf_bytes = buffer.getvalue()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="recibo_stock_bajo_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf"'
    )
    return response


@login_required
def perfil_usuario(request):
    usuario = request.user
    tipo_doc_habilitado = usuario_supports_tipo_doc(Usuario)
    if request.method == 'POST':
        form = UsuarioPerfilForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('perfil_usuario')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = UsuarioPerfilForm(instance=usuario)

    pedidos_qs = Pedido.objects.filter(id_usuario_fk=usuario).order_by('-fch_registro', '-id_pedido')
    pedido_stats = {
        'total': pedidos_qs.count(),
        'pendientes': pedidos_qs.filter(estado='pendiente').count(),
        'entregados': pedidos_qs.filter(estado__in=['entregado', 'devuelto']).count(),
        'cancelados': pedidos_qs.filter(estado__in=['cancelado', 'rechazado']).count(),
    }
    pedidos_recientes = list(pedidos_qs[:5])

    return render(request, 'inventario/usuario/perfil_usuario.html', {
        'form': form,
        'password_form': CambioPasswordPerfilForm(usuario),
        'usuario': usuario,
        'tipo_doc_habilitado': tipo_doc_habilitado,
        'pedido_stats': pedido_stats,
        'pedidos_recientes': pedidos_recientes,
    })


@login_required
def perfil_cambiar_password(request):
    if request.method != 'POST':
        return redirect('perfil_usuario')

    form = CambioPasswordPerfilForm(request.user, request.POST)
    if form.is_valid():
        form.save()
        update_session_auth_hash(request, request.user)
        _registrar_auditoria(
            request,
            accion='actualizar',
            entidad='usuario',
            entidad_id=request.user.pk,
            descripcion='Cambio de contraseña desde perfil de usuario.',
        )
        messages.success(request, 'Tu contraseña se actualizó correctamente.')
    else:
        errors = []
        for field_errors in form.errors.values():
            errors.extend(field_errors)
        if errors:
            messages.error(request, ' '.join(str(err) for err in errors))
        else:
            messages.error(request, 'No fue posible cambiar la contraseña. Intenta de nuevo.')

    return redirect('perfil_usuario')


@login_required
def manual_usuario(request):
    rol = _user_role(request) or 'usuario'
    return render(request, 'inventario/usuario/manual_usuario.html', {
        'rol_manual': rol,
    })


@login_required
def perfil_actualizar_banner(request):
    """Endpoint AJAX para guardar solo el banner/portada del perfil."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)
    archivo = request.FILES.get('banner_usu')
    if not archivo:
        return JsonResponse({'ok': False, 'error': 'No se recibió ninguna imagen.'}, status=400)
    if not archivo.content_type.startswith('image/'):
        return JsonResponse({'ok': False, 'error': 'El archivo debe ser una imagen.'}, status=400)
    usuario = request.user
    # Eliminar banner anterior para evitar archivos huérfanos
    if usuario.banner_usu:
        try:
            usuario.banner_usu.delete(save=False)
        except Exception:
            pass
    usuario.banner_usu = archivo
    usuario.save(update_fields=['banner_usu'])
    return JsonResponse({'ok': True, 'url': usuario.banner_usu.url})


@login_required
def perfil_actualizar_tema(request):
    """Endpoint AJAX para guardar la preferencia de tema del usuario en la BD."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)
    tema = (request.POST.get('tema') or '').strip().lower()
    if tema not in ('claro', 'oscuro'):
        return JsonResponse({'ok': False, 'error': 'Valor de tema inválido.'}, status=400)
    request.user.tema = tema
    request.user.save(update_fields=['tema'])
    return JsonResponse({'ok': True, 'tema': tema})


@login_required
def prestamos_panel(request):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    ahora = timezone.now()
    prestamos = list(
        Pedido.objects
        .filter(estado__in=['entregado', 'vencido', 'devuelto', 'rechazado'])
        .select_related('id_usuario_fk')
        .prefetch_related('detalles')
        .order_by('fecha_devolucion', '-fch_registro')
    )

    for prestamo in prestamos:
        estado_canonico = _estado_pedido_canonico(prestamo.estado)
        if estado_canonico and estado_canonico != prestamo.estado:
            prestamo.estado = estado_canonico
            prestamo.fch_ult_act = ahora
            prestamo.save(update_fields=['estado', 'fch_ult_act'])

        if prestamo.estado == 'devuelto' and (prestamo.codigo_entrega or prestamo.codigo_expira_en):
            prestamo.codigo_entrega = None
            prestamo.codigo_expira_en = None
            prestamo.fch_ult_act = ahora
            prestamo.save(update_fields=['codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])

        detalles = list(prestamo.detalles.all())
        prestamo.detalles_entregados = [
            detalle for detalle in detalles
            if detalle.estado_detalle not in ['no_disponible', 'rechazado', 'cancelado']
        ]
        prestamo.fecha_cierre_display = prestamo.fch_ult_act

        # Los cancelados/devueltos nunca están activos: sin vencimiento
        if prestamo.estado in ['rechazado', 'devuelto']:
            prestamo.fecha_devolucion_display = None
            prestamo.es_vencido = False
            prestamo.dias_restantes = None
            prestamo.dias_vencido = 0
            prestamo.tiempo_vencido_str = ''
            prestamo.tiempo_restante_str = ''
            prestamo.detalles_lista = detalles
            continue

        # Préstamos marcados automáticamente como vencidos
        if prestamo.estado == 'vencido':
            prestamo.fecha_devolucion_display = prestamo.fecha_devolucion
            prestamo.es_vencido = True
            if prestamo.fecha_devolucion:
                delta = prestamo.fecha_devolucion - ahora
                prestamo.dias_restantes = delta.days
                prestamo.dias_vencido = abs(delta.days)
                prestamo.tiempo_vencido_str = _tiempo_vencido(prestamo.fecha_devolucion, ahora)
            else:
                prestamo.dias_restantes = None
                prestamo.dias_vencido = 0
                prestamo.tiempo_vencido_str = ''
            prestamo.tiempo_restante_str = ''
            prestamo.detalles_lista = detalles
            continue

        if prestamo.tipo_devolucion == 'individual':
            fechas = [d.fecha_devolucion for d in detalles if d.fecha_devolucion]
            if fechas:
                fecha_ref = min(fechas)
                prestamo.fecha_devolucion_display = fecha_ref
                prestamo.es_vencido = fecha_ref < ahora
                delta = fecha_ref - ahora
                prestamo.dias_restantes = delta.days
                prestamo.dias_vencido = abs(delta.days) if prestamo.es_vencido else 0
                prestamo.tiempo_vencido_str = _tiempo_vencido(fecha_ref, ahora) if prestamo.es_vencido else ''
                prestamo.tiempo_restante_str = _tiempo_restante(fecha_ref, ahora) if not prestamo.es_vencido else ''
            else:
                prestamo.fecha_devolucion_display = None
                prestamo.es_vencido = False
                prestamo.dias_restantes = None
                prestamo.dias_vencido = 0
                prestamo.tiempo_vencido_str = ''
                prestamo.tiempo_restante_str = ''
        else:
            prestamo.fecha_devolucion_display = prestamo.fecha_devolucion
            if prestamo.fecha_devolucion:
                prestamo.es_vencido = prestamo.fecha_devolucion < ahora
                delta = prestamo.fecha_devolucion - ahora
                prestamo.dias_restantes = delta.days
                prestamo.dias_vencido = abs(delta.days) if prestamo.es_vencido else 0
                prestamo.tiempo_vencido_str = _tiempo_vencido(prestamo.fecha_devolucion, ahora) if prestamo.es_vencido else ''
                prestamo.tiempo_restante_str = _tiempo_restante(prestamo.fecha_devolucion, ahora) if not prestamo.es_vencido else ''
            else:
                prestamo.es_vencido = False
                prestamo.dias_restantes = None
                prestamo.dias_vencido = 0
                prestamo.tiempo_vencido_str = ''
                prestamo.tiempo_restante_str = ''
        prestamo.detalles_lista = detalles

    # Ordenar: vencidos primero, luego activos al día, después devueltos y cancelados.
    prestamos.sort(key=lambda p: (
        0 if p.estado == 'vencido' or (p.estado == 'entregado' and p.es_vencido) else 1 if p.estado == 'entregado' else 2 if p.estado == 'devuelto' else 3,
        p.fecha_devolucion_display or ahora.replace(year=9999),
        -(p.fecha_cierre_display.timestamp()) if p.fecha_cierre_display else 0,
    ))

    q_busqueda = (request.GET.get('q') or '').strip()
    fecha_desde_raw = (request.GET.get('fecha_desde') or '').strip()
    fecha_hasta_raw = (request.GET.get('fecha_hasta') or '').strip()

    fecha_desde = None
    fecha_hasta = None
    try:
        if fecha_desde_raw:
            fecha_desde = date.fromisoformat(fecha_desde_raw)
    except ValueError:
        fecha_desde = None
    try:
        if fecha_hasta_raw:
            fecha_hasta = date.fromisoformat(fecha_hasta_raw)
    except ValueError:
        fecha_hasta = None

    if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    def _normalize_search_text(value):
        text = str(value or '').strip().lower()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r'[^a-z0-9]+', '', text)
        return text

    normalized_query = _normalize_search_text(q_busqueda)

    def _tipo_devolucion_label(tipo):
        if tipo == 'individual':
            return 'individual'
        if tipo == 'mismo_dia':
            return 'mismo dia'
        if tipo == 'por_dias':
            return 'por dias'
        return 'global'

    def _estado_label(prestamo):
        if prestamo.estado == 'rechazado':
            return 'cancelado'
        if prestamo.estado == 'devuelto':
            return 'devuelto'
        if prestamo.estado == 'vencido' or (prestamo.estado == 'entregado' and prestamo.es_vencido):
            return 'vencido'
        if prestamo.estado == 'entregado':
            return 'al dia'
        return prestamo.estado or ''

    prestamos_filtrables = []
    for prestamo in prestamos:
        fecha_registro_local = timezone.localtime(prestamo.fch_registro) if prestamo.fch_registro else None
        fecha_registro_date = fecha_registro_local.date() if fecha_registro_local else None

        if fecha_desde and (not fecha_registro_date or fecha_registro_date < fecha_desde):
            continue
        if fecha_hasta and (not fecha_registro_date or fecha_registro_date > fecha_hasta):
            continue

        if normalized_query:
            usuario = prestamo.id_usuario_fk
            detalle_textos = []
            for det in prestamo.detalles_entregados:
                detalle_textos.extend([
                    det.nombre_producto,
                    det.nombre_catalogo,
                ])

            estado_texto = _estado_label(prestamo)
            tipo_texto = _tipo_devolucion_label(prestamo.tipo_devolucion)
            fecha_devolucion = prestamo.fecha_devolucion_display
            fecha_cierre = prestamo.fecha_cierre_display

            search_fields = [
                prestamo.id_pedido,
                usuario.cc if usuario else '',
                usuario.nombre if usuario else '',
                usuario.apellido if usuario else '',
                f'{(usuario.nombre or "") if usuario else ""} {(usuario.apellido or "") if usuario else ""}',
                usuario.correo if usuario else '',
                prestamo.area_ubicacion,
                prestamo.estado,
                estado_texto,
                tipo_texto,
                fecha_registro_local.strftime('%d/%m/%Y %H:%M') if fecha_registro_local else '',
                fecha_registro_local.strftime('%Y-%m-%d') if fecha_registro_local else '',
                fecha_devolucion.strftime('%d/%m/%Y %H:%M') if fecha_devolucion else '',
                fecha_devolucion.strftime('%Y-%m-%d') if fecha_devolucion else '',
                fecha_cierre.strftime('%d/%m/%Y %H:%M') if fecha_cierre else '',
                fecha_cierre.strftime('%Y-%m-%d') if fecha_cierre else '',
                'prestamo',
                'pedido',
            ] + detalle_textos

            joined = _normalize_search_text(' '.join(str(x or '') for x in search_fields))
            if normalized_query not in joined:
                continue

        prestamos_filtrables.append(prestamo)

    total_cancelados = sum(1 for p in prestamos_filtrables if p.estado == 'rechazado')
    total_devueltos = sum(1 for p in prestamos_filtrables if p.estado == 'devuelto')
    total_vencidos = sum(1 for p in prestamos_filtrables if p.estado == 'vencido' or (p.estado == 'entregado' and p.es_vencido))
    total_activos = sum(1 for p in prestamos_filtrables if p.estado in ('entregado', 'vencido'))
    total_al_dia = total_activos - total_vencidos

    filtro = (request.GET.get('filtro') or 'todos').strip().lower()
    if filtro not in {'todos', 'vencido', 'al-dia', 'devuelto', 'cancelado'}:
        filtro = 'todos'

    prestamos = list(prestamos_filtrables)
    if filtro == 'vencido':
        prestamos = [p for p in prestamos if p.estado == 'vencido' or (p.estado == 'entregado' and p.es_vencido)]
    elif filtro == 'al-dia':
        prestamos = [p for p in prestamos if p.estado == 'entregado' and not p.es_vencido]
    elif filtro == 'devuelto':
        prestamos = [p for p in prestamos if p.estado == 'devuelto']
    elif filtro == 'cancelado':
        prestamos = [p for p in prestamos if p.estado == 'rechazado']

    return render(request, 'inventario/prestamos/panel_prestamos.html', {
        'prestamos': prestamos,
        'filtro_activo': filtro,
        'total_vencidos': total_vencidos,
        'total_al_dia': total_al_dia,
        'total_cancelados': total_cancelados,
        'total_devueltos': total_devueltos,
        'total_activos': total_activos,
        'ahora': ahora,
        'q_busqueda': q_busqueda,
        'fecha_desde': fecha_desde.isoformat() if fecha_desde else '',
        'fecha_hasta': fecha_hasta.isoformat() if fecha_hasta else '',
    })

@login_required
def pedidos_panel(request):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    _auto_cancelar_pedidos_pendientes_vencidos()

    pedidos = (
        Pedido.objects
        .filter(estado__in=['pendiente', 'esperando entrega'])
        .select_related('id_usuario_fk')
        .prefetch_related('detalles')
        .order_by('-fch_registro', '-id_pedido')
    )
    return render(request, 'inventario/pedidos/panel_pedidos.html', {
        'pedidos': pedidos,
    })


@login_required
def pedido_detalle_panel(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    _auto_cancelar_pedidos_pendientes_vencidos()

    pedido = get_object_or_404(
        Pedido.objects.select_related('id_usuario_fk').prefetch_related('detalles__id_prod_fk__subcategorias', 'evidencias'),
        pk=pedido_id,
    )

    for detalle in pedido.detalles.all():
        detalle.cantidad_disponible_actual = 0
        detalle.subcategorias_lista = []
        if detalle.id_prod_fk_id:
            disp_actual = (
                Disponibilidad.objects
                .filter(id_prod_fk_id=detalle.id_prod_fk_id)
                .order_by('-id_disp')
                .first()
            )
            if disp_actual:
                detalle.cantidad_disponible_actual = (
                    disp_actual.cantidad if disp_actual.cantidad is not None else (disp_actual.stock or 0)
                )
            detalle.subcategorias_lista = [
                subcat.nombre_subcategoria
                for subcat in detalle.id_prod_fk.subcategorias.all()
                if getattr(subcat, 'nombre_subcategoria', None)
            ]

    consumo_total = 0
    devolutivo_total = 0
    for detalle in pedido.detalles.all():
        cantidad = detalle.cantidad_solicitada or 0
        tipo_bien = getattr(detalle.id_prod_fk, 'tipo_bien', '') if detalle.id_prod_fk else ''
        if tipo_bien == 'consumo':
            consumo_total += cantidad
        else:
            devolutivo_total += cantidad

    tipo_solicitud_items = []
    if consumo_total:
        tipo_solicitud_items.append({
            'cantidad': consumo_total,
            'texto': 'producto de consumo' if consumo_total == 1 else 'productos de consumo',
        })
    if devolutivo_total:
        tipo_solicitud_items.append({
            'cantidad': devolutivo_total,
            'texto': 'producto devolvible' if devolutivo_total == 1 else 'productos devolvibles',
        })

    if consumo_total and not devolutivo_total:
        tipo_solicitud_ayuda = 'Este producto es para un solo uso y no es devolvible.'
    elif devolutivo_total and not consumo_total:
        tipo_solicitud_ayuda = 'Este producto es devolvible y debe retornarse después de su uso.'
    else:
        tipo_solicitud_ayuda = (
            'Producto de consumo: es para un solo uso y no es devolvible. '
            'Producto devolvible: debe retornarse después de su uso.'
        )

    ahora = timezone.now()
    pedido_es_vencido = (
        pedido.estado in ['entregado', 'vencido']
        and pedido.fecha_devolucion is not None
        and pedido.fecha_devolucion <= ahora
    )
    estado_ui = 'vencido' if pedido_es_vencido else pedido.estado
    estado_ui_label = 'Vencido' if estado_ui == 'vencido' else pedido.estado.title()

    return render(request, 'inventario/pedidos/pedido_detalle.html', {
        'pedido': pedido,
        'estado_ui': estado_ui,
        'estado_ui_label': estado_ui_label,
        'tipo_solicitud_items': tipo_solicitud_items,
        'tipo_solicitud_ayuda': tipo_solicitud_ayuda,
    })


@login_required
def pedido_marcar_esperando_entrega(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    detalle_ids_raw = request.POST.getlist('detalle_no_disponible')
    motivo_no_disponible = (request.POST.get('motivo_no_disponible') or '').strip()
    try:
        detalle_ids_no_disponibles = sorted({int(d) for d in detalle_ids_raw if str(d).strip()})
    except (ValueError, TypeError):
        messages.error(request, 'La selección de productos no disponibles es inválida.')
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    with transaction.atomic():
        pedido = get_object_or_404(
            Pedido.objects.select_for_update(),
            pk=pedido_id,
        )

        if pedido.estado != 'pendiente':
            messages.error(request, 'Este pedido ya fue procesado previamente.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        ids_validos = set(
            DetallePedido.objects
            .filter(id_pedido_fk=pedido)
            .values_list('id_det_pedido', flat=True)
        )
        if not set(detalle_ids_no_disponibles).issubset(ids_validos):
            messages.error(request, 'Hay productos no disponibles seleccionados que no pertenecen al pedido.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        now = timezone.now()

        # Revertir marcas previas antiguas para que la selección actual sea la única fuente de verdad.
        DetallePedido.objects.filter(
            id_pedido_fk=pedido,
            estado_detalle='no_disponible',
        ).exclude(id_det_pedido__in=detalle_ids_no_disponibles).update(
            estado_detalle='pendiente',
            fch_ult_act=now,
        )

        if detalle_ids_no_disponibles:
            DetallePedido.objects.filter(
                id_pedido_fk=pedido,
                id_det_pedido__in=detalle_ids_no_disponibles,
            ).update(
                estado_detalle='no_disponible',
                fch_ult_act=now,
            )

        detalles = list(
            DetallePedido.objects
            .select_for_update()
            .select_related('id_prod_fk')
            .filter(id_pedido_fk=pedido)
            .exclude(estado_detalle__in=['no_disponible', 'rechazado', 'cancelado'])
            .order_by('id_det_pedido')
        )

        if not detalles:
            messages.error(request, 'El pedido no tiene productos disponibles para procesar.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        disponibilidad_por_detalle = {}
        errores = []

        for detalle in detalles:
            if not detalle.id_prod_fk_id:
                errores.append(f'Producto sin referencia en detalle #{detalle.id_det_pedido}.')
                continue

            disp = (
                Disponibilidad.objects
                .select_for_update()
                .filter(id_prod_fk_id=detalle.id_prod_fk_id)
                .order_by('-id_disp')
                .first()
            )

            if not disp:
                errores.append(f'Sin disponibilidad para {detalle.nombre_producto}.')
                continue

            disponible = disp.cantidad if disp.cantidad is not None else (disp.stock or 0)
            if disponible < detalle.cantidad_solicitada:
                errores.append(
                    f'Cantidad insuficiente en {detalle.nombre_producto} (solicita {detalle.cantidad_solicitada}, disponible {disponible}).'
                )
                continue

            disponibilidad_por_detalle[detalle.id_det_pedido] = disp

        if errores:
            messages.error(request, 'No se pudo procesar el pedido: ' + ' '.join(errores))
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        for detalle in detalles:
            disp = disponibilidad_por_detalle.get(detalle.id_det_pedido)
            if not disp:
                continue

            _ajustar_stock_disponibilidad(detalle, now, -int(detalle.cantidad_solicitada or 0))

            detalle.estado_detalle = 'esperando entrega'
            detalle.fch_ult_act = now
            detalle.save(update_fields=['estado_detalle', 'fch_ult_act'])

        codigo_entrega = f'{secrets.randbelow(1000000):06d}'
        pedido.estado = 'esperando entrega'
        pedido.codigo_entrega = codigo_entrega
        pedido.codigo_expira_en = now + timedelta(hours=2)
        pedido.motivo_rechazo = None
        pedido.fch_ult_act = now
        pedido.save(update_fields=['estado', 'codigo_entrega', 'codigo_expira_en', 'motivo_rechazo', 'fch_ult_act'])

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='pedido',
        entidad_id=pedido.id_pedido,
        descripcion=f'Pedido #{pedido.id_pedido} pasó a esperando entrega.',
    )
    messages.success(request, f'Pedido #{pedido.id_pedido} procesado. Codigo de entrega generado por 2 horas.')
    _crear_notificacion(
        usuario=pedido.id_usuario_fk,
        tipo='esperando_entrega',
        titulo='Tu pedido está listo para entrega',
        mensaje=f'Tu pedido #{pedido.id_pedido} fue aprobado y está esperando ser entregado. '
                f'Dirígete al almacén con tu código de entrega.',
        pedido_id=pedido.id_pedido,
    )

    if detalle_ids_no_disponibles:
        mensaje_no_disp = (
            f'En tu pedido #{pedido.id_pedido}, {len(detalle_ids_no_disponibles)} '
            + ('productos no están disponibles. ' if len(detalle_ids_no_disponibles) != 1 else 'producto no está disponible. ')
            + 'El resto del pedido continúa en proceso.'
        )
        if motivo_no_disponible:
            mensaje_no_disp += f' Motivo informado: {motivo_no_disponible}'

        _crear_notificacion(
            usuario=pedido.id_usuario_fk,
            tipo='no_disponible',
            titulo='Algunos productos no están disponibles',
            mensaje=mensaje_no_disp,
            pedido_id=pedido.id_pedido,
        )

    # ── Correo: pedido listo para recoger ────────────────────────────────
    try:
        from django.core.mail import EmailMultiAlternatives
        usuario = pedido.id_usuario_fk
        correo_dest = getattr(usuario, 'correo', None) or getattr(usuario, 'email', None)
        if correo_dest:
            nombre = getattr(usuario, 'nombre', '') or str(usuario)
            fecha_str = pedido.fecha_devolucion.strftime('%d/%m/%Y a las %H:%M') if pedido.fecha_devolucion else 'Sin fecha definida'
            base_url = 'https://almacensedelacolonia.pythonanywhere.com'
            detalles_list = list(pedido.detalles.exclude(estado_detalle__in=['no_disponible', 'rechazado', 'cancelado']).select_related('id_prod_fk'))
            filas_html = ''
            lista_txt = ''
            for d in detalles_list:
                prod = d.id_prod_fk
                img_url = f'{base_url}{settings.MEDIA_URL}{prod.fot_prod}' if prod and prod.fot_prod else ''
                img_tag = (f'<img src="{img_url}" width="44" height="44" style="border-radius:6px;object-fit:cover;">'
                           if img_url else '<div style="width:44px;height:44px;background:#e8f5e9;border-radius:6px;display:inline-block;">📦</div>')
                filas_html += f'<tr><td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;width:60px;">{img_tag}</td><td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;font-size:14px;color:#333;">{d.nombre_producto}</td><td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;font-size:14px;color:#555;text-align:center;">x{d.cantidad_solicitada}</td></tr>'
                lista_txt += f'  - {d.nombre_producto} x{d.cantidad_solicitada}\n'
            tabla = f'<p style="font-size:15px;font-weight:700;color:#1a2e1a;margin:20px 0 8px;">📦 Productos a recoger:</p><table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e0e0e0;border-radius:8px;overflow:hidden;"><thead><tr style="background:#f5f5f5;"><th style="padding:10px 12px;text-align:left;font-size:13px;color:#666;width:60px;">Foto</th><th style="padding:10px 12px;text-align:left;font-size:13px;color:#666;">Producto</th><th style="padding:10px 12px;text-align:center;font-size:13px;color:#666;">Cant.</th></tr></thead><tbody>{filas_html}</tbody></table>' if filas_html else ''
            asunto = f'🎉 Tu pedido #{pedido.id_pedido} está listo para recoger | Almacén SENA Sibaté'
            txt = f'Hola {nombre},\n\nTu pedido #{pedido.id_pedido} fue aprobado y está listo para ser retirado en el almacén.\n\nProductos a recoger:\n{lista_txt}\nFecha de devolución: {fecha_str}\n\nDirígete al almacén y muestra tu código de entrega.\n\n— Almacén SENA Sibaté'
            html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;padding:32px 0;">
<tr><td align="center"><table width="600" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);max-width:600px;width:100%;">
<tr><td style="background:#2196F3;padding:28px 32px;text-align:center;">
  <p style="margin:0;color:#fff;font-size:13px;opacity:0.85;">SENA — Almacén Sibaté</p>
  <h1 style="margin:8px 0 0;color:#fff;font-size:24px;">🎉 ¡Tu pedido está listo!</h1>
</td></tr>
<tr><td style="padding:32px;">
  <p style="font-size:16px;color:#333;">Hola <strong>{nombre}</strong>,</p>
  <p style="font-size:15px;color:#444;line-height:1.6;">Tu pedido <strong>#{pedido.id_pedido}</strong> fue <strong>aprobado</strong> y ya está listo para ser retirado en el almacén.</p>
  {tabla}
  <table width="100%" cellpadding="0" cellspacing="0" style="margin:20px 0;"><tr>
    <td style="background:#e3f2fd;border-left:4px solid #2196F3;border-radius:6px;padding:14px 18px;">
      <p style="margin:0 0 6px;font-size:14px;color:#333;">📅 Fecha de devolución: <strong>{fecha_str}</strong></p>
      <p style="margin:0;font-size:14px;color:#333;">🏪 Dirígete al almacén y muestra tu <strong>código de entrega</strong>.</p>
    </td>
  </tr></table>
  <p style="font-size:13px;color:#888;margin-top:28px;">— Almacén SENA Sibaté</p>
</td></tr>
<tr><td style="background:#f9f9f9;padding:14px 32px;text-align:center;border-top:1px solid #eee;">
  <p style="margin:0;font-size:12px;color:#aaa;">Centro Industrial y de Desarrollo Empresarial – Sibaté, Cundinamarca</p>
</td></tr>
</table></td></tr></table>
</body></html>"""
            msg = EmailMultiAlternatives(asunto, txt, settings.DEFAULT_FROM_EMAIL, [correo_dest])
            msg.attach_alternative(html, 'text/html')
            msg.send()
    except Exception:
        pass  # No bloquear el flujo si el correo falla

    return redirect('pedido_detalle_panel', pedido_id=pedido_id)


@login_required
def pedido_confirmar_entrega_codigo(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    codigo_ingresado = (request.POST.get('codigo_entrega') or '').strip()
    if not (len(codigo_ingresado) == 6 and codigo_ingresado.isdigit()):
        messages.error(request, 'El codigo debe tener 6 digitos numericos.')
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    evidencias_subidas = request.FILES.getlist('evidencias_entrega')
    if len(evidencias_subidas) > 8:
        messages.error(request, 'Solo puedes subir hasta 8 fotos de evidencia por entrega.')
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    for archivo in evidencias_subidas:
        if not getattr(archivo, 'content_type', '').startswith('image/'):
            messages.error(request, 'Todos los archivos de evidencia deben ser imagenes.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    with transaction.atomic():
        pedido = get_object_or_404(
            Pedido.objects.select_for_update().prefetch_related('detalles', 'evidencias'),
            pk=pedido_id,
        )

        evidencias_existentes = pedido.evidencias.count()
        if evidencias_existentes + len(evidencias_subidas) > 8:
            messages.error(request, 'Este pedido ya tiene evidencias. El maximo total permitido es 8 fotos.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        if pedido.estado != 'esperando entrega':
            messages.error(request, 'Solo se puede confirmar entrega en pedidos en estado esperando entrega.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        if not pedido.codigo_entrega or not pedido.codigo_expira_en:
            messages.error(request, 'Este pedido no tiene codigo de entrega activo.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        now = timezone.now()
        if now > pedido.codigo_expira_en:
            pedido.codigo_entrega = f'{secrets.randbelow(1000000):06d}'
            pedido.codigo_expira_en = now + timedelta(hours=2)
            pedido.fch_ult_act = now
            pedido.save(update_fields=['codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])
            messages.error(request, 'El codigo estaba vencido. Se genero uno nuevo con vigencia de 2 horas.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        if codigo_ingresado != pedido.codigo_entrega:
            messages.error(request, 'Codigo incorrecto. No se pudo confirmar la entrega.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        now = timezone.now()
        if evidencias_subidas:
            PedidoEvidencia.objects.bulk_create([
                PedidoEvidencia(
                    id_pedido_fk=pedido,
                    foto_evidencia=archivo,
                    fch_registro=now,
                )
                for archivo in evidencias_subidas
            ])

        detalles_entregables = list(
            DetallePedido.objects
            .select_related('id_prod_fk')
            .filter(id_pedido_fk=pedido)
            .exclude(estado_detalle__in=['no_disponible', 'rechazado', 'cancelado'])
        )
        detalles_consumo_ids = [
            d.id_det_pedido
            for d in detalles_entregables
            if d.id_prod_fk and d.id_prod_fk.tipo_bien == 'consumo'
        ]
        detalles_devolutivos_ids = [
            d.id_det_pedido
            for d in detalles_entregables
            if not d.id_prod_fk or d.id_prod_fk.tipo_bien != 'consumo'
        ]

        if detalles_devolutivos_ids:
            DetallePedido.objects.filter(id_det_pedido__in=detalles_devolutivos_ids).update(
                estado_detalle='entregado',
                fch_ult_act=now,
            )
        if detalles_consumo_ids:
            DetallePedido.objects.filter(id_det_pedido__in=detalles_consumo_ids).update(
                estado_detalle='devuelto',
                fch_ult_act=now,
            )

        pedido.estado = 'entregado' if detalles_devolutivos_ids else 'devuelto'
        pedido.codigo_entrega = None
        pedido.codigo_expira_en = None
        pedido.fch_ult_act = now
        pedido.save(update_fields=['estado', 'codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='pedido',
        entidad_id=pedido.id_pedido,
        descripcion=f'Pedido #{pedido.id_pedido} fue confirmado como entregado en almacén.',
    )
    messages.success(
        request,
        (
            f'Pedido #{pedido.id_pedido} marcado como entregado.'
            if pedido.estado == 'entregado'
            else f'Pedido #{pedido.id_pedido} entregado como consumo (sin devolución obligatoria).'
        ),
    )
    _crear_notificacion(
        usuario=pedido.id_usuario_fk,
        tipo='entregado',
        titulo='Pedido entregado',
        mensaje=(
            f'Tu pedido #{pedido.id_pedido} fue entregado correctamente. Recuerda devolver los materiales en la fecha acordada.'
            if pedido.estado == 'entregado'
            else f'Tu pedido #{pedido.id_pedido} fue entregado y corresponde a material de consumo, no requiere devolución.'
        ),
        pedido_id=pedido.id_pedido,
    )
    _notificar_staff(
        tipo='staff_pedido_entregado',
        titulo=f'Pedido #{pedido.id_pedido} entregado',
        mensaje=f'El pedido #{pedido.id_pedido} fue confirmado como entregado por {request.user.nombre or request.user.correo}.',
        pedido_id=pedido.id_pedido,
    )
    return redirect('pedido_detalle_panel', pedido_id=pedido_id)


@login_required
def pedido_marcar_devuelto(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('prestamos_panel')

    codigo_ingresado = (request.POST.get('codigo_devolucion') or '').strip()
    if not (len(codigo_ingresado) == 6 and codigo_ingresado.isdigit()):
        messages.error(request, 'Debes ingresar un código de devolución válido de 6 dígitos.')
        return redirect('prestamos_panel')

    with transaction.atomic():
        pedido = get_object_or_404(
            Pedido.objects.select_for_update().prefetch_related('detalles__id_prod_fk'),
            pk=pedido_id,
        )

        now = timezone.now()
        estado_canonico = _estado_pedido_canonico(pedido.estado)
        if estado_canonico and estado_canonico != pedido.estado:
            pedido.estado = estado_canonico
            pedido.fch_ult_act = now
            pedido.save(update_fields=['estado', 'fch_ult_act'])

        if pedido.estado == 'devuelto':
            if pedido.codigo_entrega or pedido.codigo_expira_en:
                pedido.codigo_entrega = None
                pedido.codigo_expira_en = None
                pedido.fch_ult_act = now
                pedido.save(update_fields=['codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])

            DetallePedido.objects.filter(
                id_pedido_fk=pedido,
                estado_detalle__in=['entregado', 'vencido', 'deuelto', 'debuelto'],
            ).update(
                estado_detalle='devuelto',
                fch_ult_act=now,
            )

            messages.info(request, f'El préstamo #{pedido.id_pedido} ya estaba marcado como devuelto.')
            return redirect('prestamos_panel')

        if pedido.estado not in ('entregado', 'vencido'):
            messages.error(request, 'Solo puedes marcar como devuelto un préstamo actualmente entregado.')
            return redirect('prestamos_panel')

        if not pedido.codigo_entrega or not pedido.codigo_expira_en:
            _renovar_codigo_devolucion(pedido, now)
            messages.error(request, 'No había un código activo. Se generó uno nuevo para el usuario.')
            return redirect('prestamos_panel')

        if now > pedido.codigo_expira_en:
            _renovar_codigo_devolucion(pedido, now)
            messages.error(request, 'El código de devolución venció. Pide al usuario el nuevo código dinámico.')
            return redirect('prestamos_panel')

        if codigo_ingresado != pedido.codigo_entrega:
            messages.error(request, 'Código de devolución incorrecto. Verifica la clave dinámica del usuario.')
            return redirect('prestamos_panel')

        detalles_entregados = list(
            DetallePedido.objects
            .select_for_update()
            .select_related('id_prod_fk')
            .filter(id_pedido_fk=pedido)
            .filter(estado_detalle='entregado')
        )

        restaurados = 0
        for detalle in detalles_entregados:
            if detalle.id_prod_fk and detalle.id_prod_fk.tipo_bien == 'consumo':
                continue
            _sumar_stock_disponibilidad(detalle, now)
            restaurados += 1

        pedido.estado = 'devuelto'
        pedido.codigo_entrega = None
        pedido.codigo_expira_en = None
        pedido.fch_ult_act = now
        pedido.save(update_fields=['estado', 'codigo_entrega', 'codigo_expira_en', 'fch_ult_act'])

        DetallePedido.objects.filter(id_pedido_fk=pedido, estado_detalle='entregado').update(
            estado_detalle='devuelto',
            fch_ult_act=now,
        )

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='prestamo',
        entidad_id=pedido.id_pedido,
        descripcion=f'Préstamo #{pedido.id_pedido} recibido en devolución y stock restaurado.',
    )
    if restaurados > 0:
        messages.success(request, f'Préstamo #{pedido.id_pedido} marcado como devuelto y el stock fue restaurado.')
    else:
        messages.success(request, f'Pedido #{pedido.id_pedido} cerrado. No hubo ítems devolutivos para restaurar stock.')
    return redirect('prestamos_panel')


@login_required
def pedido_rechazar(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    if request.method != 'POST':
        return redirect('pedido_detalle_panel', pedido_id=pedido_id)

    motivo_rechazo = (request.POST.get('motivo_rechazo') or '').strip()
    motivo_guardado = motivo_rechazo or 'El pedido fue rechazado por no disponibilidad.'

    with transaction.atomic():
        pedido = get_object_or_404(Pedido.objects.select_for_update(), pk=pedido_id)

        if pedido.estado != 'pendiente':
            messages.error(request, 'Solo se pueden rechazar pedidos en estado pendiente.')
            return redirect('pedido_detalle_panel', pedido_id=pedido_id)

        now = timezone.now()
        pedido.estado = 'rechazado'
        pedido.motivo_rechazo = motivo_guardado
        pedido.fch_ult_act = now
        pedido.save(update_fields=['estado', 'motivo_rechazo', 'fch_ult_act'])

        DetallePedido.objects.filter(id_pedido_fk=pedido).update(
            estado_detalle='rechazado',
            fch_ult_act=now,
        )

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='pedido',
        entidad_id=pedido.id_pedido,
        descripcion=f'Pedido #{pedido.id_pedido} fue rechazado por personal de almacén. Motivo: {motivo_guardado}',
    )
    messages.success(request, f'Pedido #{pedido_id} rechazado correctamente.')
    _crear_notificacion(
        usuario=pedido.id_usuario_fk,
        tipo='rechazado',
        titulo='Pedido rechazado',
        mensaje=(
            f'Tu pedido #{pedido.id_pedido} fue rechazado por el almacenista. '
            f'Motivo: {motivo_guardado}'
        ),
        pedido_id=pedido.id_pedido,
    )
    return redirect('pedido_detalle_panel', pedido_id=pedido_id)


@login_required
def auditorias_panel(request):
    if not _is_admin_or_almacenista(request):
        return redirect('dashboard')

    q = (request.GET.get('q') or '').strip()
    accion = (request.GET.get('accion') or '').strip().lower()
    entidad = (request.GET.get('entidad') or '').strip().lower()
    rol = (request.GET.get('rol') or '').strip().lower()

    logs = AuditoriaLog.objects.select_related('id_usuario_fk__id_rol_fk')

    if q:
        logs = logs.filter(
            models.Q(descripcion__icontains=q)
            | models.Q(entidad_id__icontains=q)
            | models.Q(id_usuario_fk__correo__icontains=q)
            | models.Q(id_usuario_fk__nombre__icontains=q)
            | models.Q(id_usuario_fk__apellido__icontains=q)
        )

    if accion:
        logs = logs.filter(accion=accion)
    if entidad:
        logs = logs.filter(entidad=entidad)
    if rol:
        logs = logs.filter(rol_usuario=rol)

    logs = list(logs.order_by('-fch_registro', '-id_log')[:300])
    resumen_accion = {
        'crear': sum(1 for item in logs if item.accion == 'crear'),
        'actualizar': sum(1 for item in logs if item.accion == 'actualizar'),
        'eliminar': sum(1 for item in logs if item.accion == 'eliminar'),
    }

    return render(request, 'inventario/auditorias/panel_auditorias.html', {
        'logs': logs,
        'q': q,
        'accion_activa': accion,
        'entidad_activa': entidad,
        'rol_activo': rol,
        'resumen_accion': resumen_accion,
    })

@login_required
def gestion_usuarios_panel(request):
    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede gestionar usuarios.')
        return redirect('dashboard')

    _expirar_solicitudes_validacion_manual()
    _reabrir_solicitudes_con_enlace_vencido()

    query = request.GET.get('q', '').strip()
    base_usuarios = Usuario.objects.all().select_related('id_rol_fk')
    usuarios = base_usuarios.order_by('nombre', 'apellido')
    if query:
        usuarios = usuarios.filter(
            models.Q(nombre__icontains=query) |
            models.Q(apellido__icontains=query) |
            models.Q(correo__icontains=query) |
            models.Q(cc__icontains=query)
        )
    roles = Rol.objects.all().order_by('nombre_rol')
    resumen = {
        'total_usuarios': base_usuarios.count(),
        'total_activos': base_usuarios.filter(is_active=True).count(),
        'total_admins': base_usuarios.filter(id_rol_fk__nombre_rol='admin').count(),
        'total_visibles': usuarios.count(),
    }
    return render(request, 'inventario/usuarios/panel_usuarios.html', {
        'usuarios': usuarios,
        'query': query,
        'roles': roles,
        'resumen': resumen,
    })


from django.views.decorators.http import require_POST
@login_required
@require_POST
def crear_usuario(request):
    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede crear usuarios.')
        return redirect('gestion_usuarios_panel')

    cc = request.POST.get('cc', '').strip()
    nombre = request.POST.get('nombre', '').strip()
    apellido = request.POST.get('apellido', '').strip()
    correo = request.POST.get('correo', '').strip()
    password = request.POST.get('password', '').strip()
    id_rol_fk = request.POST.get('id_rol_fk')
    if not (cc and nombre and apellido and correo and password and id_rol_fk):
        messages.error(request, 'Todos los campos son obligatorios.')
        return redirect('gestion_usuarios_panel')
    if Usuario.objects.filter(correo=correo).exists():
        messages.error(request, 'Ya existe un usuario con ese correo.')
        return redirect('gestion_usuarios_panel')
    if Usuario.objects.filter(cc=cc).exists():
        messages.error(request, 'Ya existe un usuario con esa cédula.')
        return redirect('gestion_usuarios_panel')
    try:
        rol = Rol.objects.get(pk=id_rol_fk)
    except Rol.DoesNotExist:
        messages.error(request, 'Rol inválido.')
        return redirect('gestion_usuarios_panel')
    usuario = Usuario(
        cc=cc,
        nombre=nombre,
        apellido=apellido,
        correo=correo,
        id_rol_fk=rol,
        is_active=True,
    )
    usuario.set_password(password)
    usuario.save()
    _registrar_auditoria(
        request,
        accion='crear',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se creó el usuario {usuario.correo} con rol {rol.nombre_rol}.',
    )
    return redirect('gestion_usuarios_panel')

@login_required
@require_POST
def editar_rol_usuario(request, usuario_id):
    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede editar roles.')
        return redirect('gestion_usuarios_panel')

    usuario = Usuario.objects.get(pk=usuario_id)
    nuevo_rol_id = request.POST.get('id_rol_fk')
    if not nuevo_rol_id:
        messages.error(request, 'Debes seleccionar un rol.')
        return redirect('gestion_usuarios_panel')
    try:
        nuevo_rol = Rol.objects.get(pk=nuevo_rol_id)
    except Rol.DoesNotExist:
        messages.error(request, 'Rol inválido.')
        return redirect('gestion_usuarios_panel')
    # No permitir cambiar el rol de admin
    if usuario.id_rol_fk and usuario.id_rol_fk.nombre_rol == 'admin':
        messages.error(request, 'No puedes editar el rol de un usuario admin.')
        return redirect('gestion_usuarios_panel')
    usuario.id_rol_fk = nuevo_rol
    usuario.save()
    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se actualizó el rol del usuario {usuario.correo} a {nuevo_rol.nombre_rol}.',
    )
    return redirect('gestion_usuarios_panel')


@login_required
@require_POST
def toggle_estado_usuario(request, usuario_id):
    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'No tienes permisos para cambiar el estado de usuarios.')
        return redirect('gestion_usuarios_panel')

    usuario = get_object_or_404(Usuario, pk=usuario_id)

    if usuario.id_usu == request.user.id_usu:
        messages.error(request, 'No puedes desactivar tu propia cuenta desde esta sesión.')
        return redirect('gestion_usuarios_panel')

    usuario.is_active = not usuario.is_active
    usuario.save(update_fields=['is_active'])

    accion = 'activado' if usuario.is_active else 'desactivado'
    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se dejó {accion} el acceso del usuario {usuario.correo}.',
    )
    return redirect('gestion_usuarios_panel')


@login_required
@require_POST
def eliminar_usuario(request, usuario_id):
    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'No tienes permisos para eliminar usuarios.')
        return redirect('gestion_usuarios_panel')

    usuario = get_object_or_404(Usuario, pk=usuario_id)

    if usuario.id_usu == request.user.id_usu:
        messages.error(request, 'No puedes eliminar tu propia cuenta desde esta sesión.')
        return redirect('gestion_usuarios_panel')

    nombre_completo = ' '.join(filter(None, [usuario.nombre, usuario.apellido])).strip() or 'Sin nombre'
    correo = usuario.correo
    entidad_id = usuario.id_usu

    try:
        usuario.delete()
    except Exception:
        messages.error(request, 'No se pudo eliminar el usuario. Verifica si tiene información relacionada.')
        return redirect('gestion_usuarios_panel')

    _registrar_auditoria(
        request,
        accion='eliminar',
        entidad='usuario',
        entidad_id=entidad_id,
        descripcion=f'Se eliminó el usuario {nombre_completo} ({correo}).',
    )
    messages.success(request, f'Usuario eliminado: {nombre_completo} ({correo}).')
    return redirect('gestion_usuarios_panel')


@login_required
@require_POST
def enviar_enlace_validacion_sena(request, usuario_id):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    _expirar_solicitudes_validacion_manual()
    _reabrir_solicitudes_con_enlace_vencido()

    def _redirect_admin_default():
        if next_url and next_url.startswith('/') and not next_url.startswith('//'):
            return redirect(next_url)
        return redirect('gestion_usuarios_panel')

    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede enviar enlaces de validación SENA.')
        return _redirect_admin_default()

    usuario = get_object_or_404(Usuario, pk=usuario_id)
    if usuario.verificacion_sena_estado == 'validado':
        messages.success(request, 'Ese usuario ya tiene la validación SENA aprobada.')
        return _redirect_admin_default()

    if usuario.verificacion_sena_estado == 'documento_cargado':
        messages.error(request, 'Este usuario ya cargó documento y está pendiente de aprobación/rechazo.')
        return _redirect_admin_default()

    estados_permitidos = {'solicitada', 'enlace_enviado', 'pendiente', 'rechazada'}
    if usuario.verificacion_sena_estado not in estados_permitidos:
        messages.error(request, 'Este usuario no está habilitado para envío manual de enlace en su estado actual.')
        return _redirect_admin_default()

    correo = getattr(usuario, 'correo', None)
    if not correo:
        messages.error(request, 'Este usuario no tiene correo registrado para enviar el enlace manual.')
        return _redirect_admin_default()

    # Invalida cualquier enlace previo sin usar y genera uno nuevo.
    token = VerificacionSenaToken.create_for_user(usuario)
    upload_url = request.build_absolute_uri(reverse('validacion_sena_carga_manual', args=[token.token]))
    try:
        from django.core.mail import EmailMultiAlternatives

        subject = 'Enlace de validación manual SENA'
        nombre_usuario = usuario.nombre or usuario.correo
        text_content = (
            f'Hola {nombre_usuario},\n\n'
            'El administrador aprobó tu solicitud de validación manual.\n'
            'Usa este enlace único para cargar la foto de tu carnet o un certificado vigente del SENA:\n'
            f'{upload_url}\n\n'
            'El enlace vencerá en 4 horas y solo podrá usarse una vez.'
        )
        html_content = f"""
<!DOCTYPE html>
<html lang=\"es\">
<head><meta charset=\"UTF-8\"><meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\"></head>
<body style=\"margin:0;padding:0;background:#f3f7f2;font-family:Arial,Helvetica,sans-serif;color:#1f2937;\">
  <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"padding:24px 12px;\">
    <tr><td align=\"center\">
      <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"max-width:640px;background:#ffffff;border-radius:22px;overflow:hidden;box-shadow:0 12px 30px rgba(11,71,55,.12);\">
        <tr><td style=\"background:linear-gradient(135deg,#0b4737,#39A900);padding:28px 32px;color:#fff;\">
          <p style=\"margin:0 0 8px;font-size:13px;letter-spacing:1.6px;font-weight:bold;text-transform:uppercase;opacity:.9;\">SENA · Inventario</p>
          <h1 style=\"margin:0;font-size:28px;line-height:1.15;\">Carga tu evidencia manual</h1>
        </td></tr>
        <tr><td style=\"padding:32px;\">
          <p style=\"margin:0 0 14px;font-size:16px;line-height:1.6;\">Hola <strong>{nombre_usuario}</strong>,</p>
          <p style=\"margin:0 0 18px;font-size:15px;line-height:1.7;color:#475569;\">Ya puedes cargar la foto de tu carnet SENA o un certificado que confirme que estudias en el SENA. Esta revisión será manual.</p>
          <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"margin:24px 0;\"><tr><td align=\"center\"><a href=\"{upload_url}\" style=\"display:inline-block;background:#39A900;color:#fff;text-decoration:none;font-weight:700;padding:14px 26px;border-radius:999px;font-size:15px;\">Cargar documento</a></td></tr></table>
          <p style=\"margin:0 0 10px;font-size:14px;line-height:1.7;color:#64748b;\">Si el botón no funciona, usa este enlace:</p>
          <p style=\"margin:0;font-size:13px;line-height:1.7;word-break:break-all;\"><a href=\"{upload_url}\" style=\"color:#0b4737;\">{upload_url}</a></p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
"""
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
            to=[correo],
        )
        email.attach_alternative(html_content, 'text/html')
        email.send(fail_silently=False)
    except Exception:
        if getattr(settings, 'IS_PYTHONANYWHERE', False):
            token.usado_en = timezone.now()
            token.save(update_fields=['usado_en'])
            messages.error(request, 'No se pudo enviar el correo. El usuario sigue en solicitud pendiente para reintentar envío.')
            return _redirect_admin_default()

        # En desarrollo local: dejar enlace utilizable aunque falle SMTP.
        usuario.verificacion_sena_estado = 'enlace_enviado'
        usuario.verificacion_sena_solicitada_en = timezone.now()
        usuario.verificacion_sena_observacion = 'Enlace manual generado en localhost; envío por correo no disponible.'
        usuario.save(update_fields=[
            'verificacion_sena_estado',
            'verificacion_sena_solicitada_en',
            'verificacion_sena_observacion',
        ])
        messages.warning(request, f'Correo no enviado en localhost. Usa este enlace manual para compartir con el usuario: {upload_url}')
        return _redirect_admin_default()

    ahora = timezone.now()
    usuario.verificacion_sena_estado = 'enlace_enviado'
    usuario.verificacion_sena_solicitada_en = ahora
    usuario.verificacion_sena_observacion = 'Administración envió enlace manual para cargar carnet o certificado SENA.'
    usuario.save(update_fields=[
        'verificacion_sena_estado',
        'verificacion_sena_solicitada_en',
        'verificacion_sena_observacion',
    ])

    _crear_notificacion(
        usuario=usuario,
        tipo='enlace_validacion_sena',
        titulo='Enlace de validación SENA enviado',
        mensaje='Revisa tu correo. Te enviamos un enlace único para cargar la foto del carnet o un certificado vigente del SENA.',
    )
    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se envió enlace manual de validación SENA al usuario {usuario.correo}.',
    )
    messages.success(request, f'Se envió el enlace manual de validación a {usuario.correo}.')
    return _redirect_admin_default()


@login_required
@require_POST
def aprobar_validacion_sena(request, usuario_id):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()

    def _redirect_admin_default():
        if next_url and next_url.startswith('/') and not next_url.startswith('//'):
            return redirect(next_url)
        return redirect('gestion_usuarios_panel')

    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede aprobar validaciones SENA.')
        return _redirect_admin_default()

    usuario = get_object_or_404(Usuario, pk=usuario_id)
    if not usuario.verificacion_sena_documento and not usuario.verificacion_sena_imagen:
        messages.error(request, 'Ese usuario todavía no ha cargado ninguna evidencia para revisar.')
        return _redirect_admin_default()

    usuario.verificacion_sena_estado = 'validado'
    usuario.verificacion_sena_validada_en = timezone.now()
    usuario.verificacion_sena_observacion = 'Validación manual aprobada por administración.'
    usuario.save(update_fields=[
        'verificacion_sena_estado',
        'verificacion_sena_validada_en',
        'verificacion_sena_observacion',
    ])

    _crear_notificacion(
        usuario=usuario,
        tipo='verificacion_sena_aprobada',
        titulo='Validación SENA aprobada',
        mensaje='El administrador aprobó tu verificación manual. Ya puedes realizar pedidos normalmente.',
    )

    correo = getattr(usuario, 'correo', None)
    if correo:
        try:
            from django.core.mail import EmailMultiAlternatives

            nombre_usuario = (f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() or usuario.correo)
            subject = 'Validación SENA aprobada'
            text_content = (
                f'Hola {nombre_usuario},\n\n'
                'Tu validación manual SENA fue aprobada por el administrador.\n'
                'Ya puedes realizar pedidos normalmente.\n'
            )
            html_content = f"""
<!DOCTYPE html>
<html lang=\"es\">
<head><meta charset=\"UTF-8\"><meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\"></head>
<body style=\"margin:0;padding:0;background:#f3f7f2;font-family:Arial,Helvetica,sans-serif;color:#1f2937;\">
    <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"padding:24px 12px;\">
        <tr><td align=\"center\">
            <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"max-width:640px;background:#ffffff;border-radius:22px;overflow:hidden;box-shadow:0 12px 30px rgba(11,71,55,.12);\">
                <tr><td style=\"background:linear-gradient(135deg,#0b4737,#39A900);padding:28px 32px;color:#fff;\">
                    <p style=\"margin:0 0 8px;font-size:13px;letter-spacing:1.6px;font-weight:bold;text-transform:uppercase;opacity:.9;\">SENA · Inventario</p>
                    <h1 style=\"margin:0;font-size:28px;line-height:1.15;\">Validación aprobada</h1>
                </td></tr>
                <tr><td style=\"padding:32px;\">
                    <p style=\"margin:0 0 14px;font-size:16px;line-height:1.6;\">Hola <strong>{nombre_usuario}</strong>,</p>
                    <p style=\"margin:0;font-size:15px;line-height:1.7;color:#475569;\">Tu validación manual SENA fue aprobada por el administrador. Ya puedes realizar pedidos normalmente.</p>
                </td></tr>
            </table>
        </td></tr>
    </table>
</body>
</html>
"""
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                to=[correo],
            )
            email.attach_alternative(html_content, 'text/html')
            email.send(fail_silently=False)
        except Exception:
            messages.warning(request, f'Se aprobó la validación, pero no se pudo enviar correo a {correo}.')

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se aprobó manualmente la validación SENA del usuario {usuario.correo}.',
    )
    messages.success(request, f'La validación SENA de {usuario.correo} fue aprobada.')
    return _redirect_admin_default()


@login_required
@require_POST
def rechazar_validacion_sena(request, usuario_id):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()

    def _redirect_admin_default():
        if next_url and next_url.startswith('/') and not next_url.startswith('//'):
            return redirect(next_url)
        return redirect('gestion_usuarios_panel')

    if not (request.user.id_rol_fk and request.user.id_rol_fk.nombre_rol == 'admin'):
        messages.error(request, 'Solo el administrador puede rechazar validaciones SENA.')
        return _redirect_admin_default()

    usuario = get_object_or_404(Usuario, pk=usuario_id)
    if not usuario.verificacion_sena_documento and not usuario.verificacion_sena_imagen:
        messages.error(request, 'Ese usuario todavía no ha cargado ninguna evidencia para revisar.')
        return _redirect_admin_default()

    motivo_rechazo = (request.POST.get('motivo_rechazo') or '').strip()
    observacion = 'Validación manual rechazada por administración.'
    if motivo_rechazo:
        observacion = f'{observacion} Motivo: {motivo_rechazo}'

    usuario.verificacion_sena_estado = 'rechazada'
    usuario.verificacion_sena_observacion = observacion
    usuario.save(update_fields=[
        'verificacion_sena_estado',
        'verificacion_sena_observacion',
    ])

    mensaje_rechazo = 'La revisión manual de tu validación SENA fue rechazada por el administrador.'
    if motivo_rechazo:
        mensaje_rechazo = f'{mensaje_rechazo} Motivo: {motivo_rechazo}'

    _crear_notificacion(
        usuario=usuario,
        tipo='verificacion_sena_rechazada',
        titulo='Validación SENA rechazada',
        mensaje=mensaje_rechazo,
    )

    correo = getattr(usuario, 'correo', None)
    if correo:
        try:
            from django.core.mail import EmailMultiAlternatives

            nombre_usuario = (f'{usuario.nombre or ""} {usuario.apellido or ""}'.strip() or usuario.correo)
            subject = 'Validación SENA rechazada'
            motivo_linea = f'Motivo: {motivo_rechazo}\n' if motivo_rechazo else ''
            text_content = (
                f'Hola {nombre_usuario},\n\n'
                'Tu validación manual SENA fue rechazada por el administrador.\n'
                f'{motivo_linea}'
                'Puedes volver a solicitar validación manual cuando tengas una nueva evidencia.\n'
            )
            motivo_html = f'<p style="margin:0 0 14px;font-size:15px;line-height:1.7;color:#b45309;"><strong>Motivo:</strong> {motivo_rechazo}</p>' if motivo_rechazo else ''
            html_content = f"""
<!DOCTYPE html>
<html lang=\"es\">
<head><meta charset=\"UTF-8\"><meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\"></head>
<body style=\"margin:0;padding:0;background:#f3f7f2;font-family:Arial,Helvetica,sans-serif;color:#1f2937;\">
    <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"padding:24px 12px;\">
        <tr><td align=\"center\">
            <table role=\"presentation\" width=\"100%\" cellspacing=\"0\" cellpadding=\"0\" style=\"max-width:640px;background:#ffffff;border-radius:22px;overflow:hidden;box-shadow:0 12px 30px rgba(11,71,55,.12);\">
                <tr><td style=\"background:linear-gradient(135deg,#7a1f1f,#c2410c);padding:28px 32px;color:#fff;\">
                    <p style=\"margin:0 0 8px;font-size:13px;letter-spacing:1.6px;font-weight:bold;text-transform:uppercase;opacity:.9;\">SENA · Inventario</p>
                    <h1 style=\"margin:0;font-size:28px;line-height:1.15;\">Validación rechazada</h1>
                </td></tr>
                <tr><td style=\"padding:32px;\">
                    <p style=\"margin:0 0 14px;font-size:16px;line-height:1.6;\">Hola <strong>{nombre_usuario}</strong>,</p>
                    <p style=\"margin:0 0 14px;font-size:15px;line-height:1.7;color:#475569;\">Tu validación manual SENA fue rechazada por el administrador.</p>
                    {motivo_html}
                    <p style=\"margin:0;font-size:15px;line-height:1.7;color:#475569;\">Puedes volver a solicitar validación manual cuando tengas una nueva evidencia.</p>
                </td></tr>
            </table>
        </td></tr>
    </table>
</body>
</html>
"""
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                to=[correo],
            )
            email.attach_alternative(html_content, 'text/html')
            email.send(fail_silently=False)
        except Exception:
            messages.warning(request, f'Se rechazó la validación, pero no se pudo enviar correo a {correo}.')

    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='usuario',
        entidad_id=usuario.id_usu,
        descripcion=f'Se rechazó manualmente la validación SENA del usuario {usuario.correo}. Motivo: {motivo_rechazo or "sin motivo"}.',
    )
    messages.success(request, f'La validación SENA de {usuario.correo} fue rechazada.')
    return _redirect_admin_default()


# ─────────────────────────────────────────────────────────────────────────────
# Panel de Notificaciones (usuario)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def notificaciones_panel(request):
    # Auto-marcar todas como leídas al entrar al panel
    Notificacion.objects.filter(id_usuario_fk=request.user, leida=False).update(leida=True)
    notificaciones = (
        Notificacion.objects
        .filter(id_usuario_fk=request.user)
        .order_by('-fch_registro')
    )
    return render(request, 'inventario/usuario/panel_notificaciones.html', {
        'notificaciones': notificaciones,
    })


@login_required
def live_sync_status(request):
    """Devuelve una firma de cambios para refresco ligero del frontend."""
    if request.method != 'GET':
        return JsonResponse({'ok': False, 'error': 'method_not_allowed'}, status=405)

    try:
        from django.db.models import Count

        # Throttle: evitar ejecutar mantenimientos pesados en cada poll del frontend.
        now = timezone.now()
        last_maintenance_at = cache.get('live_sync:last_maintenance_at')
        should_run_maintenance = (
            not last_maintenance_at
            or (now - last_maintenance_at).total_seconds() >= 60
        )
        if should_run_maintenance and cache.add('live_sync:maintenance_lock', 1, timeout=20):
            try:
                _auto_cancelar_pedidos_pendientes_vencidos()
                _auto_marcar_prestamos_vencidos()
                cache.set('live_sync:last_maintenance_at', now, timeout=120)
            finally:
                cache.delete('live_sync:maintenance_lock')

        usuario = request.user
        rol = getattr(getattr(usuario, 'id_rol_fk', None), 'nombre_rol', '') or ''

        noti_qs = Notificacion.objects.filter(id_usuario_fk=usuario)
        noti_ultima_id = noti_qs.order_by('-id_noti').values_list('id_noti', flat=True).first() or 0
        noti_no_leidas = noti_qs.filter(leida=False).count()

        firma_partes = [
            f'noti:{noti_ultima_id}:{noti_no_leidas}',
            (
                'sena:'
                f'{usuario.verificacion_sena_estado}:'
                f'{1 if bool(usuario.verificacion_sena_imagen) else 0}:'
                f'{1 if bool(usuario.verificacion_sena_documento) else 0}'
            ),
        ]

        if rol in ['admin', 'almacenista']:
            pedidos_staff = Pedido.objects.filter(estado__in=['pendiente', 'esperando entrega'])
            pedidos_staff_total = pedidos_staff.count()
            pedidos_staff_ultimo = pedidos_staff.order_by('-id_pedido').values_list('id_pedido', flat=True).first() or 0
            pedidos_vencidos_count = Pedido.objects.filter(estado='vencido').count()
            firma_partes.append(f'staff_pedidos:{pedidos_staff_total}:{pedidos_staff_ultimo}:{pedidos_vencidos_count}')
        else:
            pedidos_usuario = Pedido.objects.filter(id_usuario_fk=usuario)
            pedidos_usuario_ultimo = pedidos_usuario.order_by('-id_pedido').values_list('id_pedido', flat=True).first() or 0
            conteo_estados = {
                row['estado']: row['total']
                for row in pedidos_usuario.values('estado').annotate(total=Count('id_pedido')).order_by('estado')
            }
            estados_firma = ';'.join(f'{estado}:{total}' for estado, total in conteo_estados.items())
            firma_partes.append(f'usuario_pedidos:{pedidos_usuario_ultimo}:{estados_firma}')

        firma = '|'.join(firma_partes)
        return JsonResponse({
            'ok': True,
            'signature': firma,
            'unread_notifications': noti_no_leidas,
        })
    except Exception:
        # Evita romper la navegación del usuario si falla el endpoint.
        return JsonResponse({'ok': False, 'error': 'sync_unavailable'})


@login_required
def staff_alerts_api(request):
    """Devuelve alertas de acciones pendientes para admin y almacenista."""
    if request.method != 'GET':
        return JsonResponse({'ok': False}, status=405)

    _expirar_solicitudes_validacion_manual()
    _reabrir_solicitudes_con_enlace_vencido()

    rol = _user_role(request)
    if rol not in ('admin', 'almacenista'):
        return JsonResponse({'ok': False, 'alerts': []}, status=403)

    alerts = []

    # Pedidos pendientes de aprobación
    n = Pedido.objects.filter(estado='pendiente').count()
    if n:
        alerts.append({
            'mensaje': f'{n} pedido{"s" if n > 1 else ""} pendiente{"s" if n > 1 else ""} de aprobación',
            'url': reverse('pedidos_panel'),
            'icono': 'cube-outline',
            'tipo': 'warning',
        })

    # Pedidos esperando entrega
    n = Pedido.objects.filter(estado='esperando entrega').count()
    if n:
        alerts.append({
            'mensaje': f'{n} pedido{"s" if n > 1 else ""} esperando entrega',
            'url': reverse('pedidos_panel'),
            'icono': 'bicycle-outline',
            'tipo': 'info',
        })

    # Préstamos vencidos
    n = Pedido.objects.filter(estado='vencido').count()
    if n:
        alerts.append({
            'mensaje': f'{n} préstamo{"s" if n > 1 else ""} vencido{"s" if n > 1 else ""}',
            'url': reverse('prestamos_panel'),
            'icono': 'alert-circle-outline',
            'tipo': 'danger',
        })

    # Validaciones SENA pendientes — solo admin (almacenista no tiene acceso a /usuarios/)
    if rol == 'admin':
        n = Usuario.objects.filter(
            verificacion_sena_estado__in=['solicitada', 'documento_cargado']
        ).count()
        if n:
            alerts.append({
                'mensaje': f'{n} validación{"es" if n > 1 else ""} SENA pendiente{"s" if n > 1 else ""} de revisión',
                'url': '#',  # No redirige, solo abre modal si el frontend lo decide
                'icono': 'person-add-outline',
                'tipo': 'sena_validacion',
            })

    return JsonResponse({'ok': True, 'alerts': alerts})


@login_required
@require_POST
def notificacion_marcar_leida(request, noti_id):
    noti = get_object_or_404(Notificacion, pk=noti_id, id_usuario_fk=request.user)
    noti.leida = True
    noti.save(update_fields=['leida'])
    return redirect('notificaciones_panel')


@login_required
@require_POST
def notificaciones_marcar_todas_leidas(request):
    Notificacion.objects.filter(id_usuario_fk=request.user, leida=False).update(leida=True)
    return redirect('notificaciones_panel')


# ─────────────────────────────────────────────────────────────────────────────
# Aviso de devolución (almacenista → usuario)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def pedido_aviso_devolucion(request, pedido_id):
    if not request.user.id_rol_fk or request.user.id_rol_fk.nombre_rol not in ['admin', 'almacenista']:
        return redirect('dashboard')

    pedido = get_object_or_404(Pedido, pk=pedido_id, estado='entregado')
    usuario = pedido.id_usuario_fk
    _registrar_auditoria(
        request,
        accion='actualizar',
        entidad='prestamo',
        entidad_id=pedido.id_pedido,
        descripcion=f'Se envió aviso de devolución para el préstamo #{pedido.id_pedido}.',
    )
    _crear_notificacion(
        usuario=usuario,
        tipo='aviso_devolucion',
        titulo='Aviso de devolución pendiente',
        mensaje=f'El almacenista solicita que devuelvas los materiales del pedido #{pedido.id_pedido}. '
                f'Por favor, acércate al almacén a la brevedad posible.',
        pedido_id=pedido.id_pedido,
    )

    # ── Correo al usuario ─────────────────────────────────────────────────
    correo = getattr(usuario, 'correo', None) or getattr(usuario, 'email', None)
    if correo:
        try:
            from django.core.mail import EmailMultiAlternatives
            nombre = getattr(usuario, 'nombre', '') or str(usuario)
            fecha_str = pedido.fecha_devolucion.strftime('%d/%m/%Y') if pedido.fecha_devolucion else '—'
            remitente = settings.DEFAULT_FROM_EMAIL

            # Productos a devolver (solo los activos, no rechazados/cancelados)
            detalles = list(pedido.detalles.exclude(
                estado_detalle__in=['no_disponible', 'rechazado', 'cancelado']
            ).select_related('id_prod_fk'))

            # URL base para imágenes (usar dominio absoluto para que funcione en correos)
            base_url = 'https://almacensedelacolonia.pythonanywhere.com'

            # Construir filas de productos para el correo
            filas_html = ''
            lista_texto = ''
            for d in detalles:
                prod = d.id_prod_fk
                img_url = (
                    f'{base_url}{settings.MEDIA_URL}{prod.fot_prod}'
                    if prod and prod.fot_prod else ''
                )
                img_tag = (
                    f'<img src="{img_url}" alt="{d.nombre_producto}" '
                    f'width="48" height="48" '
                    f'style="border-radius:6px;object-fit:cover;display:block;">'
                    if img_url else
                    '<div style="width:48px;height:48px;background:#e8f5e9;'
                    'border-radius:6px;display:flex;align-items:center;'
                    'justify-content:center;font-size:20px;">📦</div>'
                )
                filas_html += f"""
                <tr>
                  <td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;width:68px;">
                    {img_tag}
                  </td>
                  <td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;
                              font-size:14px;color:#333;">{d.nombre_producto}</td>
                  <td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;
                              font-size:14px;color:#555;text-align:center;
                              white-space:nowrap;">x{d.cantidad_solicitada}</td>
                </tr>"""
                lista_texto += f'  - {d.nombre_producto} x{d.cantidad_solicitada}\n'

            tabla_productos = f"""
            <p style="font-size:15px;font-weight:700;color:#1a2e1a;margin:24px 0 8px;">
              📋 Productos a devolver:
            </p>
            <table width="100%" cellpadding="0" cellspacing="0"
                   style="border:1px solid #e0e0e0;border-radius:8px;overflow:hidden;">
              <thead>
                <tr style="background:#f5f5f5;">
                  <th style="padding:10px 12px;text-align:left;font-size:13px;
                              color:#666;font-weight:600;width:68px;">Foto</th>
                  <th style="padding:10px 12px;text-align:left;font-size:13px;
                              color:#666;font-weight:600;">Producto</th>
                  <th style="padding:10px 12px;text-align:center;font-size:13px;
                              color:#666;font-weight:600;">Cant.</th>
                </tr>
              </thead>
              <tbody>{filas_html}</tbody>
            </table>""" if detalles else ''

            asunto = f'📦 Recordatorio de devolución – Pedido #{pedido.id_pedido} | Almacén SENA Sibaté'
            texto_plano = (
                f'Hola {nombre},\n\n'
                f'El almacenista te recuerda que debes devolver los materiales del '
                f'pedido #{pedido.id_pedido} (fecha límite: {fecha_str}).\n\n'
                + (f'Productos a devolver:\n{lista_texto}\n' if lista_texto else '')
                + 'Por favor, acércate al almacén a la brevedad posible.\n\n'
                '— Almacén SENA Sibaté'
            )
            html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f4f4;padding:32px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0"
             style="background:#fff;border-radius:12px;overflow:hidden;
                    box-shadow:0 2px 12px rgba(0,0,0,0.08);max-width:600px;width:100%;">
        <tr>
          <td style="background:#39A900;padding:28px 32px;text-align:center;">
            <p style="margin:0;color:#fff;font-size:13px;opacity:0.85;">SENA — Almacén Sibaté</p>
            <h1 style="margin:8px 0 0;color:#fff;font-size:24px;">📦 Recordatorio de devolución</h1>
          </td>
        </tr>
        <tr>
          <td style="padding:32px;">
            <p style="font-size:16px;color:#333;">Hola <strong>{nombre}</strong>,</p>
            <p style="font-size:15px;color:#444;line-height:1.6;">
              El almacenista te recuerda que tienes pendiente la devolución de los materiales
              del préstamo <strong>#{pedido.id_pedido}</strong>
              {"(fecha límite: <strong>" + fecha_str + "</strong>)" if pedido.fecha_devolucion else ""}.
            </p>
            {tabla_productos}
            <table width="100%" cellpadding="0" cellspacing="0" style="margin:24px 0;">
              <tr>
                <td style="background:#e8f5e9;border-left:4px solid #39A900;
                            border-radius:6px;padding:16px 20px;">
                  <p style="margin:0;font-size:15px;color:#333;">
                    Por favor <strong>acércate al almacén</strong> a la brevedad posible
                    para hacer la devolución.
                  </p>
                </td>
              </tr>
            </table>
            <p style="font-size:13px;color:#888;margin-top:32px;">
              Si ya devolviste los materiales, ignora este mensaje.<br>
              — Almacén SENA Sibaté
            </p>
          </td>
        </tr>
        <tr>
          <td style="background:#f9f9f9;padding:16px 32px;text-align:center;
                      border-top:1px solid #eee;">
            <p style="margin:0;font-size:12px;color:#aaa;">
              Centro Industrial y de Desarrollo Empresarial – Sibaté, Cundinamarca
            </p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""
            msg = EmailMultiAlternatives(asunto, texto_plano, remitente, [correo])
            msg.attach_alternative(html, 'text/html')
            msg.send()
            messages.success(request, f'Aviso enviado al usuario y correo enviado a {correo}.')
        except Exception as e:
            messages.warning(request, f'Aviso interno enviado, pero el correo falló: {e}')
    else:
        messages.success(request, f'Aviso de devolución enviado al usuario del pedido #{pedido_id}.')

    return redirect('prestamos_panel')


